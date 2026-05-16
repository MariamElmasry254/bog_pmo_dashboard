"""
BOG Digital Transformation - PMO Dashboard v3
Modular templates + service mapping + filters
"""
from flask import Flask, render_template, jsonify, request, Response, session, redirect, url_for
import pandas as pd
import os
import sys
import xmlrpc.client
import traceback
import csv
import io
from datetime import datetime, timedelta, date
import logging
from roadmap_data import ROADMAP, MILESTONES, PROJECT_INFO
from db import DB, migrate_from_json
from positions_catalog import (
    seed_positions_if_empty,
    get_all_positions, get_position_by_name,
    get_all_tunis_rates, get_tunis_rate_by_name,
    upsert_position, upsert_tunis_rate,
    delete_position, delete_tunis_rate,
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s', stream=sys.stdout)
logger = logging.getLogger(__name__)

# Configuration
ODOO_URL = os.environ.get('ODOO_URL', 'https://erp.envnt.co')
ODOO_DB = os.environ.get('ODOO_DB', 'envnt')
ODOO_USERNAME = os.environ.get('ODOO_USERNAME', '')
ODOO_PASSWORD = os.environ.get('ODOO_PASSWORD', '')
PROJECT_NAME = os.environ.get('PROJECT_NAME', 'BOG Digital Transformation')

WORK_HOURS_PER_DAY = 8
WEEKEND_DAYS = [4, 5]  # Friday, Saturday

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Persistent storage path. On Railway, set DATA_PATH env var to mounted volume.
# Falls back to BASE_DIR/data for local dev.
PERSIST_DIR = os.environ.get('DATA_PATH', os.path.join(BASE_DIR, 'data'))
os.makedirs(PERSIST_DIR, exist_ok=True)
logger.info(f"Persistent storage at: {PERSIST_DIR}")
DATA_FILE = os.path.join(BASE_DIR, 'data', 'services.xlsx')

# =================================================================
# DATABASE - PostgreSQL (Railway plugin) or SQLite (fallback)
# Set DATABASE_URL env var in Railway to use PostgreSQL.
# Falls back to SQLite at PERSIST_DIR/pmo.db otherwise.
# =================================================================
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    db = DB(DATABASE_URL)
else:
    DB_PATH = os.path.join(PERSIST_DIR, 'pmo.db')
    db = DB(DB_PATH)

# One-time migration from JSON files (safe to run on every boot - it's idempotent)
try:
    migrated = migrate_from_json(db, PERSIST_DIR)
    if migrated:
        logger.info(f"Migrated from JSON: {migrated}")
except Exception as _e:
    logger.warning(f"Migration step had issues (continuing): {_e}")

# Seed positions catalog on first boot (idempotent - only inserts if empty)
try:
    seeded = seed_positions_if_empty(db)
    if seeded > 0:
        logger.info(f"Seeded positions catalog: {seeded} rows (positions + Tunis rates)")
except Exception as _e:
    logger.warning(f"Position seeding had issues (continuing): {_e}")

# Always update positions with latest rates from catalog (handles rate changes)
try:
    from positions_catalog import POSITIONS_SEED, TUNIS_RATES_SEED
    for _p in POSITIONS_SEED:
        upsert_position(db, _p['position'], _p['hour_rate'], _p.get('md_rate'),
                        _p.get('country'), _p.get('is_onsite', False))
    for _t in TUNIS_RATES_SEED:
        upsert_tunis_rate(db, _t['name'], _t['hour_rate'])
    logger.info(f"Refreshed {len(POSITIONS_SEED)} positions + {len(TUNIS_RATES_SEED)} Tunis rates from catalog")
except Exception as _e:
    logger.warning(f"Position refresh had issues (continuing): {_e}")

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'templates'),
    static_folder=os.path.join(BASE_DIR, 'static')
)
app.secret_key = os.environ.get('SECRET_KEY', 'codelab-pmo-2026-secret')

def get_active_project():
    """Returns (project_id, project_name) from session, falling back to env default."""
    pid  = session.get('project_id')
    name = session.get('project_name', PROJECT_NAME)
    return pid, name

def active_project_name():
    """Shortcut: returns just the project name for Odoo queries."""
    return session.get('project_name', PROJECT_NAME)

def active_db_prefix():
    """Returns DB namespace prefix for this project.
    BOG Digital Transformation (id=228) uses empty prefix to preserve legacy data.
    All other projects use 'proj_{id}_' prefix for isolation.
    """
    pid = session.get('project_id')
    if not pid or str(pid) == '228':
        return ''  # BOG legacy — no prefix
    return f'proj_{pid}'

def proj_get_override(namespace, subkey, key):
    """Get DB override with project prefix."""
    pfx = active_db_prefix()
    full_ns = f'{pfx}_{namespace}' if pfx else namespace
    return db.get_override(full_ns, subkey, key)

def proj_set_override(namespace, subkey, key, value):
    """Set DB override with project prefix."""
    pfx = active_db_prefix()
    full_ns = f'{pfx}_{namespace}' if pfx else namespace
    db.set_override(full_ns, subkey, key, value)

def proj_get_namespace(namespace, subkey):
    """Get all overrides in namespace+subkey with project prefix."""
    pfx = active_db_prefix()
    full_ns = f'{pfx}_{namespace}' if pfx else namespace
    return db.get_namespace_overrides(full_ns, subkey)


# ============================================================
# ODOO CLIENT
# ============================================================
class OdooClient:
    def __init__(self):
        self.uid = None
        self.models = None
        self.last_error = None

    def connect(self):
        if not ODOO_USERNAME or not ODOO_PASSWORD:
            self.last_error = "Credentials not set"
            return False
        try:
            common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
            self.uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})
            if self.uid:
                self.models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
                logger.info(f"Odoo connected: uid={self.uid}")
                self.last_error = None
                return True
            self.last_error = "Authentication failed"
            return False
        except Exception as e:
            self.last_error = f"{type(e).__name__}: {str(e)}"
            logger.error(f"Odoo connect: {e}")
            return False

    def get_phases(self, project_name=PROJECT_NAME):
        """Get list of phases for the project"""
        if not self.uid and not self.connect():
            return None
        try:
            # First find project ID
            projects = self.models.execute_kw(
                ODOO_DB, self.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[('name', 'ilike', project_name)]],
                {'fields': ['id', 'name'], 'limit': 5}
            )
            if not projects:
                return []
            project_ids = [p['id'] for p in projects]

            # Get phases linked to this project
            phases = self.models.execute_kw(
                ODOO_DB, self.uid, ODOO_PASSWORD,
                'project.phase', 'search_read',
                [[('project_id', 'in', project_ids)]],
                {'fields': ['id', 'name', 'project_id']}
            )
            return phases
        except Exception as e:
            logger.error(f"Odoo phases: {e}")
            return None

    def get_timesheets(self, project_name=PROJECT_NAME, date_from=None, date_to=None,
                       phase_filter=None, all_projects=False):
        """Get timesheets. phase_filter can be a single name or list of names."""
        if not self.uid:
            if not self.connect():
                return None
        try:
            domain = []
            if not all_projects and project_name:
                domain.append(('project_id.name', 'ilike', project_name))
            if date_from:
                domain.append(('date', '>=', date_from))
            if date_to:
                domain.append(('date', '<=', date_to))

            # Phase filter — can be string or list
            if phase_filter:
                phase_names = phase_filter if isinstance(phase_filter, list) else [phase_filter]
                phase_names = [p for p in phase_names if p]  # remove empties
                if phase_names:
                    try:
                        # Try project.phase first (BOG), then project.task.type (non-BOG)
                        phases = self.models.execute_kw(
                            ODOO_DB, self.uid, ODOO_PASSWORD,
                            'project.phase', 'search_read',
                            [[('name', 'in', phase_names)]],
                            {'fields': ['id'], 'limit': 50}
                        )
                        # If no project.phase found, try task.type and filter by stage_id
                        if not phases:
                            task_types = self.models.execute_kw(
                                ODOO_DB, self.uid, ODOO_PASSWORD,
                                'project.task.type', 'search_read',
                                [[('name', 'in', phase_names)]],
                                {'fields': ['id'], 'limit': 50}
                            )
                            if task_types:
                                type_ids = [t['id'] for t in task_types]
                                type_tasks = self.models.execute_kw(
                                    ODOO_DB, self.uid, ODOO_PASSWORD,
                                    'project.task', 'search_read',
                                    [[('stage_id', 'in', type_ids),
                                      ('project_id.name', 'ilike', project_name)]],
                                    {'fields': ['id'], 'limit': 5000}
                                )
                                if type_tasks:
                                    type_task_ids = [t['id'] for t in type_tasks]
                                    domain.append(('task_id', 'in', type_task_ids))
                                phases = []  # skip the phase_ids block below
                        phase_ids = [p['id'] for p in phases]
                        if phase_ids:
                            # Get ALL tasks with phase_id in our phases (not just parent tasks)
                            direct_tasks = self.models.execute_kw(
                                ODOO_DB, self.uid, ODOO_PASSWORD,
                                'project.task', 'search_read',
                                [[('phase_id', 'in', phase_ids)]],
                                {'fields': ['id'], 'limit': 5000}
                            )
                            direct_task_ids = {t['id'] for t in direct_tasks}

                            # Also get ALL sub-tasks of those tasks (walk down)
                            if direct_task_ids:
                                all_proj_tasks = self.models.execute_kw(
                                    ODOO_DB, self.uid, ODOO_PASSWORD,
                                    'project.task', 'search_read',
                                    [[('project_id.name', 'ilike', project_name)]],
                                    {'fields': ['id', 'parent_id', 'phase_id'], 'limit': 10000}
                                )
                                task_map = {t['id']: t for t in all_proj_tasks}
                                relevant_ids = set(direct_task_ids)
                                # Add tasks whose phase_id is in our phases directly
                                for t in all_proj_tasks:
                                    ph = t.get('phase_id')
                                    if ph and isinstance(ph, list) and ph[0] in phase_ids:
                                        relevant_ids.add(t['id'])
                                # Walk up parent chains to include sub-tasks
                                for t in all_proj_tasks:
                                    if t['id'] in relevant_ids:
                                        continue
                                    cur = t
                                    visited = set()
                                    while cur and cur['id'] not in visited:
                                        visited.add(cur['id'])
                                        if cur['id'] in direct_task_ids:
                                            relevant_ids.add(t['id'])
                                            break
                                        if not cur.get('parent_id'):
                                            break
                                        pid = cur['parent_id'][0] if isinstance(cur['parent_id'], list) else cur['parent_id']
                                        cur = task_map.get(pid)

                                domain.append(('task_id', 'in', list(relevant_ids)))
                            else:
                                return []
                        else:
                            return []
                    except Exception as e:
                        logger.warning(f"Phase filter failed: {e}")

            result = self.models.execute_kw(
                ODOO_DB, self.uid, ODOO_PASSWORD,
                'account.analytic.line', 'search_read', [domain],
                {'fields': ['date', 'employee_id', 'project_id', 'task_id',
                            'name', 'unit_amount'], 'limit': 5000}
            )
            logger.info(f"Odoo timesheets fetched: {len(result)} entries (filter: {domain})")

            # Enrich with parent task info (Odoo v14 doesn't have parent_task_id on timesheet)
            # We need to fetch task records to get their parent_id and phase
            task_ids = list(set(e['task_id'][0] for e in result if e.get('task_id')))
            parent_map = {}  # task_id -> parent_task_name
            phase_map = {}   # task_id -> phase_name
            if task_ids:
                try:
                    tasks = self.models.execute_kw(
                        ODOO_DB, self.uid, ODOO_PASSWORD,
                        'project.task', 'read',
                        [task_ids],
                        {'fields': ['id', 'name', 'parent_id', 'phase_id']}
                    )
                    for t in tasks:
                        parent = t.get('parent_id')
                        if parent and isinstance(parent, list) and len(parent) > 1:
                            parent_map[t['id']] = parent[1]
                        else:
                            parent_map[t['id']] = t.get('name')
                        ph = t.get('phase_id')
                        if ph and isinstance(ph, list) and len(ph) > 1:
                            phase_map[t['id']] = ph[1]
                    logger.info(f"Loaded mapping for {len(tasks)} tasks")
                except Exception as e:
                    logger.warning(f"Could not load tasks: {e}")

            # Attach parent task name + phase to each entry
            for entry in result:
                task = entry.get('task_id')
                if task and task[0] in parent_map:
                    entry['_parent_name'] = parent_map[task[0]]
                else:
                    entry['_parent_name'] = task[1] if task else ''
                entry['_phase_name'] = phase_map.get(task[0] if task else None, '')

            return result
        except Exception as e:
            self.last_error = f"{type(e).__name__}: {str(e)}"
            logger.error(f"Odoo timesheets: {e}")
            self.uid = None
            self.models = None
            return None

odoo = OdooClient()

# ============================================================
# SERVICE MAPPING
# ============================================================
DEPT_COLUMNS = ['Dev MDs', 'Analysis MD', 'UI/UX', 'QC', 'UAT', 'PM']
DEPT_LABELS = {'Dev MDs': 'Dev', 'Analysis MD': 'Analysis', 'UI/UX': 'UI/UX',
               'QC': 'QC', 'UAT': 'UAT', 'PM': 'PM'}

_services_cache = None

def load_services():
    """Load services from Excel and merge with roadmap data"""
    global _services_cache
    if _services_cache is not None:
        return _services_cache
    try:
        if not os.path.exists(DATA_FILE):
            _services_cache = []
            return _services_cache
        df = pd.read_excel(DATA_FILE)
        services = []

        # Map roadmap by name (fuzzy)
        roadmap_by_name = {item['name']: item for item in ROADMAP}

        for record in df.to_dict(orient='records'):
            clean = {}
            for k, v in record.items():
                if pd.isna(v):
                    clean[k] = None
                elif isinstance(v, (pd.Timestamp, datetime)):
                    clean[k] = v.isoformat()
                else:
                    clean[k] = v

            # Map roadmap
            name = clean.get('اسم الخدمة المستقبلي', '') or ''
            rm = roadmap_by_name.get(name)
            if not rm and name:
                for rm_name, rm_data in roadmap_by_name.items():
                    if name in rm_name or (rm_name in name and len(rm_name) > 4):
                        rm = rm_data
                        break

            clean['planned_team'] = rm['team'] if rm else None
            clean['planned_start'] = rm['start'] if rm else None
            clean['planned_end'] = rm['end'] if rm else None
            clean['planned_wd_roadmap'] = rm['wd'] if rm else None

            # Department-wise baseline (from presales)
            baseline_by_dept = {}
            for col in DEPT_COLUMNS:
                val = clean.get(col)
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    try:
                        baseline_by_dept[DEPT_LABELS[col]] = float(val)
                    except (ValueError, TypeError):
                        pass  # skip non-numeric values like "Total"
            clean['baseline_by_dept'] = baseline_by_dept

            # Placeholder for planned/actual/assignation (to be filled from Odoo or WBD)
            clean['planned_by_dept'] = {k: None for k in DEPT_LABELS.values()}
            clean['actuals_hours'] = 0  # will be computed dynamically
            clean['actuals_days'] = 0
            clean['remaining_baseline'] = clean.get('ALL') or 0
            clean['status'] = 'Not Started'  # auto, but overridable
            clean['status_manual'] = None  # if PM overrides
            clean['expected_end'] = None  # computed: today + remaining/8

            services.append(clean)

        _services_cache = services
        return services
    except Exception as e:
        logger.error(f"load_services: {e}\n{traceback.format_exc()}")
        return []

def normalize_timesheet(entry):
    """Convert Odoo timesheet entry to flat dict.
    'service' = parent task name (the umbrella service the task belongs to).
    'phase' = phase name (project.phase).
    For Odoo v14, both are computed via task lookup.
    """
    project = entry.get('project_id', [None, ''])
    task = entry.get('task_id', [None, ''])
    task_name = task[1] if task and task[1] else (entry.get('name') or '')

    # SERVICE = parent task name (computed in get_timesheets), fallback to task name itself
    service_name = entry.get('_parent_name') or task_name
    phase_name = entry.get('_phase_name') or ''

    return {
        'date': entry.get('date'),
        'employee':         entry.get('employee_id', [None, 'Unknown'])[1] if entry.get('employee_id') else 'Unknown',
        'odoo_employee_id': entry.get('employee_id', [None])[0] if entry.get('employee_id') else None,
        'project': project[1] if project else '',
        'task': task_name,
        'service': service_name,  # umbrella/parent — used to group actuals per service
        'phase': phase_name,
        'description': entry.get('name', ''),
        'hours': float(entry.get('unit_amount', 0)),
    }

def get_demo_timesheets():
    """Demo data with phase + parent service grouping"""
    return [
        # Service: إدارة الصلاحيات (Development Phase)
        {'date': '2026-04-26', 'employee': 'Ahmed Hassan', 'project': PROJECT_NAME, 'task': 'JWT Setup', 'service': 'إدارة الصلاحيات', 'phase': 'Development Phase', 'description': 'Auth flow', 'hours': 7.5},
        {'date': '2026-04-27', 'employee': 'Ahmed Hassan', 'project': PROJECT_NAME, 'task': 'Token refresh', 'service': 'إدارة الصلاحيات', 'phase': 'Development Phase', 'description': '', 'hours': 8.0},
        {'date': '2026-05-03', 'employee': 'Ahmed Hassan', 'project': PROJECT_NAME, 'task': 'Bug fix', 'service': 'إدارة الصلاحيات', 'phase': 'Development Phase', 'description': '', 'hours': 6.5},
        {'date': '2026-05-02', 'employee': 'Omar Khaled', 'project': PROJECT_NAME, 'task': 'Frontend auth', 'service': 'إدارة الصلاحيات', 'phase': 'Development Phase', 'description': '', 'hours': 7.0},
        # Service: قيد دعوى إدارية (Development Phase)
        {'date': '2026-04-28', 'employee': 'Ahmed Hassan', 'project': PROJECT_NAME, 'task': 'Cases schema', 'service': 'قيد دعوى إدارية', 'phase': 'Development Phase', 'description': '', 'hours': 6.0},
        {'date': '2026-04-29', 'employee': 'Ahmed Hassan', 'project': PROJECT_NAME, 'task': 'Cases schema', 'service': 'قيد دعوى إدارية', 'phase': 'Development Phase', 'description': '', 'hours': 8.0},
        {'date': '2026-04-28', 'employee': 'Sara Ali', 'project': PROJECT_NAME, 'task': 'Wireframes', 'service': 'قيد دعوى إدارية', 'phase': 'Development Phase', 'description': '', 'hours': 7.5},
        {'date': '2026-04-29', 'employee': 'Sara Ali', 'project': PROJECT_NAME, 'task': 'Wireframes', 'service': 'قيد دعوى إدارية', 'phase': 'Development Phase', 'description': '', 'hours': 8.0},
        # Consultation phases
        {'date': '2026-04-27', 'employee': 'Sara Ali', 'project': PROJECT_NAME, 'task': 'Stakeholder Interviews', 'service': 'الفهارس العامة', 'phase': 'Consultation phase - Analysis', 'description': '', 'hours': 7.0},
        {'date': '2026-04-29', 'employee': 'Omar Khaled', 'project': PROJECT_NAME, 'task': 'UX Research', 'service': 'الفهارس العامة', 'phase': 'Consultation phase - UX', 'description': '', 'hours': 7.5},
        {'date': '2026-04-28', 'employee': 'Omar Khaled', 'project': PROJECT_NAME, 'task': 'Initiation Meeting', 'service': 'تسجيل الدخول', 'phase': 'Consultation phase - Initiation', 'description': '', 'hours': 8.0},
        # PM activities
        {'date': '2026-04-26', 'employee': 'Mariam Elmasry', 'project': PROJECT_NAME, 'task': 'PM Review', 'service': '', 'phase': 'Development Phase', 'description': '', 'hours': 6.0},
        {'date': '2026-04-27', 'employee': 'Mariam Elmasry', 'project': PROJECT_NAME, 'task': 'Stakeholder Meeting', 'service': '', 'phase': 'Development Phase', 'description': '', 'hours': 4.5},
        {'date': '2026-04-28', 'employee': 'Mariam Elmasry', 'project': PROJECT_NAME, 'task': 'Sprint Planning', 'service': '', 'phase': 'Development Phase', 'description': '', 'hours': 5.5},
        {'date': '2026-04-29', 'employee': 'Mariam Elmasry', 'project': PROJECT_NAME, 'task': 'Documentation', 'service': '', 'phase': 'Development Phase', 'description': '', 'hours': 7.0},
        {'date': '2026-04-30', 'employee': 'Mariam Elmasry', 'project': PROJECT_NAME, 'task': 'Sprint Review', 'service': '', 'phase': 'Development Phase', 'description': '', 'hours': 4.0},
        # Cross-project (no phase)
        {'date': '2026-05-03', 'employee': 'Mariam Elmasry', 'project': 'Other Project X', 'task': 'External work', 'service': '', 'phase': '', 'description': '', 'hours': 8.0},
        {'date': '2026-05-04', 'employee': 'Omar Khaled', 'project': 'Other Project X', 'task': 'External', 'service': '', 'phase': '', 'description': '', 'hours': 8.0},
    ]

def get_all_timesheets(date_from=None, date_to=None, phases=None, all_projects=False):
    """Unified getter — Odoo if available, else demo. phases = list or None."""
    _proj_name = active_project_name()
    ts = odoo.get_timesheets(project_name=_proj_name, date_from=date_from, date_to=date_to,
                             phase_filter=phases, all_projects=all_projects)
    if ts is None:
        data = get_demo_timesheets()
        if not all_projects:
            data = [d for d in data if d['project'] == _proj_name]
        if date_from:
            data = [d for d in data if d.get('date', '') >= date_from]
        if date_to:
            data = [d for d in data if d.get('date', '') <= date_to]
        if phases:
            phase_list = phases if isinstance(phases, list) else [phases]
            data = [d for d in data if d.get('phase', '') in phase_list]
        return data, False
    return [normalize_timesheet(e) for e in ts], True

def is_working_day(d):
    if isinstance(d, str):
        d = datetime.strptime(d, '%Y-%m-%d').date()
    return d.weekday() not in WEEKEND_DAYS

def get_working_days_between(start_date, end_date):
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    days = []
    current = start_date
    while current <= end_date:
        if is_working_day(current):
            days.append(current.isoformat())
        current += timedelta(days=1)
    return days

# ============================================================
# ROUTES
# ============================================================
@app.errorhandler(Exception)
def handle_error(e):
    logger.error(f"Error: {e}\n{traceback.format_exc()}")
    return jsonify({'error': str(e)}), 500

# ════════════════════════════════════════════════════════════════════
# AUTH — simple username/password from env
# ════════════════════════════════════════════════════════════════════
DASHBOARD_USER = os.environ.get('DASHBOARD_USER', 'codelab')
DASHBOARD_PASS = os.environ.get('DASHBOARD_PASS', 'pmo2026')


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        username = data.get('username') or request.form.get('username', '')
        password = data.get('password') or request.form.get('password', '')
        if username == DASHBOARD_USER and password == DASHBOARD_PASS:
            session['logged_in'] = True
            session.permanent = True
            return jsonify({'ok': True, 'redirect': '/projects'}) if request.is_json else redirect('/projects')
        error = 'Invalid credentials'
        if request.is_json:
            return jsonify({'ok': False, 'error': error}), 401
    return render_template('partials/login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


@app.route('/')
@login_required
def index():
    # If no project selected, redirect to projects list
    if not session.get('project_name'):
        return redirect('/projects')
    _, proj_name = get_active_project()
    return render_template('index.html', project_name=proj_name)


@app.route('/projects')
@login_required
def projects_list():
    """Projects selection page — lists all Odoo projects with stage info."""
    # Try both locations
    import os
    tmpl_dir = app.template_folder
    return render_template('partials/projects.html')


@app.route('/project/select', methods=['POST'])
def project_select():
    """Set the active project in session and redirect to dashboard."""
    data = request.get_json() or {}
    project_id   = data.get('project_id')
    project_name = data.get('project_name', '')
    if not project_name:
        return jsonify({'error': 'project_name required'}), 400
    session['project_id']   = project_id
    session['project_name'] = project_name
    # Force redirect to overview tab (add hash)
    return jsonify({'ok': True, 'redirect': '/#overview'})



# ════════════════════════════════════════════════════════════════════
# GLOBAL TRAVEL & PROMOTIONS  (not project-specific)
# ════════════════════════════════════════════════════════════════════

TRAVEL_SEED = [{"id": 1, "name": "Rabab Hosney", "position": "Team Lead", "start_date": "2023-10-01", "end_date": "2023-11-09", "notes": "", "status": "confirmed"}, {"id": 2, "name": "Sara Samir", "position": "Senior", "start_date": "2023-10-07", "end_date": "2024-01-04", "notes": "", "status": "confirmed"}, {"id": 3, "name": "Ahmed Helmi", "position": "Project Manager", "start_date": "2023-10-09", "end_date": "2023-12-14", "notes": "", "status": "confirmed"}, {"id": 4, "name": "Omar Mohamed", "position": "Senior", "start_date": "2023-10-14", "end_date": "2023-12-30", "notes": "", "status": "confirmed"}, {"id": 5, "name": "Youssef Ashraf", "position": "junior", "start_date": "2023-10-14", "end_date": "2023-12-28", "notes": "", "status": "confirmed"}, {"id": 6, "name": "Islam Ashraf", "position": "Senior", "start_date": "2023-10-17", "end_date": "2024-01-12", "notes": "", "status": "confirmed"}, {"id": 7, "name": "Hanan Mohamed", "position": "Team Lead", "start_date": "2023-10-20", "end_date": "2023-12-02", "notes": "", "status": "confirmed"}, {"id": 8, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2023-10-28", "end_date": "2024-01-04", "notes": "", "status": "confirmed"}, {"id": 9, "name": "Hazem Khater", "position": "junior", "start_date": "2023-10-28", "end_date": "2024-01-26", "notes": "", "status": "confirmed"}, {"id": 10, "name": "Naiim Bassili", "position": "Senior", "start_date": "2023-10-28", "end_date": "2024-01-21", "notes": "", "status": "confirmed"}, {"id": 11, "name": "Mohamed Aql", "position": "Senior", "start_date": "2023-11-04", "end_date": "2024-03-01", "notes": "", "status": "confirmed"}, {"id": 12, "name": "Ahmed Hamdy", "position": "junior", "start_date": "2023-11-05", "end_date": "2024-01-18", "notes": "", "status": "confirmed"}, {"id": 13, "name": "Aicha Chatti", "position": "Senior", "start_date": "2023-11-11", "end_date": "2024-01-21", "notes": "", "status": "confirmed"}, {"id": 14, "name": "Mina Kamal", "position": "Project Manager", "start_date": "2023-11-17", "end_date": "2023-12-29", "notes": "", "status": "confirmed"}, {"id": 15, "name": "Hoda Nabil", "position": "Team Lead", "start_date": "2023-11-25", "end_date": "2024-01-26", "notes": "", "status": "confirmed"}, {"id": 16, "name": "Aya Khattab", "position": "Senior", "start_date": "2023-11-26", "end_date": "2023-12-31", "notes": "", "status": "confirmed"}, {"id": 17, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2023-11-26", "end_date": "2023-12-22", "notes": "", "status": "confirmed"}, {"id": 18, "name": "Rania Abdulrahman", "position": "junior", "start_date": "2023-12-09", "end_date": "2024-03-01", "notes": "", "status": "confirmed"}, {"id": 19, "name": "Mustafa Medhat", "position": "junior", "start_date": "2023-12-15", "end_date": "2024-02-15", "notes": "", "status": "confirmed"}, {"id": 20, "name": "Mohamed Nabil", "position": "Project Manager", "start_date": "2023-12-20", "end_date": "2024-03-14", "notes": "", "status": "confirmed"}, {"id": 21, "name": "Hala Ali", "position": "Project Manager", "start_date": "2023-12-21", "end_date": "2024-03-01", "notes": "", "status": "confirmed"}, {"id": 22, "name": "Essam Emad", "position": "junior", "start_date": "2023-12-23", "end_date": "2024-03-21", "notes": "", "status": "confirmed"}, {"id": 23, "name": "ZeinElabdeen Trabulsi", "position": "junior", "start_date": "2023-12-24", "end_date": "2024-03-20", "notes": "", "status": "confirmed"}, {"id": 24, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2023-12-30", "end_date": "2024-01-26", "notes": "", "status": "confirmed"}, {"id": 25, "name": "Islam Mohamed", "position": "junior", "start_date": "2024-01-02", "end_date": "2024-02-15", "notes": "", "status": "confirmed"}, {"id": 26, "name": "Mostafa Ahmed", "position": "Senior", "start_date": "2024-01-02", "end_date": "2024-02-15", "notes": "", "status": "confirmed"}, {"id": 27, "name": "Aya Khattab", "position": "Senior", "start_date": "2024-01-05", "end_date": "2024-03-08", "notes": "", "status": "confirmed"}, {"id": 28, "name": "Hazem Ezz", "position": "junior", "start_date": "2024-01-05", "end_date": "2024-01-19", "notes": "", "status": "confirmed"}, {"id": 29, "name": "Assma Ibrahim", "position": "Team Lead", "start_date": "2024-01-06", "end_date": "2024-03-21", "notes": "", "status": "confirmed"}, {"id": 30, "name": "Hazem Emam", "position": "Senior", "start_date": "2024-01-06", "end_date": "2024-04-03", "notes": "", "status": "confirmed"}, {"id": 31, "name": "Mohamed Nagy", "position": "junior", "start_date": "2024-01-06", "end_date": "2024-03-30", "notes": "", "status": "confirmed"}, {"id": 32, "name": "Mohamed Shafik", "position": "junior", "start_date": "2024-01-06", "end_date": "2024-04-03", "notes": "", "status": "confirmed"}, {"id": 33, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2024-01-13", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 34, "name": "Ahmed Khaled", "position": "junior", "start_date": "2024-01-13", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 35, "name": "Mina Kamal", "position": "Project Manager", "start_date": "2024-01-13", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 36, "name": "Youssef Ashraf", "position": "junior", "start_date": "2024-01-13", "end_date": "2024-03-26", "notes": "", "status": "confirmed"}, {"id": 37, "name": "Samah Elsayed", "position": "Project Manager", "start_date": "2024-01-14", "end_date": "2024-02-09", "notes": "", "status": "confirmed"}, {"id": 38, "name": "Menna Elshtery", "position": "Senior", "start_date": "2024-01-18", "end_date": "2024-03-31", "notes": "", "status": "confirmed"}, {"id": 39, "name": "Elhussien Elsayed", "position": "Senior", "start_date": "2024-01-20", "end_date": "2024-03-28", "notes": "", "status": "confirmed"}, {"id": 40, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2024-01-21", "end_date": "2024-04-02", "notes": "", "status": "confirmed"}, {"id": 41, "name": "Naiim Bassili", "position": "Senior", "start_date": "2024-01-26", "end_date": "2024-02-15", "notes": "", "status": "confirmed"}, {"id": 42, "name": "Ahmed Helmi", "position": "Project Manager", "start_date": "2024-01-27", "end_date": "2024-03-03", "notes": "", "status": "confirmed"}, {"id": 43, "name": "Aicha Chatti", "position": "Senior", "start_date": "2024-01-30", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 44, "name": "Omar Mohamed", "position": "Senior", "start_date": "2024-01-30", "end_date": "2024-03-02", "notes": "", "status": "confirmed"}, {"id": 45, "name": "Laila Khaled", "position": "Project Manager", "start_date": "2024-01-31", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 46, "name": "Hazem Khater", "position": "junior", "start_date": "2024-02-03", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 47, "name": "Diaa Waheed", "position": "Team Lead", "start_date": "2024-02-16", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 48, "name": "Mohamed Tarek", "position": "junior", "start_date": "2024-02-16", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 49, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2024-02-17", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 50, "name": "Ahmed Gamel", "position": "junior", "start_date": "2024-02-17", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 51, "name": "Ali Mohamed", "position": "Senior", "start_date": "2024-02-17", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 52, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2024-02-17", "end_date": "2024-03-07", "notes": "", "status": "confirmed"}, {"id": 53, "name": "Mahmoud Shaban", "position": "Project Manager", "start_date": "2024-02-23", "end_date": "2024-03-09", "notes": "", "status": "confirmed"}, {"id": 54, "name": "Moamen Ibrahim", "position": "junior", "start_date": "2024-02-24", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 55, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2024-02-24", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 56, "name": "Rania Abdulrahman", "position": "junior", "start_date": "2024-03-09", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 57, "name": "Ahmed Hamdy", "position": "Senior", "start_date": "2024-03-13", "end_date": "2024-04-04", "notes": "", "status": "confirmed"}, {"id": 58, "name": "Hala Ali", "position": "Project Manager", "start_date": "2024-03-15", "end_date": "2024-04-05", "notes": "", "status": "confirmed"}, {"id": 59, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2024-04-13", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 60, "name": "Elhussien Elsayed", "position": "Senior", "start_date": "2024-04-13", "end_date": "2024-06-07", "notes": "", "status": "confirmed"}, {"id": 61, "name": "Hala Ali", "position": "Project Manager", "start_date": "2024-04-13", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 62, "name": "Hazem Emam", "position": "Senior", "start_date": "2024-04-13", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 63, "name": "Hazem Khater", "position": "junior", "start_date": "2024-04-13", "end_date": "2024-06-13", "notes": "", "status": "confirmed"}, {"id": 64, "name": "Mohamed Tarek", "position": "Junior", "start_date": "2024-04-13", "end_date": "2024-05-11", "notes": "", "status": "confirmed"}, {"id": 65, "name": "Rania Abdulrahman", "position": "Senior", "start_date": "2024-04-13", "end_date": "2024-05-11", "notes": "", "status": "confirmed"}, {"id": 66, "name": "Mohamed Nagy", "position": "junior", "start_date": "2024-04-14", "end_date": "2024-06-13", "notes": "", "status": "confirmed"}, {"id": 67, "name": "Abdelaziem Ashraf", "position": "Junior", "start_date": "2024-04-15", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 68, "name": "Youssef Ashraf", "position": "junior", "start_date": "2024-04-15", "end_date": "2024-06-13", "notes": "", "status": "confirmed"}, {"id": 69, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2024-04-16", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 70, "name": "Menna Elshtery", "position": "Senior", "start_date": "2024-04-17", "end_date": "2024-06-04", "notes": "", "status": "confirmed"}, {"id": 71, "name": "Diaa Waheed", "position": "Team Lead", "start_date": "2024-04-19", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 72, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2024-04-20", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 73, "name": "Moamen Ibrahim", "position": "Junior", "start_date": "2024-04-20", "end_date": "2024-07-16", "notes": "", "status": "confirmed"}, {"id": 74, "name": "ZienElabdeen Trabelsi", "position": "junior", "start_date": "2024-04-20", "end_date": "2024-06-10", "notes": "", "status": "confirmed"}, {"id": 75, "name": "Laila Khaled", "position": "Project Manager", "start_date": "2024-04-22", "end_date": "2024-05-11", "notes": "", "status": "confirmed"}, {"id": 76, "name": "Karim Khamis", "position": "Senior", "start_date": "2024-04-26", "end_date": "2024-08-02", "notes": "", "status": "confirmed"}, {"id": 77, "name": "Nada Hany", "position": "junior", "start_date": "2024-04-27", "end_date": "2024-06-15", "notes": "", "status": "confirmed"}, {"id": 78, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2024-05-03", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 79, "name": "Mohamed Aql", "position": "Senior", "start_date": "2024-05-04", "end_date": "2024-06-13", "notes": "", "status": "confirmed"}, {"id": 80, "name": "Mohannad Hany", "position": "Senior", "start_date": "2024-05-04", "end_date": "2024-06-13", "notes": "", "status": "confirmed"}, {"id": 81, "name": "Marwan Asem", "position": "Senior", "start_date": "2024-05-10", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 82, "name": "Ahmed Darwish", "position": "Team Lead", "start_date": "2024-05-18", "end_date": "2024-08-08", "notes": "", "status": "confirmed"}, {"id": 83, "name": "Eslam Hamada", "position": "Senior", "start_date": "2024-05-18", "end_date": "2024-08-15", "notes": "", "status": "confirmed"}, {"id": 84, "name": "Hazem Ezz", "position": "Senior", "start_date": "2024-05-18", "end_date": "2024-08-15", "notes": "", "status": "confirmed"}, {"id": 85, "name": "Mohamed Salah", "position": "Senior", "start_date": "2024-05-18", "end_date": "2024-06-15", "notes": "", "status": "confirmed"}, {"id": 86, "name": "Sara Samir", "position": "Senior", "start_date": "2024-05-18", "end_date": "2024-08-15", "notes": "", "status": "confirmed"}, {"id": 87, "name": "Mohamed Nabil", "position": "Project Manager", "start_date": "2024-05-20", "end_date": "2024-08-01", "notes": "", "status": "confirmed"}, {"id": 88, "name": "Mohamed Shafik", "position": "junior", "start_date": "2024-05-25", "end_date": "2024-08-09", "notes": "", "status": "confirmed"}, {"id": 89, "name": "Essam Emad", "position": "junior", "start_date": "2024-05-26", "end_date": "2024-06-14", "notes": "", "status": "confirmed"}, {"id": 90, "name": "Rabab Hosney", "position": "Team Lead", "start_date": "2024-06-07", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 91, "name": "Fady Ayman", "position": "junior", "start_date": "2024-06-19", "end_date": "2024-08-01", "notes": "", "status": "confirmed"}, {"id": 92, "name": "Mohamed Anter", "position": "Team Lead", "start_date": "2024-06-21", "end_date": "2024-08-30", "notes": "", "status": "confirmed"}, {"id": 93, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2024-06-21", "end_date": "2024-09-14", "notes": "", "status": "confirmed"}, {"id": 94, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2024-06-22", "end_date": "2024-09-16", "notes": "", "status": "confirmed"}, {"id": 95, "name": "Mohamed Salah", "position": "Senior", "start_date": "2024-06-22", "end_date": "2024-09-14", "notes": "", "status": "confirmed"}, {"id": 96, "name": "Nada Hany", "position": "junior", "start_date": "2024-06-22", "end_date": "2024-09-14", "notes": "", "status": "confirmed"}, {"id": 97, "name": "Omar Mahrous", "position": "Senior", "start_date": "2024-06-22", "end_date": "2024-08-13", "notes": "", "status": "confirmed"}, {"id": 98, "name": "ZienElabdeen Trabelsi", "position": "Senior", "start_date": "2024-06-22", "end_date": "2024-08-30", "notes": "", "status": "confirmed"}, {"id": 99, "name": "Hazem Emam", "position": "Senior", "start_date": "2024-06-23", "end_date": "2024-07-18", "notes": "", "status": "confirmed"}, {"id": 100, "name": "Ahmed Kamel", "position": "junior", "start_date": "2024-06-25", "end_date": "2024-08-01", "notes": "", "status": "confirmed"}, {"id": 101, "name": "Mohamed Ali", "position": "junior", "start_date": "2024-06-25", "end_date": "2024-08-01", "notes": "", "status": "confirmed"}, {"id": 102, "name": "Mohamed Latif", "position": "Senior", "start_date": "2024-06-27", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 103, "name": "Marwan Asem", "position": "Senior", "start_date": "2024-06-28", "end_date": "2024-09-25", "notes": "", "status": "confirmed"}, {"id": 104, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2024-06-29", "end_date": "2024-08-29", "notes": "", "status": "confirmed"}, {"id": 105, "name": "Hazem Khater", "position": "junior", "start_date": "2024-06-29", "end_date": "2024-09-27", "notes": "", "status": "confirmed"}, {"id": 106, "name": "Mohamed Nagy", "position": "junior", "start_date": "2024-06-29", "end_date": "2024-09-26", "notes": "", "status": "confirmed"}, {"id": 107, "name": "Youssef Ashraf", "position": "junior", "start_date": "2024-06-29", "end_date": "2024-09-26", "notes": "", "status": "confirmed"}, {"id": 108, "name": "Ahmed Gamel", "position": "Senior", "start_date": "2024-07-06", "end_date": "2024-10-03", "notes": "", "status": "confirmed"}, {"id": 109, "name": "Essam Emad", "position": "junior", "start_date": "2024-07-06", "end_date": "2024-10-03", "notes": "", "status": "confirmed"}, {"id": 110, "name": "Mohamed Aql", "position": "Senior", "start_date": "2024-07-06", "end_date": "2024-10-03", "notes": "", "status": "confirmed"}, {"id": 111, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2024-07-13", "end_date": "2024-10-04", "notes": "", "status": "confirmed"}, {"id": 112, "name": "Hala Ali", "position": "Project Manager", "start_date": "2024-07-19", "end_date": "2024-09-17", "notes": "", "status": "confirmed"}, {"id": 113, "name": "Menna Sameh", "position": "Senior", "start_date": "2024-07-23", "end_date": "2024-08-09", "notes": "", "status": "confirmed"}, {"id": 114, "name": "Ali Mohamed", "position": "Senior", "start_date": "2024-07-26", "end_date": "2024-09-30", "notes": "", "status": "confirmed"}, {"id": 115, "name": "Mohamed Nabil", "position": "Project Manager", "start_date": "2024-08-10", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 116, "name": "Ahmed Darwish", "position": "Team Lead", "start_date": "2024-08-13", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 117, "name": "Eslam Hamada", "position": "Senior", "start_date": "2024-08-16", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 118, "name": "Hazem Ezz", "position": "Senior", "start_date": "2024-08-16", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 119, "name": "Karim Khamis", "position": "Senior", "start_date": "2024-08-16", "end_date": "2024-11-01", "notes": "", "status": "confirmed"}, {"id": 120, "name": "Sara Samir", "position": "Senior", "start_date": "2024-08-16", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 121, "name": "Nada Ramadan", "position": "junior", "start_date": "2024-08-17", "end_date": "2024-08-30", "notes": "", "status": "confirmed"}, {"id": 122, "name": "Laila Khaled", "position": "Project Manager", "start_date": "2024-08-19", "end_date": "2024-09-13", "notes": "", "status": "confirmed"}, {"id": 123, "name": "Mohamed ElDegwy", "position": "Team Lead", "start_date": "2024-08-23", "end_date": "2024-09-07", "notes": "", "status": "confirmed"}, {"id": 124, "name": "Mohamed ElDegwy", "position": "Team Lead", "start_date": "2024-09-07", "end_date": "2024-09-15", "notes": "", "status": "confirmed"}, {"id": 125, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2024-09-09", "end_date": "2024-11-28", "notes": "", "status": "confirmed"}, {"id": 126, "name": "Menna Awad", "position": "junior", "start_date": "2024-09-14", "end_date": "2024-11-30", "notes": "", "status": "confirmed"}, {"id": 127, "name": "Mohamed Nabil", "position": "Project Manager", "start_date": "2024-09-16", "end_date": "2024-10-10", "notes": "", "status": "confirmed"}, {"id": 128, "name": "Laila Khaled", "position": "Project Manager", "start_date": "2024-09-23", "end_date": "2024-10-31", "notes": "", "status": "confirmed"}, {"id": 129, "name": "Ahmed Ashraf", "position": "junior", "start_date": "2024-09-27", "end_date": "2024-12-24", "notes": "", "status": "confirmed"}, {"id": 130, "name": "Mohamed Nagy", "position": "junior", "start_date": "2024-09-27", "end_date": "2024-11-23", "notes": "", "status": "confirmed"}, {"id": 131, "name": "ZienElabdeen Trabelsi", "position": "Senior", "start_date": "2024-09-27", "end_date": "2024-12-25", "notes": "", "status": "confirmed"}, {"id": 132, "name": "Mohamed Salah", "position": "Senior", "start_date": "2024-09-28", "end_date": "2024-11-15", "notes": "", "status": "confirmed"}, {"id": 133, "name": "Hala Ali", "position": "Project Manager", "start_date": "2024-10-04", "end_date": "2024-11-10", "notes": "", "status": "confirmed"}, {"id": 134, "name": "Mohamed Anter", "position": "Team Lead", "start_date": "2024-10-04", "end_date": "2024-10-18", "notes": "", "status": "confirmed"}, {"id": 135, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2024-10-05", "end_date": "2025-01-01", "notes": "", "status": "confirmed"}, {"id": 136, "name": "Ali Ezzat", "position": "Senior", "start_date": "2024-10-06", "end_date": "2024-11-15", "notes": "", "status": "confirmed"}, {"id": 137, "name": "Youssef Ashraf", "position": "junior", "start_date": "2024-10-06", "end_date": "2024-12-26", "notes": "", "status": "confirmed"}, {"id": 138, "name": "Aya Khattab", "position": "Senior", "start_date": "2024-10-11", "end_date": "2024-12-06", "notes": "", "status": "confirmed"}, {"id": 139, "name": "Hazem Khater", "position": "junior", "start_date": "2024-10-11", "end_date": "2025-01-01", "notes": "", "status": "confirmed"}, {"id": 140, "name": "Ahmed Gamel", "position": "Senior", "start_date": "2024-10-13", "end_date": "2025-01-10", "notes": "", "status": "confirmed"}, {"id": 141, "name": "Hoda Nabil", "position": "Team Lead", "start_date": "2024-10-14", "end_date": "2024-11-15", "notes": "", "status": "confirmed"}, {"id": 142, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2024-10-19", "end_date": "2025-01-16", "notes": "", "status": "confirmed"}, {"id": 143, "name": "Heba Abdulmunem", "position": "junior", "start_date": "2024-10-25", "end_date": "2025-01-18", "notes": "", "status": "confirmed"}, {"id": 144, "name": "Menna Elshtery", "position": "Senior", "start_date": "2024-10-25", "end_date": "2025-01-16", "notes": "", "status": "confirmed"}, {"id": 145, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2024-10-29", "end_date": "2024-12-19", "notes": "", "status": "confirmed"}, {"id": 146, "name": "Basem Mohamed", "position": "Project Manager", "start_date": "2024-11-02", "end_date": "2025-01-24", "notes": "", "status": "confirmed"}, {"id": 147, "name": "Essam Emad", "position": "junior", "start_date": "2024-11-02", "end_date": "2025-01-24", "notes": "", "status": "confirmed"}, {"id": 148, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2024-11-08", "end_date": "2025-01-30", "notes": "", "status": "confirmed"}, {"id": 149, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2024-11-09", "end_date": "2025-01-31", "notes": "", "status": "confirmed"}, {"id": 150, "name": "Amr Emam", "position": "junior", "start_date": "2024-11-09", "end_date": "2024-12-20", "notes": "", "status": "confirmed"}, {"id": 151, "name": "Youssef Khairy", "position": "junior", "start_date": "2024-11-09", "end_date": "2024-12-19", "notes": "", "status": "confirmed"}, {"id": 152, "name": "Omar Mohamed", "position": "Senior", "start_date": "2024-11-10", "end_date": "2025-01-11", "notes": "", "status": "confirmed"}, {"id": 153, "name": "Omar Sami", "position": "junior", "start_date": "2024-11-12", "end_date": "2025-02-06", "notes": "", "status": "confirmed"}, {"id": 154, "name": "Ahmed Kamel", "position": "junior", "start_date": "2024-11-15", "end_date": "2025-02-14", "notes": "", "status": "confirmed"}, {"id": 155, "name": "Mohamed ElDegwy", "position": "Team Lead", "start_date": "2024-11-15", "end_date": "2025-02-07", "notes": "", "status": "confirmed"}, {"id": 156, "name": "Karim Khamis", "position": "Senior", "start_date": "2024-11-16", "end_date": "2025-01-31", "notes": "", "status": "confirmed"}, {"id": 157, "name": "Menna Sameh", "position": "Senior", "start_date": "2024-11-30", "end_date": "2024-12-13", "notes": "", "status": "confirmed"}, {"id": 158, "name": "Hazem Emam", "position": "Senior", "start_date": "2024-12-14", "end_date": "2025-03-08", "notes": "", "status": "confirmed"}, {"id": 159, "name": "Aya Saeed", "position": "Team Lead", "start_date": "2024-12-27", "end_date": "2025-03-22", "notes": "", "status": "confirmed"}, {"id": 160, "name": "Ahmed Ashraf", "position": "junior", "start_date": "2024-12-28", "end_date": "2025-03-21", "notes": "", "status": "confirmed"}, {"id": 161, "name": "Fady Ayman", "position": "junior", "start_date": "2024-12-28", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 162, "name": "Mohamed Ammar", "position": "Senior", "start_date": "2024-12-28", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 163, "name": "Mohamed Latif", "position": "Senior", "start_date": "2024-12-28", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 164, "name": "Mostafa Ahmed", "position": "Senior", "start_date": "2024-12-28", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 165, "name": "Youssed Bardisy", "position": "junior", "start_date": "2024-12-28", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 166, "name": "Youssef Morsi", "position": "junior", "start_date": "2024-12-28", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 167, "name": "Ahmed Gamel", "position": "Senior", "start_date": "2025-01-01", "end_date": "2025-01-31", "notes": "", "status": "confirmed"}, {"id": 168, "name": "Mohamed Anter", "position": "Team Lead", "start_date": "2025-01-02", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 169, "name": "Ali Ezzat", "position": "Senior", "start_date": "2025-01-04", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 170, "name": "Mohamed Aql", "position": "Senior", "start_date": "2025-01-04", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 171, "name": "Mohab Akeel", "position": "Senior", "start_date": "2025-01-10", "end_date": "2025-03-27", "notes": "", "status": "confirmed"}, {"id": 172, "name": "Mohamed Hatem", "position": "Senior", "start_date": "2025-01-10", "end_date": "2025-03-28", "notes": "", "status": "confirmed"}, {"id": 173, "name": "Amr Emam", "position": "junior", "start_date": "2025-01-11", "end_date": "2025-02-07", "notes": "", "status": "confirmed"}, {"id": 174, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2025-01-11", "end_date": "2025-02-27", "notes": "", "status": "confirmed"}, {"id": 175, "name": "Mohamed Salah", "position": "Senior", "start_date": "2025-01-11", "end_date": "2025-03-25", "notes": "", "status": "confirmed"}, {"id": 176, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2025-01-11", "end_date": "2025-02-28", "notes": "", "status": "confirmed"}, {"id": 177, "name": "Youssef Khairy", "position": "Senior", "start_date": "2025-01-11", "end_date": "2025-02-28", "notes": "", "status": "confirmed"}, {"id": 178, "name": "Ahmed Gamel", "position": "Senior", "start_date": "2025-01-12", "end_date": "2025-01-31", "notes": "", "status": "confirmed"}, {"id": 179, "name": "Omar Mahrous", "position": "Senior", "start_date": "2025-01-12", "end_date": "2025-03-27", "notes": "", "status": "confirmed"}, {"id": 180, "name": "Zein Gaber", "position": "Team Lead", "start_date": "2025-01-13", "end_date": "2025-03-27", "notes": "", "status": "confirmed"}, {"id": 181, "name": "Omar Mohamed", "position": "Senior", "start_date": "2025-01-16", "end_date": "2025-03-27", "notes": "", "status": "confirmed"}, {"id": 182, "name": "Menna Elshtery", "position": "Senior", "start_date": "2025-01-18", "end_date": "2025-01-29", "notes": "", "status": "confirmed"}, {"id": 183, "name": "Youssef Ashraf", "position": "junior", "start_date": "2025-01-18", "end_date": "2025-03-20", "notes": "", "status": "confirmed"}, {"id": 184, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2025-01-20", "end_date": "2025-03-27", "notes": "", "status": "confirmed"}, {"id": 185, "name": "Menna Awad", "position": "junior", "start_date": "2025-01-25", "end_date": "2025-03-26", "notes": "", "status": "confirmed"}, {"id": 186, "name": "Ahmed Gamel", "position": "Senior", "start_date": "2025-02-01", "end_date": "2025-02-28", "notes": "", "status": "confirmed"}, {"id": 187, "name": "Ahmed Samir", "position": "junior", "start_date": "2025-02-01", "end_date": "2025-03-24", "notes": "", "status": "confirmed"}, {"id": 188, "name": "Aya Khattab", "position": "Senior", "start_date": "2025-02-01", "end_date": "2025-03-28", "notes": "", "status": "confirmed"}, {"id": 189, "name": "Essam Emad", "position": "junior", "start_date": "2025-02-02", "end_date": "2025-03-21", "notes": "", "status": "confirmed"}, {"id": 190, "name": "Eslam Hamada", "position": "Senior", "start_date": "2025-02-07", "end_date": "2025-05-06", "notes": "", "status": "confirmed"}, {"id": 191, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2025-02-07", "end_date": "2025-03-25", "notes": "", "status": "confirmed"}, {"id": 192, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2025-02-08", "end_date": "2025-03-24", "notes": "", "status": "confirmed"}, {"id": 193, "name": "Omar Sami", "position": "junior", "start_date": "2025-02-10", "end_date": "2025-03-27", "notes": "", "status": "confirmed"}, {"id": 194, "name": "Heba Abdulmunem", "position": "junior", "start_date": "2025-02-12", "end_date": "2025-03-28", "notes": "", "status": "confirmed"}, {"id": 195, "name": "Karim Khamis", "position": "Senior", "start_date": "2025-02-15", "end_date": "2025-03-21", "notes": "", "status": "confirmed"}, {"id": 196, "name": "Ahmed Kamel", "position": "junior", "start_date": "2025-02-17", "end_date": "2025-03-28", "notes": "", "status": "confirmed"}, {"id": 197, "name": "Marwa Alaa", "position": "Senior", "start_date": "2025-02-21", "end_date": "2025-03-29", "notes": "", "status": "confirmed"}, {"id": 198, "name": "Marwan Asem", "position": "Senior", "start_date": "2025-02-21", "end_date": "2025-03-29", "notes": "", "status": "confirmed"}, {"id": 199, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2025-03-04", "end_date": "2025-03-25", "notes": "", "status": "confirmed"}, {"id": 200, "name": "Hazem Emam", "position": "Senior", "start_date": "2025-03-12", "end_date": "2025-03-25", "notes": "", "status": "confirmed"}, {"id": 201, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2025-04-03", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 202, "name": "Ahmed Samir", "position": "Senior", "start_date": "2025-04-03", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 203, "name": "Aya Khattab", "position": "Senior", "start_date": "2025-04-03", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 204, "name": "Fady Ayman", "position": "Senior", "start_date": "2025-04-03", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 205, "name": "Mostafa Ahmed", "position": "Senior", "start_date": "2025-04-03", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 206, "name": "Youssed Bardisy", "position": "junior", "start_date": "2025-04-03", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 207, "name": "Karim Khamis", "position": "Team Lead", "start_date": "2025-04-04", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 208, "name": "Ahmed Mamdoh", "position": "Senior", "start_date": "2025-04-05", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 209, "name": "Marwan Asem", "position": "Senior", "start_date": "2025-04-05", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 210, "name": "Menna Awad", "position": "junior", "start_date": "2025-04-05", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 211, "name": "Mohamed Ammar", "position": "Senior", "start_date": "2025-04-05", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 212, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2025-04-05", "end_date": "2025-06-06", "notes": "", "status": "confirmed"}, {"id": 213, "name": "Omar Sami", "position": "Senior", "start_date": "2025-04-05", "end_date": "2025-06-03", "notes": "", "status": "confirmed"}, {"id": 214, "name": "Zein Gaber", "position": "Team Lead", "start_date": "2025-04-05", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 215, "name": "Mohamed Aql", "position": "Senior", "start_date": "2025-04-06", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 216, "name": "Mohamed Latif", "position": "Team Lead", "start_date": "2025-04-06", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 217, "name": "Youssef Morsi", "position": "junior", "start_date": "2025-04-06", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 218, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2025-04-07", "end_date": "2025-05-26", "notes": "", "status": "confirmed"}, {"id": 219, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2025-04-07", "end_date": "2025-05-29", "notes": "", "status": "confirmed"}, {"id": 220, "name": "Hazem Emam", "position": "Senior", "start_date": "2025-04-07", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 221, "name": "Ahmed Kamel", "position": "junior", "start_date": "2025-04-08", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 222, "name": "Essam Emad", "position": "junior", "start_date": "2025-04-08", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 223, "name": "Mohab Akeel", "position": "Senior", "start_date": "2025-04-08", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 224, "name": "Mohamed Eldegwi", "position": "Team Lead", "start_date": "2025-04-10", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 225, "name": "Marwa Alaa", "position": "Senior", "start_date": "2025-04-11", "end_date": "2025-05-16", "notes": "", "status": "confirmed"}, {"id": 226, "name": "Mohamed Tamer", "position": "junior", "start_date": "2025-04-11", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 227, "name": "Ahmed Ashraf", "position": "junior", "start_date": "2025-04-12", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 228, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2025-04-12", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 229, "name": "Ali Ezzat", "position": "Senior", "start_date": "2025-04-12", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 230, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2025-04-19", "end_date": "2025-05-31", "notes": "", "status": "confirmed"}, {"id": 231, "name": "Youssef Khairy", "position": "Senior", "start_date": "2025-04-19", "end_date": "2025-05-30", "notes": "", "status": "confirmed"}, {"id": 232, "name": "Hazem Ezz", "position": "Senior", "start_date": "2025-04-21", "end_date": "2025-06-05", "notes": "", "status": "confirmed"}, {"id": 233, "name": "Ahmed Khaled", "position": "Senior", "start_date": "2025-04-26", "end_date": "2025-07-03", "notes": "", "status": "confirmed"}, {"id": 234, "name": "Mohamed Salah", "position": "Senior", "start_date": "2025-04-27", "end_date": "2025-06-27", "notes": "", "status": "confirmed"}, {"id": 235, "name": "Mohamed Shafik", "position": "Senior", "start_date": "2025-05-03", "end_date": "2025-07-31", "notes": "", "status": "confirmed"}, {"id": 236, "name": "Aahmed Araby", "position": "Senior", "start_date": "2025-05-10", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 237, "name": "Rabab Hosney", "position": "Team Lead", "start_date": "2025-05-12", "end_date": "2025-06-03", "notes": "", "status": "confirmed"}, {"id": 238, "name": "Amira Abdulgawad", "position": "junior", "start_date": "2025-05-13", "end_date": "2025-06-04", "notes": "", "status": "confirmed"}, {"id": 239, "name": "Ahmed Mamdoh", "position": "Senior", "start_date": "2025-06-11", "end_date": "2025-08-09", "notes": "", "status": "confirmed"}, {"id": 240, "name": "Menna Awad", "position": "junior", "start_date": "2025-06-11", "end_date": "2025-07-10", "notes": "", "status": "confirmed"}, {"id": 241, "name": "Mohamed Ammar", "position": "Senior", "start_date": "2025-06-11", "end_date": "2025-10-23", "notes": "", "status": "confirmed"}, {"id": 242, "name": "Zein Gaber", "position": "Team Lead", "start_date": "2025-06-11", "end_date": "2025-10-02", "notes": "", "status": "confirmed"}, {"id": 243, "name": "Mina Kamal", "position": "Project Manager", "start_date": "2025-06-13", "end_date": "2025-09-05", "notes": "", "status": "confirmed"}, {"id": 244, "name": "Mohamed Anter", "position": "Team Lead", "start_date": "2025-06-13", "end_date": "2025-09-04", "notes": "", "status": "confirmed"}, {"id": 245, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2025-06-14", "end_date": "2025-08-30", "notes": "", "status": "confirmed"}, {"id": 246, "name": "Ahmed Araby", "position": "Senior", "start_date": "2025-06-14", "end_date": "2025-10-15", "notes": "", "status": "confirmed"}, {"id": 247, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2025-06-14", "end_date": "2025-09-15", "notes": "", "status": "confirmed"}, {"id": 248, "name": "Ahmed Samir", "position": "Senior", "start_date": "2025-06-14", "end_date": "2025-09-10", "notes": "", "status": "confirmed"}, {"id": 249, "name": "Aya Saeed", "position": "Team Lead", "start_date": "2025-06-14", "end_date": "2025-09-10", "notes": "", "status": "confirmed"}, {"id": 250, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2025-06-14", "end_date": "2025-08-14", "notes": "", "status": "confirmed"}, {"id": 251, "name": "Issra Houssien", "position": "Project Manager", "start_date": "2025-06-14", "end_date": "2025-07-17", "notes": "", "status": "confirmed"}, {"id": 252, "name": "Karim Khamis", "position": "Team Lead", "start_date": "2025-06-14", "end_date": "2025-09-10", "notes": "", "status": "confirmed"}, {"id": 253, "name": "Mohamed Tamer", "position": "junior", "start_date": "2025-06-14", "end_date": "2025-07-31", "notes": "", "status": "confirmed"}, {"id": 254, "name": "Mostafa Ahmed", "position": "Senior", "start_date": "2025-06-14", "end_date": "2025-08-14", "notes": "", "status": "confirmed"}, {"id": 255, "name": "omar Sami", "position": "Senior", "start_date": "2025-06-14", "end_date": "2025-08-12", "notes": "", "status": "confirmed"}, {"id": 256, "name": "Youssed Bardisy", "position": "junior", "start_date": "2025-06-14", "end_date": "2025-09-03", "notes": "", "status": "confirmed"}, {"id": 257, "name": "Ali Mohamed", "position": "Senior", "start_date": "2025-06-15", "end_date": "2025-08-28", "notes": "", "status": "confirmed"}, {"id": 258, "name": "Mohamed Aql", "position": "Senior", "start_date": "2025-06-15", "end_date": "2025-09-10", "notes": "", "status": "confirmed"}, {"id": 259, "name": "Menna Sameh", "position": "Senior", "start_date": "2025-06-16", "end_date": "2025-07-09", "notes": "", "status": "confirmed"}, {"id": 260, "name": "Amr Emam", "position": "Senior", "start_date": "2025-06-17", "end_date": "2025-07-04", "notes": "", "status": "confirmed"}, {"id": 261, "name": "Hazem Emam", "position": "Senior", "start_date": "2025-06-18", "end_date": "2025-09-11", "notes": "", "status": "confirmed"}, {"id": 262, "name": "Adham Ali", "position": "Senior", "start_date": "2025-06-20", "end_date": "2025-09-17", "notes": "", "status": "confirmed"}, {"id": 263, "name": "Mohamed Ali", "position": "junior", "start_date": "2025-06-20", "end_date": "2025-08-22", "notes": "", "status": "confirmed"}, {"id": 264, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2025-06-20", "end_date": "2025-09-16", "notes": "", "status": "confirmed"}, {"id": 265, "name": "Ali Ezzat", "position": "Senior", "start_date": "2025-06-21", "end_date": "2025-08-21", "notes": "", "status": "confirmed"}, {"id": 266, "name": "Fady Ayman", "position": "Senior", "start_date": "2025-06-21", "end_date": "2025-09-12", "notes": "", "status": "confirmed"}, {"id": 267, "name": "Mohab Akeel", "position": "Senior", "start_date": "2025-06-21", "end_date": "2025-09-18", "notes": "", "status": "confirmed"}, {"id": 268, "name": "Youssef Morsi", "position": "junior", "start_date": "2025-06-21", "end_date": "2025-09-10", "notes": "", "status": "confirmed"}, {"id": 269, "name": "Naiim Bassili", "position": "Senior", "start_date": "2025-06-27", "end_date": "2025-09-21", "notes": "", "status": "confirmed"}, {"id": 270, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2025-06-28", "end_date": "2025-09-19", "notes": "", "status": "confirmed"}, {"id": 271, "name": "Aya Khattab", "position": "Senior", "start_date": "2025-07-03", "end_date": "2025-09-29", "notes": "", "status": "confirmed"}, {"id": 272, "name": "Ahmed Ashraf", "position": "junior", "start_date": "2025-07-04", "end_date": "2025-09-26", "notes": "", "status": "confirmed"}, {"id": 273, "name": "Ahmed Hamdy", "position": "Senior", "start_date": "2025-07-04", "end_date": "2025-09-26", "notes": "", "status": "confirmed"}, {"id": 274, "name": "Hazem Ezz", "position": "Senior", "start_date": "2025-07-04", "end_date": "2025-09-23", "notes": "", "status": "confirmed"}, {"id": 275, "name": "Marwan Asem", "position": "Senior", "start_date": "2025-07-06", "end_date": "2025-09-11", "notes": "", "status": "confirmed"}, {"id": 276, "name": "Hala Ali", "position": "Project Manager", "start_date": "2025-07-11", "end_date": "2025-10-03", "notes": "", "status": "confirmed"}, {"id": 277, "name": "Marwa Alaa", "position": "Senior", "start_date": "2025-07-11", "end_date": "2025-10-03", "notes": "", "status": "confirmed"}, {"id": 278, "name": "Nada Hany", "position": "junior", "start_date": "2025-07-11", "end_date": "2025-10-31", "notes": "", "status": "confirmed"}, {"id": 279, "name": "Samah Elsayed", "position": "Project Manager", "start_date": "2025-07-11", "end_date": "2025-09-03", "notes": "", "status": "confirmed"}, {"id": 280, "name": "Ahmed Gamel", "position": "Senior", "start_date": "2025-07-12", "end_date": "2025-10-09", "notes": "", "status": "confirmed"}, {"id": 281, "name": "Essam Emad", "position": "junior", "start_date": "2025-07-12", "end_date": "2025-08-29", "notes": "", "status": "confirmed"}, {"id": 282, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2025-07-12", "end_date": "2025-09-26", "notes": "", "status": "confirmed"}, {"id": 283, "name": "Waleed Mohamed", "position": "junior", "start_date": "2025-07-12", "end_date": "2025-10-09", "notes": "", "status": "confirmed"}, {"id": 284, "name": "Youssef Khairy", "position": "Senior", "start_date": "2025-07-12", "end_date": "2025-09-26", "notes": "", "status": "confirmed"}, {"id": 285, "name": "Menna Awad", "position": "junior", "start_date": "2025-07-15", "end_date": "2025-10-09", "notes": "", "status": "confirmed"}, {"id": 286, "name": "Mohamed Desoky", "position": "Senior", "start_date": "2025-07-15", "end_date": "2025-10-11", "notes": "", "status": "confirmed"}, {"id": 287, "name": "Ahmed Kamel", "position": "junior", "start_date": "2025-07-18", "end_date": "2025-10-11", "notes": "", "status": "confirmed"}, {"id": 288, "name": "Ayman Elsayed", "position": "junior", "start_date": "2025-07-18", "end_date": "2025-08-31", "notes": "", "status": "confirmed"}, {"id": 289, "name": "Elsaeed Elhosiney", "position": "junior", "start_date": "2025-07-19", "end_date": "2025-08-22", "notes": "", "status": "confirmed"}, {"id": 290, "name": "Esraa Ahmed", "position": "junior", "start_date": "2025-07-19", "end_date": "2025-08-16", "notes": "", "status": "confirmed"}, {"id": 291, "name": "Laila Khalid", "position": "Project Manager", "start_date": "2025-07-19", "end_date": "2025-09-02", "notes": "", "status": "confirmed"}, {"id": 292, "name": "Mohamed Hamdy", "position": "Team Lead", "start_date": "2025-07-19", "end_date": "2025-08-21", "notes": "", "status": "confirmed"}, {"id": 293, "name": "Mohamed Hatem", "position": "Senior", "start_date": "2025-07-26", "end_date": "2025-08-29", "notes": "", "status": "confirmed"}, {"id": 294, "name": "Omar Mohamed", "position": "Senior", "start_date": "2025-07-26", "end_date": "2025-10-10", "notes": "", "status": "confirmed"}, {"id": 295, "name": "Salah Elshenawi", "position": "junior", "start_date": "2025-08-01", "end_date": "2025-12-18", "notes": "", "status": "confirmed"}, {"id": 296, "name": "Sara Samir", "position": "Senior", "start_date": "2025-08-01", "end_date": "2025-10-04", "notes": "", "status": "confirmed"}, {"id": 297, "name": "Aya Mohamed", "position": "Project Manager", "start_date": "2025-08-02", "end_date": "2025-08-29", "notes": "", "status": "confirmed"}, {"id": 298, "name": "Menna Sameh", "position": "Senior", "start_date": "2025-08-02", "end_date": "2025-08-29", "notes": "", "status": "confirmed"}, {"id": 299, "name": "Mohamed Tamer", "position": "junior", "start_date": "2025-08-04", "end_date": "2025-10-31", "notes": "", "status": "confirmed"}, {"id": 300, "name": "Sandra Latif", "position": "senior", "start_date": "2025-08-08", "end_date": "2025-09-19", "notes": "", "status": "confirmed"}, {"id": 301, "name": "Basem Mohamed", "position": "Project Manager", "start_date": "2025-08-09", "end_date": "2025-09-12", "notes": "", "status": "confirmed"}, {"id": 302, "name": "Mohamed Salah", "position": "Senior", "start_date": "2025-08-15", "end_date": "2025-11-10", "notes": "", "status": "confirmed"}, {"id": 303, "name": "omar Sami", "position": "Senior", "start_date": "2025-08-16", "end_date": "2025-11-12", "notes": "", "status": "confirmed"}, {"id": 304, "name": "Heba Abdulmunem", "position": "junior", "start_date": "2025-08-18", "end_date": "2025-11-15", "notes": "", "status": "confirmed"}, {"id": 305, "name": "Ahmed Mamdoh", "position": "Senior", "start_date": "2025-08-22", "end_date": "2025-11-15", "notes": "", "status": "confirmed"}, {"id": 306, "name": "Salma Shammakh", "position": "Team Lead", "start_date": "2025-08-22", "end_date": "2025-10-20", "notes": "", "status": "confirmed"}, {"id": 307, "name": "Omar Elhady", "position": "Senior", "start_date": "2025-08-23", "end_date": "2025-11-15", "notes": "", "status": "confirmed"}, {"id": 308, "name": "Mohamed Tamer", "position": "junior", "start_date": "2025-08-26", "end_date": "2025-11-21", "notes": "", "status": "confirmed"}, {"id": 309, "name": "Haitham Nabil", "position": "Senior", "start_date": "2025-08-29", "end_date": "2025-10-31", "notes": "", "status": "confirmed"}, {"id": 310, "name": "Mohamed Tarek", "position": "Senior", "start_date": "2025-08-29", "end_date": "2025-11-21", "notes": "", "status": "confirmed"}, {"id": 311, "name": "Shady Mahfouz", "position": "Senior", "start_date": "2025-08-29", "end_date": "2025-11-21", "notes": "", "status": "confirmed"}, {"id": 312, "name": "Israa Houssien", "position": "Project Manager", "start_date": "2025-08-30", "end_date": "2025-10-30", "notes": "", "status": "confirmed"}, {"id": 313, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2025-09-01", "end_date": "2025-11-27", "notes": "", "status": "confirmed"}, {"id": 314, "name": "Mohamed Hatem", "position": "Senior", "start_date": "2025-09-02", "end_date": "2025-10-03", "notes": "", "status": "confirmed"}, {"id": 315, "name": "Reem Alaa", "position": "junior", "start_date": "2025-09-05", "end_date": "2025-10-09", "notes": "", "status": "confirmed"}, {"id": 316, "name": "Khaled Elbialy", "position": "Senior", "start_date": "2025-09-06", "end_date": "2025-11-06", "notes": "", "status": "confirmed"}, {"id": 317, "name": "Mohamed Hamdy", "position": "Team Lead", "start_date": "2025-09-06", "end_date": "2025-10-02", "notes": "", "status": "confirmed"}, {"id": 318, "name": "Mohamed Shafik", "position": "Senior", "start_date": "2025-09-07", "end_date": "2025-11-29", "notes": "", "status": "confirmed"}, {"id": 319, "name": "Essam Emad", "position": "junior", "start_date": "2025-09-08", "end_date": "2025-11-29", "notes": "", "status": "confirmed"}, {"id": 320, "name": "Ahmed soliman", "position": "junior", "start_date": "2025-09-12", "end_date": "2025-12-09", "notes": "", "status": "confirmed"}, {"id": 321, "name": "Mina Kamal", "position": "Project Manager", "start_date": "2025-09-13", "end_date": "2025-11-14", "notes": "", "status": "confirmed"}, {"id": 322, "name": "Hazem Emam", "position": "Senior", "start_date": "2025-09-15", "end_date": "2025-11-27", "notes": "", "status": "confirmed"}, {"id": 323, "name": "Omar Magdy", "position": "junior", "start_date": "2025-09-15", "end_date": "2025-09-26", "notes": "", "status": "confirmed"}, {"id": 324, "name": "Youssed Bardisy", "position": "junior", "start_date": "2025-09-15", "end_date": "2025-11-28", "notes": "", "status": "confirmed"}, {"id": 325, "name": "Karim Khamis", "position": "Team Lead", "start_date": "2025-09-19", "end_date": "2025-12-12", "notes": "", "status": "confirmed"}, {"id": 326, "name": "Abdelaziem Ashraf", "position": "junior", "start_date": "2025-09-20", "end_date": "2025-12-17", "notes": "", "status": "confirmed"}, {"id": 327, "name": "Ali Mohamed", "position": "Senior", "start_date": "2025-09-20", "end_date": "2025-11-27", "notes": "", "status": "confirmed"}, {"id": 328, "name": "Aya Saeed", "position": "Team Lead", "start_date": "2025-09-20", "end_date": "2025-12-17", "notes": "", "status": "confirmed"}, {"id": 329, "name": "Basem Mohamed", "position": "Project Manager", "start_date": "2025-09-20", "end_date": "2025-11-28", "notes": "", "status": "confirmed"}, {"id": 330, "name": "Fady Ayman", "position": "Senior", "start_date": "2025-09-20", "end_date": "2025-12-12", "notes": "", "status": "confirmed"}, {"id": 331, "name": "Ali Ezzat", "position": "Senior", "start_date": "2025-09-26", "end_date": "2025-12-23", "notes": "", "status": "confirmed"}, {"id": 332, "name": "Ahmed Samir", "position": "Senior", "start_date": "2025-09-27", "end_date": "2025-12-24", "notes": "", "status": "confirmed"}, {"id": 333, "name": "Youssef Morsi", "position": "junior", "start_date": "2025-09-27", "end_date": "2025-12-24", "notes": "", "status": "confirmed"}, {"id": 334, "name": "Mustafa Medhat", "position": "Senior", "start_date": "2025-10-03", "end_date": "2025-12-25", "notes": "", "status": "confirmed"}, {"id": 335, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2025-10-04", "end_date": "2025-12-31", "notes": "", "status": "confirmed"}, {"id": 336, "name": "Ahmed Abdulaziz", "position": "Team Lead", "start_date": "2025-10-04", "end_date": "2025-10-31", "notes": "", "status": "confirmed"}, {"id": 337, "name": "Mohamed Anter", "position": "Team Lead", "start_date": "2025-10-04", "end_date": "2025-12-26", "notes": "", "status": "confirmed"}, {"id": 338, "name": "Hazem Ezz", "position": "Senior", "start_date": "2025-10-05", "end_date": "2026-01-01", "notes": "", "status": "confirmed"}, {"id": 339, "name": "Ahmed Ashraf", "position": "junior", "start_date": "2025-10-11", "end_date": "2025-12-11", "notes": "", "status": "confirmed"}, {"id": 340, "name": "Amr Emam", "position": "Senior", "start_date": "2025-10-11", "end_date": "2026-01-02", "notes": "", "status": "confirmed"}, {"id": 341, "name": "Zein Gaber", "position": "Team Lead", "start_date": "2025-10-11", "end_date": "2026-01-08", "notes": "", "status": "confirmed"}, {"id": 342, "name": "Ahmed Hamdy", "position": "Senior", "start_date": "2025-10-18", "end_date": "2026-01-14", "notes": "", "status": "confirmed"}, {"id": 343, "name": "Mohab Akeel", "position": "Senior", "start_date": "2025-10-18", "end_date": "2025-11-27", "notes": "", "status": "confirmed"}, {"id": 344, "name": "Mohamed Galal", "position": "junior", "start_date": "2025-10-18", "end_date": "2026-01-08", "notes": "", "status": "confirmed"}, {"id": 345, "name": "Mohamed Hamdy", "position": "Team Lead", "start_date": "2025-10-18", "end_date": "2025-12-21", "notes": "", "status": "confirmed"}, {"id": 346, "name": "Menna Awad", "position": "junior", "start_date": "2025-10-20", "end_date": "2026-01-09", "notes": "", "status": "confirmed"}, {"id": 347, "name": "Ahmed Kamel", "position": "junior", "start_date": "2025-10-23", "end_date": "2026-01-02", "notes": "", "status": "confirmed"}, {"id": 348, "name": "Mohamed Ammar", "position": "Senior", "start_date": "2025-10-25", "end_date": "2025-11-20", "notes": "", "status": "confirmed"}, {"id": 349, "name": "Mohamed Desoky", "position": "Senior", "start_date": "2025-10-25", "end_date": "2025-12-31", "notes": "", "status": "confirmed"}, {"id": 350, "name": "Walid Mohamed", "position": "junior", "start_date": "2025-10-25", "end_date": "2026-01-17", "notes": "", "status": "confirmed"}, {"id": 351, "name": "Aya Khattab", "position": "Senior", "start_date": "2025-10-31", "end_date": "2026-01-28", "notes": "", "status": "confirmed"}, {"id": 352, "name": "Adham Ali", "position": "Senior", "start_date": "2025-11-01", "end_date": "2026-01-29", "notes": "", "status": "confirmed"}, {"id": 353, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2025-11-01", "end_date": "2026-01-20", "notes": "", "status": "confirmed"}, {"id": 354, "name": "Nada Hany", "position": "junior", "start_date": "2025-11-02", "end_date": "2025-12-05", "notes": "", "status": "confirmed"}, {"id": 355, "name": "Nourhan Khattab", "position": "Project Manager", "start_date": "2025-11-04", "end_date": "2025-11-17", "notes": "", "status": "confirmed"}, {"id": 356, "name": "Rabab Hosney", "position": "Team Lead", "start_date": "2025-11-07", "end_date": "2025-12-26", "notes": "", "status": "confirmed"}, {"id": 357, "name": "Kareem Waleed", "position": "Senior", "start_date": "2025-11-08", "end_date": "2025-12-31", "notes": "", "status": "confirmed"}, {"id": 358, "name": "Mohamed Latif", "position": "Team Lead", "start_date": "2025-11-08", "end_date": "2026-01-02", "notes": "", "status": "confirmed"}, {"id": 359, "name": "Mohamed Aql", "position": "Senior", "start_date": "2025-11-10", "end_date": "2026-01-01", "notes": "", "status": "confirmed"}, {"id": 360, "name": "Marwa Alaa", "position": "Senior", "start_date": "2025-11-11", "end_date": "2026-01-02", "notes": "", "status": "confirmed"}, {"id": 361, "name": "Haitham Nabil", "position": "Senior", "start_date": "2025-11-14", "end_date": "2026-01-29", "notes": "", "status": "confirmed"}, {"id": 362, "name": "Salma Shammakh", "position": "Team Lead", "start_date": "2025-11-14", "end_date": "2025-12-05", "notes": "", "status": "confirmed"}, {"id": 363, "name": "Eslam Hamada", "position": "Senior", "start_date": "2025-11-15", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 364, "name": "Reem Alaa", "position": "junior", "start_date": "2025-11-15", "end_date": "2026-02-12", "notes": "", "status": "confirmed"}, {"id": 365, "name": "Sara Elsayed", "position": "Project Manager", "start_date": "2025-11-15", "end_date": "2025-11-28", "notes": "", "status": "confirmed"}, {"id": 366, "name": "Mina Kamal", "position": "Project Manager", "start_date": "2025-11-22", "end_date": "2025-12-25", "notes": "", "status": "confirmed"}, {"id": 367, "name": "Omar Sami", "position": "Senior", "start_date": "2025-11-22", "end_date": "2026-02-18", "notes": "", "status": "confirmed"}, {"id": 368, "name": "Ahmed Mamdoh", "position": "Senior", "start_date": "2025-11-23", "end_date": "2026-02-21", "notes": "", "status": "confirmed"}, {"id": 369, "name": "Mohamed Eldegwy", "position": "Team Lead", "start_date": "2025-11-28", "end_date": "2025-12-26", "notes": "", "status": "confirmed"}, {"id": 370, "name": "Youssef Tarek", "position": "junior", "start_date": "2025-11-28", "end_date": "2026-02-25", "notes": "", "status": "confirmed"}, {"id": 371, "name": "Ahmed Khaled", "position": "Senior", "start_date": "2025-11-29", "end_date": "2026-02-24", "notes": "", "status": "confirmed"}, {"id": 372, "name": "Mohamed Ammar", "position": "Senior", "start_date": "2025-11-29", "end_date": "2026-02-26", "notes": "", "status": "confirmed"}, {"id": 373, "name": "Elsaeed Elhosiney", "position": "junior", "start_date": "2025-11-30", "end_date": "2026-02-26", "notes": "", "status": "confirmed"}, {"id": 374, "name": "Hazem Emam", "position": "Senior", "start_date": "2025-11-30", "end_date": "2026-02-26", "notes": "", "status": "confirmed"}, {"id": 375, "name": "Mostafa Mokhtar", "position": "Senior", "start_date": "2025-11-30", "end_date": "2026-02-26", "notes": "", "status": "confirmed"}, {"id": 376, "name": "Mohamed Tarek", "position": "Senior", "start_date": "2025-12-03", "end_date": "2026-01-30", "notes": "", "status": "confirmed"}, {"id": 377, "name": "Heba Abdulmunem", "position": "junior", "start_date": "2025-12-05", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 378, "name": "Mohamed Gouda", "position": "Senior", "start_date": "2025-12-05", "end_date": "2026-02-14", "notes": "", "status": "confirmed"}, {"id": 379, "name": "Mohamed Tamer", "position": "junior", "start_date": "2025-12-05", "end_date": "2026-02-02", "notes": "", "status": "confirmed"}, {"id": 380, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2025-12-13", "end_date": "2026-02-05", "notes": "", "status": "confirmed"}, {"id": 381, "name": "Mohamed Salah", "position": "Senior", "start_date": "2025-12-13", "end_date": "2026-03-12", "notes": "", "status": "confirmed"}, {"id": 382, "name": "Youssed Bardisy", "position": "junior", "start_date": "2025-12-13", "end_date": "2026-01-10", "notes": "", "status": "confirmed"}, {"id": 383, "name": "Hanan Mohamed", "position": "Team Lead", "start_date": "2025-12-15", "end_date": "2026-02-11", "notes": "", "status": "confirmed"}, {"id": 384, "name": "Omar Elhady", "position": "Senior", "start_date": "2025-12-18", "end_date": "2026-01-09", "notes": "", "status": "confirmed"}, {"id": 385, "name": "Shady Mahfouz", "position": "Senior", "start_date": "2025-12-20", "end_date": "2026-02-13", "notes": "", "status": "confirmed"}, {"id": 386, "name": "Nada Hany", "position": "junior", "start_date": "2026-01-04", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 387, "name": "Arafet Zouari", "position": "junior", "start_date": "2026-01-12", "end_date": "2026-03-12", "notes": "", "status": "confirmed"}, {"id": 388, "name": "Karim Khamis", "position": "Team Lead", "start_date": "2026-01-12", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 389, "name": "Salah Elshenawi", "position": "junior", "start_date": "2026-01-12", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 390, "name": "Hazem Ezz", "position": "Senior", "start_date": "2026-01-14", "end_date": "2026-04-04", "notes": "", "status": "confirmed"}, {"id": 391, "name": "Youssef Morsi", "position": "junior", "start_date": "2026-01-14", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 392, "name": "aya Saeed", "position": "Team Lead", "start_date": "2026-01-16", "end_date": "2026-04-15", "notes": "", "status": "confirmed"}, {"id": 393, "name": "Ahmed Emarah", "position": "Senior", "start_date": "2026-01-25", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 394, "name": "Ahmed Samir", "position": "Senior", "start_date": "2026-01-26", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 395, "name": "Ahmed Soliman", "position": "junior", "start_date": "2026-01-26", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 396, "name": "Ali Ezzat", "position": "Senior", "start_date": "2026-01-26", "end_date": "2026-02-19", "notes": "", "status": "confirmed"}, {"id": 397, "name": "Kareem Waleed", "position": "Senior", "start_date": "2026-01-26", "end_date": "2026-04-16", "notes": "", "status": "confirmed"}, {"id": 398, "name": "Menna Awad", "position": "junior", "start_date": "2026-01-26", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 399, "name": "Abeer Abdulhamed", "position": "junior", "start_date": "2026-01-29", "end_date": "2026-01-31", "notes": "", "status": "confirmed"}, {"id": 400, "name": "Ahmed Hamdy", "position": "Senior", "start_date": "2026-01-29", "end_date": "2026-03-19", "notes": "", "status": "confirmed"}, {"id": 401, "name": "Zein Gaber", "position": "Team Lead", "start_date": "2026-01-29", "end_date": "2026-03-17", "notes": "", "status": "confirmed"}, {"id": 402, "name": "Abeer Abdulhamed", "position": "Senior", "start_date": "2026-02-01", "end_date": "2026-04-16", "notes": "", "status": "confirmed"}, {"id": 403, "name": "Ahmed Emarah", "position": "Team Lead", "start_date": "2026-02-01", "end_date": "2026-03-16", "notes": "", "status": "confirmed"}, {"id": 404, "name": "Ahmed Soliman", "position": "Senior", "start_date": "2026-02-01", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 405, "name": "Aicha Chatti", "position": "junior", "start_date": "2026-02-01", "end_date": "2026-03-02", "notes": "", "status": "confirmed"}, {"id": 406, "name": "Menna Awad", "position": "Senior", "start_date": "2026-02-01", "end_date": "2026-04-02", "notes": "", "status": "confirmed"}, {"id": 407, "name": "Nada Hany", "position": "Senior", "start_date": "2026-02-01", "end_date": "2026-03-17", "notes": "", "status": "confirmed"}, {"id": 408, "name": "Salah Elshenawi", "position": "Senior", "start_date": "2026-02-01", "end_date": "2026-03-05", "notes": "", "status": "confirmed"}, {"id": 409, "name": "Youssef Morsi", "position": "Senior", "start_date": "2026-02-01", "end_date": "2026-03-19", "notes": "", "status": "confirmed"}, {"id": 410, "name": "Fady Ayman", "position": "Senior", "start_date": "2026-02-04", "end_date": "2026-04-17", "notes": "", "status": "confirmed"}, {"id": 411, "name": "Adham Ali", "position": "Senior", "start_date": "2026-02-10", "end_date": "2026-04-09", "notes": "", "status": "confirmed"}, {"id": 412, "name": "Ahmed Fawzy", "position": "junior", "start_date": "2026-02-10", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 413, "name": "Waleed Mohamed", "position": "junior", "start_date": "2026-02-10", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 414, "name": "Ahmed Elshehawy", "position": "junior", "start_date": "2026-02-12", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 415, "name": "Amr Emam", "position": "Senior", "start_date": "2026-02-12", "end_date": "2026-03-19", "notes": "", "status": "confirmed"}, {"id": 416, "name": "Aya Khattab", "position": "Team Lead", "start_date": "2026-02-12", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 417, "name": "Khaled Elbialy", "position": "Senior", "start_date": "2026-02-12", "end_date": "2026-04-10", "notes": "", "status": "confirmed"}, {"id": 418, "name": "Mohamed Tamer", "position": "junior", "start_date": "2026-02-12", "end_date": "2026-03-13", "notes": "", "status": "confirmed"}, {"id": 419, "name": "Eslam Hamada", "position": "Senior", "start_date": "2026-02-18", "end_date": "2026-03-18", "notes": "", "status": "confirmed"}, {"id": 420, "name": "Gasser Ashraf", "position": "Senior", "start_date": "2026-02-18", "end_date": "2026-03-13", "notes": "", "status": "confirmed"}, {"id": 421, "name": "Mohamed Hatem", "position": "Senior", "start_date": "2026-02-20", "end_date": "2026-04-02", "notes": "", "status": "confirmed"}, {"id": 422, "name": "Omar Elhady", "position": "Senior", "start_date": "2026-02-28", "end_date": "2026-04-09", "notes": "", "status": "confirmed"}, {"id": 423, "name": "Omar Sami", "position": "Senior", "start_date": "2026-02-28", "end_date": "2026-03-19", "notes": "", "status": "confirmed"}, {"id": 424, "name": "Sara Samir", "position": "Team Lead", "start_date": "2026-02-28", "end_date": "2026-04-16", "notes": "", "status": "confirmed"}, {"id": 425, "name": "Ahmed Mamdoh", "position": "Senior", "start_date": "2026-03-03", "end_date": "2026-04-02", "notes": "", "status": "confirmed"}, {"id": 426, "name": "Ahmed Emarah", "position": "Team Lead", "start_date": "2026-04-14", "end_date": "2026-05-31", "notes": "", "status": "confirmed"}, {"id": 427, "name": "Aya Saeed", "position": "Team Lead", "start_date": "2026-05-02", "end_date": "2026-05-31", "notes": "", "status": "confirmed"}, {"id": 428, "name": "Eslam Hamada", "position": "Senior", "start_date": "2026-05-02", "end_date": "2026-05-31", "notes": "", "status": "confirmed"}, {"id": 429, "name": "Hazem Emam", "position": "Senior", "start_date": "2026-05-02", "end_date": "2026-05-31", "notes": "", "status": "confirmed"}, {"id": 430, "name": "Moatasem Hatem", "position": "Team Lead", "start_date": "2026-05-03", "end_date": "2026-06-09", "notes": "", "status": "confirmed"}, {"id": 431, "name": "Ahmed Emarah", "position": "Team Lead", "start_date": "2026-06-08", "end_date": "2026-06-30", "notes": "", "status": "confirmed"}]

PROMO_SEED  = [{"id": 1, "name": "Mohamed Shafik", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2025, "notes": ""}, {"id": 2, "name": "Fady Ayman", "new_title": "Senior Software Engineer", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 3, "name": "Mohamed Mansour Desouky", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2025, "notes": ""}, {"id": 4, "name": "Mahmoud El Naggar", "new_title": "Software Engineer", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 5, "name": "Ereeny Essam", "new_title": "Lead Technology - React Native", "effective_date": "", "year": 2025, "notes": ""}, {"id": 6, "name": "Amr Emam", "new_title": "Senior Software Engineer", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 7, "name": "Ahmed Samir", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2025, "notes": ""}, {"id": 8, "name": "Sarah El Bahrawy", "new_title": "Senior Software Engineer", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 9, "name": "Ahmed Khaled", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2025, "notes": ""}, {"id": 10, "name": "El Hussein ElSayed", "new_title": "Lead Technology", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 11, "name": "Ahmed Mamdouh", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2025, "notes": ""}, {"id": 12, "name": "Mohamed Latif", "new_title": "Lead Quality Control", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 13, "name": "Salma Shammakh", "new_title": "Lead Quality Control", "effective_date": "", "year": 2025, "notes": ""}, {"id": 14, "name": "Omar Sami", "new_title": "Senior Quality Control Engineer", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 15, "name": "Anan Yehia", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2025, "notes": ""}, {"id": 16, "name": "Moamen Ibrahim", "new_title": "Senior Quality Control Engineer", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 17, "name": "Dina Said", "new_title": "Lead Business Analysis", "effective_date": "", "year": 2025, "notes": ""}, {"id": 18, "name": "Abdelrahman Refaie", "new_title": "Senior Business Analysis", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 19, "name": "Noha Saeed", "new_title": "Senior PMO", "effective_date": "", "year": 2025, "notes": ""}, {"id": 20, "name": "Farida Morsi", "new_title": "Senior Marketing and Social Media Specialist", "effective_date": "2025-02-27", "year": 2025, "notes": ""}, {"id": 21, "name": "Ahmed Emad", "new_title": "Senior Presales Consulant", "effective_date": "2025-03-11", "year": 2025, "notes": ""}, {"id": 22, "name": "Mohamed Fouad", "new_title": "Business Advisor", "effective_date": "2025-03-11", "year": 2025, "notes": ""}, {"id": 23, "name": "Nada Hany", "new_title": "Senior PMO", "effective_date": "2026-02-19", "year": 2026, "notes": ""}, {"id": 24, "name": "Esraa Ahmed", "new_title": "Senior PMO", "effective_date": "", "year": 2026, "notes": ""}, {"id": 25, "name": "Sultanah AlAnzi", "new_title": "Business Analyst", "effective_date": "2026-02-19", "year": 2026, "notes": ""}, {"id": 26, "name": "Amira AlQahtani", "new_title": "Business Analyst", "effective_date": "", "year": 2026, "notes": ""}, {"id": 27, "name": "Amani Almarzoog", "new_title": "Senior Business Analyst", "effective_date": "", "year": 2026, "notes": ""}, {"id": 28, "name": "Bushra Al Ghamdi", "new_title": "Senior Business Analyst", "effective_date": "", "year": 2026, "notes": ""}, {"id": 29, "name": "Abeer Abderlhamid", "new_title": "Senior Business Analyst", "effective_date": "", "year": 2026, "notes": ""}, {"id": 30, "name": "Aya Khattab", "new_title": "Lead Business Analyst", "effective_date": "", "year": 2026, "notes": ""}, {"id": 31, "name": "Adnan Albaba", "new_title": "Business Analyst , Manager", "effective_date": "", "year": 2026, "notes": ""}, {"id": 32, "name": "Ahmed Emad", "new_title": "Presales Manager", "effective_date": "", "year": 2026, "notes": ""}, {"id": 33, "name": "Ahmed Gouda", "new_title": "Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 34, "name": "Saif Mohamed", "new_title": "Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 35, "name": "Muntaha AlMuhareb", "new_title": "Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 36, "name": "Raya AlMuhaimed", "new_title": "Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 37, "name": "Nada AlDosari", "new_title": "Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 38, "name": "Reem Alaa", "new_title": "Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 39, "name": "Mahmoud Elnaggar", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 40, "name": "Omar Magdy", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 41, "name": "Nada Ramdan", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 42, "name": "Abdelazim Ashraf", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 43, "name": "Amira AbdelGawad", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 44, "name": "Essam Emad", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 45, "name": "Ahmed Ashraf", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 46, "name": "Nada Anies", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 47, "name": "Alaa Seif", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 48, "name": "Mohamed Tamer", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 49, "name": "Mahmoud Helmy", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 50, "name": "Ahmed Soliman", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 51, "name": "Heba Abdelmoneim", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 52, "name": "Menna Awad", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 53, "name": "Shatha Alrowisan", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 54, "name": "Salah El Shenawy", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 55, "name": "Njoud Alnajem", "new_title": "Senior Software Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 56, "name": "Ahmed Emara", "new_title": "Lead Technology", "effective_date": "", "year": 2026, "notes": ""}, {"id": 57, "name": "Abdulrahman Abduallah", "new_title": "Lead Technology", "effective_date": "", "year": 2026, "notes": ""}, {"id": 58, "name": "Sara Samir", "new_title": "Lead Technology", "effective_date": "", "year": 2026, "notes": ""}, {"id": 59, "name": "Ali Mohamed", "new_title": "Lead Technology", "effective_date": "", "year": 2026, "notes": ""}, {"id": 60, "name": "Ahmed El Shehawy", "new_title": "Senior NLP Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 61, "name": "Ahmed Kamel", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 62, "name": "Aly Hesham", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 63, "name": "Mohamed Youssef", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 64, "name": "Salma Alaa", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 65, "name": "Fatema Elnaggar", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 66, "name": "Mohamed Nagy", "new_title": "Senior Quality Control Engineer", "effective_date": "", "year": 2026, "notes": ""}, {"id": 67, "name": "Youssef Morsi", "new_title": "Security Expert", "effective_date": "", "year": 2026, "notes": ""}, {"id": 68, "name": "Abdelrahman Anwer", "new_title": "Lead Quality Control", "effective_date": "", "year": 2026, "notes": ""}, {"id": 69, "name": "Alaa El Kordi", "new_title": "Lead Quality Control", "effective_date": "", "year": 2026, "notes": ""}]


def _global_get(ns):
    """Read a global (non-project-prefixed) list from DB."""
    raw = db.get_override('global', ns, 'data')
    if isinstance(raw, list):
        return raw
    return None   # not yet seeded


def _global_set(ns, records):
    db.set_override('global', ns, 'data', records)


def _seed_if_empty(ns, seed):
    existing = _global_get(ns)
    if existing is None:
        _global_set(ns, seed)
        return True
    return False


def _next_id(records):
    if not records:
        return 1
    return max((r.get('id') or 0) for r in records) + 1


# ── Travel ───────────────────────────────────────────────────────────

@app.route('/api/global/travel', methods=['GET'])
def api_global_travel_get():
    _seed_if_empty('travel', TRAVEL_SEED)
    records = _global_get('travel') or []
    q = (request.args.get('q') or '').lower()
    linked_filter = request.args.get('linked')  # '1'=linked only, '0'=unlinked only
    if q:
        records = [r for r in records if q in (r.get('name') or '').lower()]
    if linked_filter == '1':
        records = [r for r in records if r.get('odoo_employee_id')]
    elif linked_filter == '0':
        records = [r for r in records if not r.get('odoo_employee_id')]
    return jsonify({'records': records, 'total': len(records)})


@app.route('/api/global/travel', methods=['POST'])
def api_global_travel_add():
    _seed_if_empty('travel', TRAVEL_SEED)
    records = _global_get('travel') or []
    body = request.json or {}
    name  = (body.get('name') or '').strip()
    start = (body.get('start_date') or '').strip()
    if not name or not start:
        return jsonify({'error': 'name and start_date required'}), 400
    name_lower = name.lower()
    for r in records:
        if r.get('name','').lower() != name_lower:
            continue
        r_start = r.get('start_date') or ''
        r_end   = r.get('end_date')   or '9999-12-31'
        new_end = (body.get('end_date') or '').strip() or '9999-12-31'
        # Overlap: two periods overlap if one starts before the other ends
        if r_start <= new_end and start <= r_end:
            return jsonify({'error': 'duplicate', 'existing': r,
                'message': f'{name} already has a travel record overlapping this period ({r_start} → {r.get("end_date") or "open"})'}), 409
    rec = {'id': _next_id(records), 'name': name,
        'position': (body.get('position') or '').strip(),
        'start_date': start, 'end_date': (body.get('end_date') or '').strip(),
        'notes': (body.get('notes') or '').strip(), 'status': body.get('status','confirmed'),
        'odoo_employee_id': body.get('odoo_employee_id'),
        'odoo_employee_code': body.get('odoo_employee_code'),
        'odoo_employee_name': body.get('odoo_employee_name')}
    records.append(rec); _global_set('travel', records)
    return jsonify({'ok': True, 'record': rec})


@app.route('/api/global/travel/<int:rec_id>', methods=['PUT'])
def api_global_travel_update(rec_id):
    records = _global_get('travel') or []
    body = request.json or {}

    # Find the record being updated
    target = next((r for r in records if r.get('id') == rec_id), None)
    if not target:
        return jsonify({'error': 'not found'}), 404

    # If linking to an Odoo employee, check for overlapping records with same Odoo ID
    new_odoo_id = body.get('odoo_employee_id')
    if new_odoo_id and not body.get('force'):
        t_start = target.get('start_date') or ''
        t_end   = target.get('end_date')   or '9999-12-31'
        conflicts = []
        for r in records:
            if r.get('id') == rec_id:
                continue
            # Match by Odoo ID
            if str(r.get('odoo_employee_id') or '') != str(new_odoo_id):
                continue
            r_start = r.get('start_date') or ''
            r_end   = r.get('end_date')   or '9999-12-31'
            # Overlap check
            if r_start <= t_end and t_start <= r_end:
                conflicts.append({
                    'id':         r.get('id'),
                    'name':       r.get('name'),
                    'position':   r.get('position',''),
                    'start_date': r.get('start_date',''),
                    'end_date':   r.get('end_date',''),
                })
        if conflicts:
            return jsonify({
                'ok': False,
                'conflict': True,
                'conflicts': conflicts,
                'message': f'This employee already has {len(conflicts)} record(s) overlapping this travel period.'
            }), 409

    # Apply update
    for i, r in enumerate(records):
        if r.get('id') == rec_id:
            records[i] = {**r, **{k: v for k, v in body.items() if k not in ('id', 'force')}}
            _global_set('travel', records)
            return jsonify({'ok': True, 'record': records[i]})
    return jsonify({'error': 'not found'}), 404


@app.route('/api/global/travel/<int:rec_id>', methods=['DELETE'])
def api_global_travel_delete(rec_id):
    records = _global_get('travel') or []
    new = [r for r in records if r.get('id') != rec_id]
    if len(new) == len(records): return jsonify({'error': 'not found'}), 404
    _global_set('travel', new)
    return jsonify({'ok': True})


@app.route('/api/global/travel/import', methods=['POST'])
def api_global_travel_import():
    """Bulk import travel records from JSON array. Skips duplicates."""
    _seed_if_empty('travel', TRAVEL_SEED)
    records = _global_get('travel') or []
    incoming = request.json or []
    if not isinstance(incoming, list):
        incoming = incoming.get('records', [])
    added = skipped = 0
    for row in incoming:
        name  = (row.get('name') or '').strip()
        start = (row.get('start_date') or row.get('From') or '').strip()
        if not name or not start: continue
        dup = any(r.get('name','').lower()==name.lower() and r.get('start_date')==start for r in records)
        if dup: skipped += 1; continue
        records.append({
            'id': _next_id(records),
            'name': name,
            'position': (row.get('position') or row.get('Level') or '').strip(),
            'start_date': start[:10],
            'end_date': str(row.get('end_date') or row.get('To') or '')[:10],
            'notes': (row.get('notes') or '').strip(),
            'status': row.get('status', 'confirmed'),
        })
        added += 1
    _global_set('travel', records)
    return jsonify({'ok': True, 'added': added, 'skipped': skipped})


# ── Promotions ───────────────────────────────────────────────────────

@app.route('/api/global/promotions', methods=['GET'])
def api_global_promos_get():
    _seed_if_empty('promotions', PROMO_SEED)
    records = _global_get('promotions') or []
    year = request.args.get('year')
    if year:
        records = [r for r in records if str(r.get('year','')) == year]
    q = (request.args.get('q') or '').lower()
    if q:
        records = [r for r in records if q in (r.get('name') or '').lower()]
    return jsonify({'records': records, 'total': len(records)})


@app.route('/api/global/promotions', methods=['POST'])
def api_global_promos_add():
    _seed_if_empty('promotions', PROMO_SEED)
    records = _global_get('promotions') or []
    body = request.json or {}
    name = (body.get('name') or '').strip()
    year = body.get('year') or ''
    if not name: return jsonify({'error': 'name required'}), 400
    for r in records:
        if r.get('name','').lower()==name.lower() and str(r.get('year',''))==str(year):
            return jsonify({'error': 'duplicate', 'existing': r}), 409
    rec = {'id': _next_id(records), 'name': name,
        'old_title':      (body.get('old_title')      or '').strip(),
        'new_title':      (body.get('new_title')      or '').strip(),
        'effective_date': (body.get('effective_date') or '').strip(),
        'year': int(year) if year else None, 'notes': (body.get('notes') or '').strip(),
        'odoo_employee_id':   body.get('odoo_employee_id'),
        'odoo_employee_code': body.get('odoo_employee_code')}
    records.append(rec); _global_set('promotions', records)
    return jsonify({'ok': True, 'record': rec})


@app.route('/api/global/promotions/<int:rec_id>', methods=['PUT'])
def api_global_promos_update(rec_id):
    records = _global_get('promotions') or []
    body = request.json or {}
    for i, r in enumerate(records):
        if r.get('id') == rec_id:
            records[i] = {**r, **{k: v for k, v in body.items() if k != 'id'}}
            _global_set('promotions', records)
            return jsonify({'ok': True, 'record': records[i]})
    return jsonify({'error': 'not found'}), 404


@app.route('/api/global/promotions/<int:rec_id>', methods=['DELETE'])
def api_global_promos_delete(rec_id):
    records = _global_get('promotions') or []
    new = [r for r in records if r.get('id') != rec_id]
    if len(new) == len(records): return jsonify({'error': 'not found'}), 404
    _global_set('promotions', new)
    return jsonify({'ok': True})


@app.route('/api/global/promotions/import', methods=['POST'])
def api_global_promos_import():
    """Bulk import promotions. Skips duplicates (same name+year)."""
    _seed_if_empty('promotions', PROMO_SEED)
    records = _global_get('promotions') or []
    incoming = request.json or []
    if not isinstance(incoming, list):
        incoming = incoming.get('records', [])
    added = skipped = 0
    for row in incoming:
        name = (row.get('name') or row.get('Name') or '').strip()
        year = row.get('year') or ''
        if not name: continue
        dup = any(r.get('name','').lower()==name.lower() and str(r.get('year',''))==str(year) for r in records)
        if dup: skipped += 1; continue
        records.append({
            'id': _next_id(records),
            'name':           name,
            'new_title':      (row.get('new_title') or row.get('Title') or row.get('Tittle') or '').strip(),
            'effective_date': (row.get('effective_date') or row.get('Promotion Date') or '').strip(),
            'year':           int(year) if year else None,
            'notes':          (row.get('notes') or '').strip(),
        })
        added += 1
    _global_set('promotions', records)
    return jsonify({'ok': True, 'added': added, 'skipped': skipped})


@app.route('/manage')
@login_required
def manage_page():
    """Global Travel & Promotions management page."""
    return render_template('partials/manage.html')



@app.route('/api/global/travel/parse-ticket', methods=['POST'])
def api_parse_ticket():
    import re, io
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    raw = f.read()
    text = ''
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or '') + chr(10)
    except Exception:
        try:
            import pypdf
            rdr = pypdf.PdfReader(io.BytesIO(raw))
            text = chr(10).join(pg.extract_text() or '' for pg in rdr.pages)
        except Exception as e2:
            return jsonify({'error': str(e2)}), 500

    result = {}

    # ── Passenger Name — handles multiple ticket formats ──
    _LABEL_WORDS = {'PASSENGER','PASSPORT','AIRLINE','AGENT','FLIGHT','CITY',
                    'ARRIVAL','DEPARTURE','LUGGAGE','FARE','TAXES','TOTAL',
                    'ATTENTION','RECEIPT','ITINERARY','TICKET','BOOKING','NAME'}

    def _extract_name(text):
        # 1. Saudia: "MR . Kareem Elsafy - ADT" / "MRS. Name/ADT"
        m = re.search(r'(?:MR|MRS|MS|DR)\s*\.?\s+([A-Za-z][A-Za-z ]{2,35}?)\s*[-/]\s*ADT', text, re.IGNORECASE)
        if m:
            return m.group(1).strip().title()
        # 2. "Passenger Name : Kareem Elsafy - ADT"
        m = re.search(r'(?:Passenger\s+)?Name\s*[:\-]\s*(?:MR|MRS|MS|DR)?\s*\.?\s*([A-Za-z][A-Za-z ]{2,35}?)\s*[-/]\s*ADT', text, re.IGNORECASE)
        if m:
            return m.group(1).strip().title()
        # 3. Tumodo: line after "PASSENGER, PASSPORT AIRLINE AGENT" header
        #    Next line: "KAREEM ELSAFY Saudia (SV) Tumodo" — take leading ALL-CAPS words
        lines = text.split('\n')
        for i, line in enumerate(lines):
            if re.search(r'PASSENGER.*AGENT', line, re.IGNORECASE) and i+1 < len(lines):
                next_line = lines[i+1]
                words = next_line.split()
                name_words = []
                for w in words:
                    wc = re.sub(r'[(),./]', '', w)
                    if re.match(r'^[A-Z]{2,}$', wc) and wc not in _LABEL_WORDS:
                        name_words.append(wc)
                    else:
                        break
                if len(name_words) >= 1:
                    return ' '.join(name_words).title()
        # 4. Tumodo variant: name on same line before airline "(SV)" or "(MS)" code
        m = re.search(r'\n([A-Z][A-Z ]{2,35}?)\s+(?:Saudia|EgyptAir|flydubai|Emirates|Qatar|Air\s+Arabia|Nile\s+Air|\w+\s*\([A-Z]{2}\))', text)
        if m:
            raw = m.group(1).strip()
            if len(raw) > 3 and raw.upper() not in _LABEL_WORDS:
                return raw.title()
        # 5. ALL-CAPS name before passport number (letter + 6+ digits)
        m = re.search(r'\n([A-Z][A-Z ]{3,35})\n[A-Z]\d{6,}', text)
        if m:
            raw = re.sub(r'^(?:' + '|'.join(_LABEL_WORDS) + r')[,\s]*', '', m.group(1).strip(), flags=re.IGNORECASE).strip()
            if raw and len(raw) > 3:
                return raw.title()
        # 6. Generic: "Passenger Name: KAREEM ELSAFY"
        m = re.search(r'(?:Passenger\s+)?Name\s*:\s*\n?\s*([A-Za-z][A-Za-z ]{2,35}?)(?:\n|$)', text, re.IGNORECASE)
        if m:
            return m.group(1).strip().title()
        return None

    _name = _extract_name(text)
    if _name:
        # Clean trailing noise (ADT/ADU/etc)
        _name = re.sub(r'\s*(Adt|Adu|Chd|Inf)\s*$', '', _name, flags=re.IGNORECASE).strip()
        if len(_name) > 3:
            result['name'] = _name

    # ── VALIDATION: reject if no passenger name found ──
    if not result.get('name'):
        return jsonify({
            'ok': False,
            'error': 'Could not extract passenger name from this ticket. Please make sure the PDF is a valid flight ticket with a passenger name, or add the record manually.'
        }), 422

    # ── IATA airport code → city name mapping ──
    IATA_MAP = {
        'RUH':'Riyadh','JED':'Jeddah','DMM':'Dammam','MED':'Medina','TIF':'Taif',
        'CAI':'Cairo','ALY':'Alexandria','HRG':'Hurghada','LXR':'Luxor','SSH':'Sharm El Sheikh',
        'DXB':'Dubai','AUH':'Abu Dhabi','DOH':'Doha','AMM':'Amman','BEY':'Beirut',
        'IST':'Istanbul','LHR':'London','FRA':'Frankfurt','CDG':'Paris',
    }
    EGYPT_IATA = {'CAI','ALY','HRG','LXR','SSH','HBE'}
    KSA_IATA   = {'RUH','JED','DMM','MED','TIF','AHB','GIZ','TUU'}
    EGYPT_CITIES = ['Cairo','Alexandria','Hurghada','Luxor','Sharm','Egypt']
    KSA_CITIES   = ['Riyadh','Jeddah','Dammam','Medina','Mecca','Khobar','Saudi Arabia','Saudi']
    ALL_CITIES   = EGYPT_CITIES + KSA_CITIES + ['Dubai','Doha','Amman','Beirut','Istanbul','London','Frankfurt','Paris','Abu Dhabi']

    # Expand IATA codes in text for city detection
    text_exp = text
    for code, city in IATA_MAP.items():
        text_exp = re.sub(r'\b' + code + r'\b', city, text_exp)

    # ── Dates — all formats ──
    mo_map = {'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
               'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'}
    dates_found = []

    # Extract flight date from route line first: "RUH — CAI, 16.04.2026" or "CAI-RUH 26Jan2026"
    flight_date = None
    route_date_m = re.search(
        r'\b(?:' + '|'.join(list(IATA_MAP.keys())) + r')\s*[—\-]+\s*(?:' + '|'.join(list(IATA_MAP.keys())) + r')[,\s]+(\d{2})\.(\d{2})\.(20\d{2})',
        text, re.IGNORECASE)
    if route_date_m:
        flight_date = f"{route_date_m.group(3)}-{route_date_m.group(2)}-{route_date_m.group(1)}"

    # DD.MM.YYYY
    for m in re.finditer(r'\b(\d{2})\.(\d{2})\.(20\d{2})\b', text):
        iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        if '2020' <= iso[:4] <= '2030' and iso not in dates_found:
            dates_found.append(iso)
    # DDMonYYYY / DD Mon YYYY
    for m in re.finditer(r'(\d{1,2})\s*-?\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*-?\s*(\d{4})', text, re.IGNORECASE):
        iso = '{}-{}-{:02d}'.format(m.group(3), mo_map[m.group(2).lower()], int(m.group(1)))
        if '2020' <= iso[:4] <= '2030' and iso not in dates_found:
            dates_found.append(iso)
    # YYYY-MM-DD
    for m in re.finditer(r'\b(20\d{2}-\d{2}-\d{2})\b', text):
        if m.group(1) not in dates_found: dates_found.append(m.group(1))
    # DD/MM/YYYY
    for m in re.finditer(r'\b(\d{2})/(\d{2})/(20\d{2})\b', text):
        iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        if iso not in dates_found: dates_found.append(iso)
    dates_found = sorted(set(dates_found))

    # ── Cities ──
    city_hits = []
    for city in ALL_CITIES:
        for m in re.finditer(r'\b' + re.escape(city) + r'\b', text_exp, re.IGNORECASE):
            city_hits.append((m.start(), city.title()))
    city_hits.sort(key=lambda x: x[0])
    seen_order = []
    for _, city in city_hits:
        if not seen_order or seen_order[-1].lower() != city.lower():
            seen_order.append(city)

    from_city = seen_order[0] if seen_order else None
    to_city   = seen_order[1] if len(seen_order) > 1 else None
    result['from_city'] = from_city or ''
    result['to_city']   = to_city   or ''

    # ── Direction ──
    def _is_egypt(c):
        if not c: return False
        return any(x.lower() in c.lower() for x in EGYPT_CITIES + ['egypt'])
    def _is_ksa(c):
        if not c: return False
        return any(x.lower() in c.lower() for x in KSA_CITIES + ['saudi'])

    is_outbound = _is_egypt(from_city) and not _is_egypt(to_city)
    is_return   = _is_egypt(to_city)   and not _is_egypt(from_city)

    # Fallback: check IATA route pattern in raw text "XXX — YYY"
    if not is_outbound and not is_return:
        route_m = re.search(r'\b([A-Z]{3})\s*[—\-]+\s*([A-Z]{3})\b', text)
        if route_m:
            orig, dest = route_m.group(1), route_m.group(2)
            is_outbound = orig in EGYPT_IATA and dest in KSA_IATA
            is_return   = orig in KSA_IATA   and dest in EGYPT_IATA
            if is_outbound: from_city, to_city = IATA_MAP.get(orig, orig), IATA_MAP.get(dest, dest)
            if is_return:   from_city, to_city = IATA_MAP.get(orig, orig), IATA_MAP.get(dest, dest)
            result['from_city'] = from_city or ''
            result['to_city']   = to_city   or ''

    result['ok'] = True
    if is_outbound:
        result['direction']  = 'outbound'
        result['start_date'] = flight_date or (dates_found[0] if dates_found else '')
        result['end_date']   = ''
    elif is_return:
        result['direction'] = 'return'
        result['end_date']  = flight_date or (dates_found[-1] if dates_found else '')
        result['start_date'] = ''
        # Find matching travel record — open-ended OR overlapping the return date
        if result.get('name'):
            travel_records = _global_get('travel') or []
            clean_name = result['name'].lower().strip()
            return_date = result.get('end_date') or ''

            def _nm(n1, n2):
                w1 = set(n1.lower().split())
                w2 = set(n2.lower().split())
                return w1 == w2 or w1.issubset(w2) or w2.issubset(w1)

            # Priority 1: open-ended record (no end_date)
            open_record = None
            for tr in travel_records:
                if _nm(clean_name, (tr.get('name') or '').lower().strip()) and not tr.get('end_date'):
                    open_record = tr; break

            # Priority 2: record where return_date falls within [start_date, end_date]
            overlap_record = None
            if not open_record and return_date:
                for tr in travel_records:
                    if not _nm(clean_name, (tr.get('name') or '').lower().strip()):
                        continue
                    tr_start = tr.get('start_date') or ''
                    tr_end   = tr.get('end_date')   or '9999-12-31'
                    if tr_start <= return_date <= tr_end:
                        overlap_record = tr; break

            matched = open_record or overlap_record
            if matched:
                existing_end = (matched.get('end_date') or '').strip()
                result['open_travel_id']    = matched['id']
                result['open_travel_start'] = matched.get('start_date', '')
                result['open_travel_end']   = existing_end
                if existing_end:
                    result['warning'] = f'This person already has a travel record ({matched.get("start_date","")} → {existing_end}). The return date is already set.'
            else:
                result['warning'] = 'No outbound travel record found for this person covering the return date. Please add the outbound ticket first.'
    else:
        result['direction']  = 'unknown'
        result['start_date'] = flight_date or (dates_found[0] if dates_found else '')
        result['end_date']   = ''

    # ── Odoo employee match + position ──
    if result.get('name') and odoo.uid:
        try:
            first = result['name'].split()[0]
            emps = odoo.models.execute_kw(ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'hr.employee', 'search_read',
                [[('name', 'ilike', first), ('active', '=', True)]],
                {'fields': ['id', 'name', 'job_title', 'job_id', 'number', 'barcode', 'identification_id'], 'limit': 5})
            result['odoo_matches'] = []
            for e in emps:
                job = e.get('job_id')
                title = e.get('job_title') or (job[1] if isinstance(job, list) and len(job) > 1 else '')
                # Get employee code
                import re as _re_t
                code = ''
                for field in ['number', 'barcode', 'identification_id']:
                    val = (e.get(field) or '').strip()
                    if val and _re_t.match(r'^[ERT]\d+$', val, _re_t.IGNORECASE):
                        code = val.upper(); break
                result['odoo_matches'].append({
                    'id': e['id'], 'name': e['name'],
                    'job': title, 'code': code
                })
            # Auto-fill position from best match (exact name match preferred)
            name_lower = result['name'].lower()
            best = None
            for e in result['odoo_matches']:
                if e['name'].lower() == name_lower:
                    best = e; break
            if not best and result['odoo_matches']:
                best = result['odoo_matches'][0]
            if best:
                result['odoo_id']   = best['id']
                result['odoo_code'] = best['code']
                # Map Odoo job_title to DB position key
                raw_job = best['job'] or ''
                # Try to find matching DB position (fuzzy: normalize Sr/Senior/Lead)
                import re as _re_pos2
                def _norm_pos(s):
                    s = _re_pos2.sub(r'\bsenior\b', 'Sr.', s, flags=_re_pos2.IGNORECASE)
                    s = _re_pos2.sub(r'\bsr\b(?!\.)', 'Sr.', s, flags=_re_pos2.IGNORECASE)
                    s = _re_pos2.sub(r'\blead\b', 'Lead', s, flags=_re_pos2.IGNORECASE)
                    return s.strip()
                all_db_pos = get_all_positions(db)
                matched_pos = None
                job_norm = _norm_pos(raw_job).lower()
                for p in all_db_pos:
                    pname = (p.get('position') or '').replace('EGY - ', '').replace('KSA - ', '')
                    if _norm_pos(pname).lower() == job_norm:
                        matched_pos = p.get('position')
                        break
                result['position'] = matched_pos or raw_job
        except Exception as _e:
            logger.warning('Odoo match: %s', _e)

    return jsonify(result)




@app.route('/api/global/employees/search')
def api_search_employees():
    """Search Odoo employees by name for linking to travel/promotions records."""
    try:
        if not odoo.uid: odoo.connect()
        q = (request.args.get('q') or '').strip()
        if not q or len(q) < 2:
            return jsonify({'employees': []})
        employees = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('name', 'ilike', q), ('active', '=', True)]],
            {'fields': ['id', 'name', 'job_title', 'job_id', 'barcode', 'identification_id', 'number'], 'limit': 20}
        )
        result = []
        for e in employees:
            job = e.get('job_id')
            # Try barcode first, then identification_id for the employee code
            import re as _re_emp
            def _get_code(e):
                for field in ['number', 'barcode', 'identification_id']:
                    val = (e.get(field) or '').strip()
                    if val and _re_emp.match(r'^[ERT]\d+$', val, _re_emp.IGNORECASE):
                        return val.upper()
                return ''
            code = _get_code(e)
            result.append({
                'id':    e['id'],
                'name':  e['name'],
                'code':  code,
                'title': e.get('job_title') or (job[1] if isinstance(job, list) and len(job)>1 else ''),
            })
        return jsonify({'employees': result})
    except Exception as ex:
        return jsonify({'employees': [], 'error': str(ex)})


@app.route('/api/global/travel/auto-link', methods=['POST'])
def api_travel_auto_link():
    """Auto-match travel records to Odoo employees with fuzzy matching."""
    try:
        if not odoo.uid: odoo.connect()
        records = _global_get('travel') or []
        import re as _re
        import unicodedata

        def norm(s):
            """Aggressive normalization: remove spaces, dashes, normalize Arabic transliterations"""
            s = (s or '').lower().strip()
            # Normalize common transliteration variants
            s = s.replace('abdel', 'abd').replace('abdal', 'abd').replace('abder', 'abd')
            s = s.replace('elh', 'elh').replace('el-', 'el').replace('al-', 'al')
            s = _re.sub(r"[\s\-_'.]+", '', s)
            return s

        def norm2(s):
            """Lighter norm: just lowercase + remove spaces"""
            return _re.sub(r'\s+', '', (s or '').lower().strip())

        def valid_code(s):
            s = (s or '').strip().upper()
            return s if (s and _re.match(r'^[ERT]\d+$', s)) else None

        all_emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('active', '=', True)]],
            {'fields': ['id', 'name', 'barcode', 'identification_id'], 'limit': 2000}
        )

        # Build multiple indexes for different matching strategies
        emp_by_norm  = {}   # aggressive norm
        emp_by_norm2 = {}   # light norm (just remove spaces)
        emp_by_fl    = {}   # first word + last word
        emp_by_code  = {}   # E/R/T code
        emp_list     = all_emps  # for fuzzy scan

        for e in all_emps:
            emp_by_norm[norm(e['name'])]   = e
            emp_by_norm2[norm2(e['name'])] = e
            words = e['name'].lower().split()
            if len(words) >= 2:
                emp_by_fl[words[0] + words[-1]] = e
            code = valid_code(e.get('barcode')) or valid_code(e.get('identification_id'))
            if code:
                emp_by_code[code] = e
                e['_code'] = code

        matched = already = unmatched = 0
        for r in records:
            found = None
            rname = r.get('name', '').strip()

            # 1. Code in record fields
            for field in ['name', 'position', 'notes']:
                m = _re.search(r'\b([ERT]\d+)\b', r.get(field, ''), _re.IGNORECASE)
                if m:
                    found = emp_by_code.get(m.group(1).upper())
                    if found: break

            # 2. Exact normalized name
            if not found:
                found = emp_by_norm.get(norm(rname))

            # 3. Light norm (just spaces removed)
            if not found:
                found = emp_by_norm2.get(norm2(rname))

            # 4. First + last word
            if not found:
                words = rname.lower().split()
                if len(words) >= 2:
                    found = emp_by_fl.get(words[0] + words[-1])

            # 5. Fuzzy: first word match + last word similar (handles spelling variants)
            if not found and len(rname.split()) >= 2:
                rwords = rname.lower().split()
                r_first = rwords[0]
                r_last  = rwords[-1]
                for e in emp_list:
                    ewords = e['name'].lower().split()
                    if len(ewords) < 2: continue
                    e_first = ewords[0]
                    e_last  = ewords[-1]
                    # First word must match exactly OR be very similar
                    if r_first != e_first:
                        continue
                    # Last word: check if one contains the other (handles Abdulhamed vs AbdelHamed)
                    rl = r_last.replace('el','').replace('ul','').replace('al','')
                    el = e_last.replace('el','').replace('ul','').replace('al','')
                    if rl == el or r_last in e_last or e_last in r_last:
                        found = e
                        break

            if found:
                new_code = found.get('_code') or None
                old_id = r.get('odoo_employee_id')
                r['odoo_employee_id']   = found['id']
                r['odoo_employee_name'] = found['name']
                r['odoo_employee_code'] = new_code
                if old_id == found['id']:
                    already += 1
                else:
                    matched += 1
            else:
                unmatched += 1

        _global_set('travel', records)
        return jsonify({'ok': True, 'matched': matched, 'already': already,
                        'unmatched': unmatched, 'total': len(records)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/global/promotions/auto-link', methods=['POST'])
def api_promotions_auto_link():
    """Auto-match promotion records to Odoo employees with fuzzy matching."""
    try:
        if not odoo.uid: odoo.connect()
        records = _global_get('promotions') or []
        import re as _re

        def norm(s):
            s = (s or '').lower().strip()
            s = s.replace('abdel', 'abd').replace('abdal', 'abd')
            s = _re.sub(r"[\s\-_'.]+", '', s)
            return s

        def valid_code(s):
            s = (s or '').strip().upper()
            return s if (s and _re.match(r'^[ERT]\d+$', s)) else None

        all_emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('active', '=', True)]],
            {'fields': ['id', 'name', 'barcode', 'identification_id'], 'limit': 2000}
        )
        emp_by_norm = {}
        emp_by_fl   = {}
        emp_list    = all_emps
        for e in all_emps:
            emp_by_norm[norm(e['name'])] = e
            words = e['name'].lower().split()
            if len(words) >= 2:
                emp_by_fl[words[0]+words[-1]] = e
            code = valid_code(e.get('barcode')) or valid_code(e.get('identification_id'))
            if code:
                e['_code'] = code

        matched = already = unmatched = 0
        for r in records:
            rname = r.get('name','').strip()
            found = emp_by_norm.get(norm(rname))
            if not found:
                words = rname.lower().split()
                if len(words) >= 2:
                    found = emp_by_fl.get(words[0]+words[-1])
            if not found and len(rname.split()) >= 2:
                rwords = rname.lower().split()
                for e in emp_list:
                    ewords = e['name'].lower().split()
                    if len(ewords) < 2: continue
                    if rwords[0] != ewords[0]: continue
                    rl = rwords[-1].replace('el','').replace('ul','').replace('al','')
                    el = ewords[-1].replace('el','').replace('ul','').replace('al','')
                    if rl == el or rwords[-1] in ewords[-1] or ewords[-1] in rwords[-1]:
                        found = e; break

            if found:
                old_id = r.get('odoo_employee_id')
                r['odoo_employee_id']   = found['id']
                r['odoo_employee_name'] = found['name']
                r['odoo_employee_code'] = found.get('_code') or None
                if old_id == found['id']: already += 1
                else: matched += 1
            else:
                unmatched += 1

        _global_set('promotions', records)
        return jsonify({'ok': True, 'matched': matched, 'already': already,
                        'unmatched': unmatched, 'total': len(records)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/debug/emp-fields')
def debug_emp_fields():
    """Check what fields Odoo employees have for code identification"""
    try:
        if not odoo.uid: odoo.connect()
        emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('name', 'in', ['Sara Samir', 'Rabab Hosney', 'Ahmed Helmi', 'Omar Mohamed'])]],
            {'fields': ['id', 'name', 'barcode', 'identification_id', 'employee_id', 'pin']}
        )
        return jsonify({'employees': emps})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/debug/autolink-test')
def debug_autolink_test():
    """Debug: show why auto-link fails - compare travel names vs Odoo names"""
    try:
        if not odoo.uid: odoo.connect()
        import re as _re

        def norm(s):
            return _re.sub(r"[\s\-_'.]+", '', (s or '').lower().strip())

        # Get Odoo employees
        all_emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('active', '=', True)]],
            {'fields': ['id', 'name', 'barcode'], 'limit': 2000}
        )
        emp_by_norm = {norm(e['name']): e['name'] for e in all_emps}
        emp_by_fl = {}
        for e in all_emps:
            words = e['name'].lower().split()
            if len(words) >= 2:
                emp_by_fl[words[0]+words[-1]] = e['name']

        # Get travel records
        records = _global_get('travel') or []

        results = []
        for r in records[:50]:  # first 50
            name = r.get('name','')
            n = norm(name)
            words = name.lower().split()
            fl_key = words[0]+words[-1] if len(words)>=2 else ''

            exact = emp_by_norm.get(n)
            fl    = emp_by_fl.get(fl_key)
            results.append({
                'travel_name': name,
                'norm':        n,
                'exact_match': exact,
                'fl_match':    fl,
                'linked':      bool(r.get('odoo_employee_id')),
            })

        # Show sample of Odoo names
        sample_odoo = [e['name'] for e in all_emps[:20]]

        return jsonify({
            'odoo_count': len(all_emps),
            'travel_count': len(records),
            'sample_odoo_names': sample_odoo,
            'travel_results': results,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/debug/effort-employee')
def debug_effort_employee():
    """Debug: show travel + promotions matching for a specific employee"""
    name = request.args.get('name', 'Abeer AbdelHamed')
    try:
        if not odoo.uid: odoo.connect()
        import re as _re

        # Get Odoo employee
        emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('name', 'ilike', name.split()[0])]],
            {'fields': ['id', 'name', 'barcode', 'job_title'], 'limit': 10}
        )

        # Load travel + promotions
        travel    = _global_get('travel') or []
        promos    = _global_get('promotions') or []

        def norm(s):
            return _re.sub(r'\s+', '', (s or '').lower())

        clean = _re.sub(r'\[[A-Z]\d+\]\s*', '', name).strip().lower()

        matched_travel = []
        for r in travel:
            if r.get('name','').lower().replace(' ','') == clean.replace(' ','') or                clean in r.get('name','').lower() or r.get('name','').lower() in clean:
                matched_travel.append(r)

        matched_promos = []
        for r in promos:
            rname = (r.get('name') or '').lower()
            emp_id = None
            for e in emps:
                if e['name'].lower() == name.lower():
                    emp_id = e['id']
            id_match = emp_id and r.get('odoo_employee_id') and int(r.get('odoo_employee_id',0)) == emp_id
            name_match = clean in rname or rname in clean
            if id_match or name_match:
                matched_promos.append(r)

        return jsonify({
            'searched_name': name,
            'odoo_employees': emps,
            'travel_matches': matched_travel,
            'promo_matches': matched_promos,
            'total_travel': len(travel),
            'total_promos': len(promos),
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/global/travel/fetch-codes', methods=['POST'])
def api_fetch_employee_codes():
    """Fetch employee codes ([E403] format) from Odoo timesheets and update travel/promo records."""
    try:
        if not odoo.uid: odoo.connect()

        import re as _re
        id_to_code = {}

        # Method 1: Get 'number' field directly from hr.employee (Job Number = R346, E259 etc)
        all_emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('active', '=', True)]],
            {'fields': ['id', 'name', 'number', 'barcode'], 'limit': 2000}
        )
        for e in all_emps:
            emp_id = e['id']
            # Try 'number' field first (Job Number like R346, E259)
            num = (e.get('number') or '').strip()
            if num and _re.match(r'^[ERT]\d+$', num, _re.IGNORECASE):
                id_to_code[emp_id] = num.upper()
            # Try barcode as fallback
            elif (e.get('barcode') or '').strip():
                bc = e['barcode'].strip()
                if _re.match(r'^[ERT]\d+$', bc, _re.IGNORECASE):
                    id_to_code[emp_id] = bc.upper()

        # Method 2: Get from timesheet names [E403] format for any missing
        timesheets = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.analytic.line', 'search_read',
            [[('project_id', '!=', False)]],
            {'fields': ['employee_id'], 'limit': 5000}
        )
        for ts in timesheets:
            emp = ts.get('employee_id')
            if not isinstance(emp, list) or len(emp) < 2: continue
            emp_id   = emp[0]
            emp_name = emp[1]
            if emp_id not in id_to_code:
                m = _re.match(r'^\[([ERT]\d+)\]', emp_name)
                if m:
                    id_to_code[emp_id] = m.group(1)

        # Update travel records
        travel = _global_get('travel') or []
        updated_travel = 0
        for r in travel:
            oid = r.get('odoo_employee_id')
            if oid and int(oid) in id_to_code:
                new_code = id_to_code[int(oid)]
                if r.get('odoo_employee_code') != new_code:
                    r['odoo_employee_code'] = new_code
                    updated_travel += 1

        # Update promotion records
        promos = _global_get('promotions') or []
        updated_promos = 0
        for r in promos:
            oid = r.get('odoo_employee_id')
            if oid and int(oid) in id_to_code:
                new_code = id_to_code[int(oid)]
                if r.get('odoo_employee_code') != new_code:
                    r['odoo_employee_code'] = new_code
                    updated_promos += 1

        _global_set('travel', travel)
        _global_set('promotions', promos)

        return jsonify({
            'ok': True,
            'codes_found': len(id_to_code),
            'travel_updated': updated_travel,
            'promos_updated': updated_promos,
            'sample_codes': dict(list(id_to_code.items())[:10])
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/global/travel/fetch-positions', methods=['POST'])
def api_fetch_employee_positions():
    """Fetch job positions from Odoo hr.employee (job_id field) and update travel + promo records."""
    try:
        if not odoo.uid: odoo.connect()

        # Get job_id (Job Position) + job_title for all active employees
        all_emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('active', '=', True)]],
            {'fields': ['id', 'name', 'job_id', 'job_title', 'number', 'barcode'], 'limit': 2000}
        )

        import re as _re_pos
        # Build map: odoo_id → position string
        id_to_position = {}
        id_to_code = {}
        for e in all_emps:
            eid = e['id']
            # Job position (hr.job name) — most reliable
            job = e.get('job_id')
            job_name = job[1] if isinstance(job, list) and len(job) > 1 else ''
            # Job title fallback
            job_title = (e.get('job_title') or '').strip()
            pos = job_name or job_title or ''
            if pos:
                id_to_position[eid] = pos
            # Also collect codes
            num = (e.get('number') or '').strip()
            if num and _re_pos.match(r'^[ERT]\d+$', num, _re_pos.IGNORECASE):
                id_to_code[eid] = num.upper()
            elif (e.get('barcode') or '').strip():
                bc = e['barcode'].strip()
                if _re_pos.match(r'^[ERT]\d+$', bc, _re_pos.IGNORECASE):
                    id_to_code[eid] = bc.upper()

        # Map Odoo job title → DB position key (EGY - Sr. Software Engineer etc)
        all_db_pos = get_all_positions(db)
        def _map_to_db_pos(raw_job, country='EGY'):
            if not raw_job: return raw_job
            def _norm(s):
                import re as _rn
                s = _rn.sub(r'senior', 'Sr.', s, flags=_rn.IGNORECASE)
                s = _rn.sub(r'sr(?!\.)', 'Sr.', s, flags=_rn.IGNORECASE)
                return s.strip().lower()
            job_norm = _norm(raw_job)
            for p in all_db_pos:
                pname = (p.get('position') or '').replace('EGY - ', '').replace('KSA - ', '').replace('TUN - ', '')
                if _norm(pname) == job_norm:
                    return p.get('position')
            return raw_job  # fallback: keep raw

        # Update travel records (only those linked to Odoo)
        travel = _global_get('travel') or []
        updated_travel = 0
        for r in travel:
            oid = r.get('odoo_employee_id')
            if not oid: continue
            oid_int = int(oid)
            changed = False
            # Update position
            if oid_int in id_to_position:
                raw = id_to_position[oid_int]
                # Try to detect country from existing record or code
                code = r.get('odoo_employee_code') or id_to_code.get(oid_int, '')
                country = 'KSA' if str(code).startswith('R') else 'EGY'
                new_pos = _map_to_db_pos(raw, country)
                if r.get('position') != new_pos:
                    r['position'] = new_pos; changed = True
            # Update code if missing
            if not r.get('odoo_employee_code') and oid_int in id_to_code:
                r['odoo_employee_code'] = id_to_code[oid_int]; changed = True
            if changed: updated_travel += 1

        _global_set('travel', travel)

        # Update promotion records
        promos = _global_get('promotions') or []
        updated_promos = 0
        for r in promos:
            oid = r.get('odoo_employee_id')
            if not oid: continue
            oid_int = int(oid)
            # Only update position if new_title is empty (don't override promo data)
            if oid_int in id_to_position and not r.get('new_title') and not r.get('new_position'):
                r['new_title'] = id_to_position[oid_int]
                updated_promos += 1
            if not r.get('odoo_employee_code') and oid_int in id_to_code:
                r['odoo_employee_code'] = id_to_code[oid_int]

        _global_set('promotions', promos)

        return jsonify({
            'ok': True,
            'positions_found': len(id_to_position),
            'travel_updated': updated_travel,
            'promos_updated': updated_promos,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/estimated-rows/import-summary-excel', methods=['POST'])
def api_import_summary_excel():
    """Read sheets from the summary Excel and return sheet names + optionally parse one sheet."""
    import io, re as _re_s
    f = request.files.get('file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file'}), 400
    raw = f.read()
    phase = (request.form.get('phase') or 'development').strip()
    sheet = (request.form.get('sheet') or '').strip()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    def _parse_num(v):
        if v is None: return None
        if isinstance(v, (int, float)): return float(v)
        s = str(v).replace(',','').replace('SAR','').replace('USD','').strip()
        try: return float(s)
        except: return None

    def _read_summary(ws):
        result = {}
        labels = {
            'total_usd_per_month': ['TOTAL USD per month'],
            'total_sar_per_month': ['TOTAL Estimated Cost per month'],
            'est_mds_per_month':   ['Estimated MDs / Month'],
            'cost_per_md_sar':     ['TOTAL Estimated Cost per MD'],
            'total_cost_sar':      ['Total project estimated Cost'],
            'total_mds':           ['Total No of MDs'],
            'total_cost_usd':      ['Total cost USD'],
        }
        for row in ws.iter_rows(values_only=True):
            label = str(row[0] or '').strip()
            if not label: continue
            for key, keywords in labels.items():
                if any(kw.lower() in label.lower() for kw in keywords):
                    for v in row[1:]:
                        n = _parse_num(v)
                        if n is not None:
                            result[key] = n; break
                    break
        return result

    # If sheet specified → parse and save
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
        summary = _read_summary(ws)
        if not summary:
            return jsonify({'ok': False, 'error': f'No summary data found in sheet "{sheet}"'}), 400
        # Save to DB per project+phase
        proj_id = session.get('project_id') or ''
        db.set_override('estimated_summary', str(proj_id), phase, summary)
        return jsonify({'ok': True, 'phase': phase, 'sheet': sheet, 'summary': summary})

    # No sheet specified → return list of sheets
    return jsonify({'ok': True, 'sheets': wb.sheetnames, 'needs_sheet_selection': True})


@app.route('/api/overview/phase-progress')
def api_overview_phase_progress():
    """Return latest % completion + remaining MDs per phase from plan_overrides."""
    _proj_id = session.get('project_id')
    _is_bog  = not _proj_id or str(_proj_id) == '228'
    phases = ['development', 'consultation', 'support'] if _is_bog else ['services', 'support']
    result = {}
    for phase in phases:
        try:
            overrides = load_plan_overrides().get('plan_overrides', {}).get(phase, {})
            if not overrides:
                result[phase] = {}
                continue
            # Find last month with completion or remaining entered
            latest = {}
            latest_month = ''
            for month_key, fields in sorted(overrides.items()):
                pct = fields.get('completion')
                rem = fields.get('remaining')
                if pct or rem:
                    latest = fields
                    latest_month = month_key
            result[phase] = {
                'completion': float(latest.get('completion', 0) or 0),
                'remaining':  float(latest.get('remaining',  0) or 0),
                'month_key':  latest_month,
            }
        except Exception as e:
            result[phase] = {}
    return jsonify({'ok': True, 'phases': result})


@app.route('/api/project-phases-available')
def api_project_phases_available():
    """Check which phase groups exist for this project — uses Excel config first."""
    proj_id = session.get('project_id')
    is_bog  = not proj_id or str(proj_id) == '228'
    if is_bog:
        return jsonify({'is_bog': True, 'has_services': True, 'has_support': True})

    _proj_name = active_project_name()

    # Check saved Excel config first
    configs = db.get_override('project_config', 'global', 'phase_configs') or {}
    def _match_cfg(pn, cfgs):
        pl = pn.lower().strip()
        for k, v in cfgs.items():
            kl = k.lower().strip()
            if kl == pl or kl in pl or pl in kl:
                return v
        return None

    cfg = _match_cfg(_proj_name, configs)
    if cfg:
        return jsonify({
            'is_bog':       False,
            'has_services': cfg.get('has_services', True),
            'has_support':  cfg.get('has_support', False),
            'sheets':       cfg.get('sheets', {}),
            'source':       'excel_config',
        })

    # Fallback: check Odoo project.phase records directly
    try:
        pid = get_project_odoo_id(_proj_name)
        if not pid:
            return jsonify({'has_services': True, 'has_support': False})
        # Get all project.phase for this project
        all_phases = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('project_id', '=', pid)]],
            {'fields': ['id', 'name'], 'limit': 100}
        )
        support_phases = [p for p in all_phases
                          if any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
        service_phases = [p for p in all_phases
                          if not any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
        return jsonify({
            'is_bog':        False,
            'has_services':  len(service_phases) > 0 or len(all_phases) == 0,
            'has_support':   len(support_phases) > 0,
            'source':        'odoo_phases',
            'all_phases':    [p['name'] for p in all_phases],
        })
    except Exception as e:
        return jsonify({'has_services': True, 'has_support': False, 'error': str(e)})


@app.route('/api/estimated-rows/import-excel', methods=['POST'])
def api_estimated_rows_import_excel():
    """Import Estimated Cost rows from Excel sheet.
    Reads columns: Position | Hour Rate | Actual Time (MH) | Cost per month | Estimated # of Months
    Hour rate is auto-filled from DB positions catalog if not in Excel.
    """
    import io
    f = request.files.get('file')
    phase = (request.form.get('phase') or 'development').strip()
    if not f:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Cannot read Excel: {e}'}), 400

    # Find the right sheet — prefer phase-specific sheets
    phase_sheet_map = {
        'development':  ['Estimated Cost - Development', 'Estimated Cost - Dev', 'Development'],
        'consultation': ['Estimated Cost - Consultation', 'Estimated Cost - Con', 'Consultation'],
        'support':      ['Estimated Cost - Support', 'Support'],
    }
    preferred = phase_sheet_map.get(phase, [])
    ws = None
    for name in preferred:
        if name in wb.sheetnames:
            ws = wb[name]; break
    if ws is None:
        # Try any sheet with "Estimated Cost" in name
        for sname in wb.sheetnames:
            if 'estimated cost' in sname.lower():
                ws = wb[sname]; break
    if ws is None:
        ws = wb.active  # fallback: first sheet

    # Parse rows — find header row first
    all_pos = get_all_positions(db)
    def _find_rate(pos_name):
        if not pos_name: return None
        import re as _re_r
        def _norm(s):
            s = _re_r.sub(r'senior', 'Sr.', s, flags=_re_r.IGNORECASE)
            s = _re_r.sub(r'sr(?!\.)', 'Sr.', s, flags=_re_r.IGNORECASE)
            s = _re_r.sub(r'\s*-\s*', ' - ', s)
            return s.strip().lower()
        target = _norm(pos_name)
        # Exact match
        for p in all_pos:
            if _norm(p.get('position','')) == target and p.get('hour_rate'):
                return float(p['hour_rate'])
        # Partial: strip country prefix and try
        stripped = _re_r.sub(r'^(EGY|KSA|TUN)\s*-\s*', '', pos_name, flags=_re_r.IGNORECASE).strip()
        stripped_norm = _norm(stripped)
        for p in all_pos:
            pstripped = _re_r.sub(r'^(EGY|KSA|TUN)\s*-\s*', '', p.get('position',''), flags=_re_r.IGNORECASE).strip()
            if _norm(pstripped) == stripped_norm and p.get('hour_rate'):
                return float(p['hour_rate'])
        return None

    # Find header row (contains "Position" or "Hour Rate")
    header_row = None
    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v or '').strip().lower() for v in row]
        if any('position' in v for v in vals) and any('hour' in v or 'actual' in v or 'month' in v for v in vals):
            header_row = [str(v or '').strip() for v in row]
            header_idx = i
            break

    if not header_row:
        return jsonify({'ok': False, 'error': 'Could not find header row with Position/Hour Rate columns'}), 400

    # Map column indices
    def _col(keywords):
        for kw in keywords:
            for j, h in enumerate(header_row):
                if kw.lower() in h.lower():
                    return j
        return None

    col_pos    = _col(['Position', 'position'])
    col_rate   = _col(['Hour Rate', 'Rate'])
    col_time   = _col(['Actual Time', 'MH', 'Hours/month'])
    col_months = _col(['Estimated #', 'No of Month', '# of Month', 'Months'])

    if col_pos is None:
        return jsonify({'ok': False, 'error': 'Position column not found'}), 400

    rows = []
    import time
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        pos = str(row[col_pos] if col_pos is not None and col_pos < len(row) and row[col_pos] else '').strip()
        if not pos or pos.lower() in ('total', 'position', 'none', ''):
            continue
        # Skip summary rows
        if pos.upper() == pos and len(pos) > 30:
            continue

        hr_val = row[col_rate] if col_rate is not None and col_rate < len(row) else None
        at_val = row[col_time] if col_time is not None and col_time < len(row) else 176
        em_val = row[col_months] if col_months is not None and col_months < len(row) else None

        try: hr = float(hr_val) if hr_val else None
        except: hr = None
        try: at = float(at_val) if at_val else 176
        except: at = 176
        try: em = float(em_val) if em_val else None
        except: em = None

        # Auto-fill hour rate from DB if not in Excel
        if not hr:
            hr = _find_rate(pos)

        rows.append({
            'id':         time.time() + len(rows),
            'position':   pos,
            'hourRate':   hr or '',
            'actualTime': at,
            'estMonths':  em or '',
        })

    if not rows:
        return jsonify({'ok': False, 'error': 'No data rows found in sheet'}), 400

    # Save to DB
    proj_id = session.get('project_id') or ''
    db.set_override('estimated_rows', str(proj_id), phase, rows)

    return jsonify({'ok': True, 'phase': phase, 'rows': rows, 'count': len(rows),
                    'sheet_used': ws.title})


@app.route('/api/project-config/upload-excel', methods=['POST'])
def api_upload_project_config_excel():
    """Upload the Estimated Cost Excel to auto-configure project phase configs.
    Reads sheet names to determine which projects have Services/Support.
    Saves config to DB per project name match.
    """
    import io, re as _re
    f = request.files.get('file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    def _is_support(name):
        return bool(_re.search(r'support|suppo\b', name, _re.IGNORECASE))

    def _is_service(name):
        return bool(_re.search(r'service|serv\b|\bCR\b', name, _re.IGNORECASE))

    def _proj_key(name):
        n = name.strip()
        n = _re.sub(r'[\s_]*(support|suppo|service|serv|CR)\s*$', '', n, flags=_re.IGNORECASE).strip()
        return n

    projects = {}
    for sheet in wb.sheetnames:
        is_sup = _is_support(sheet)
        is_svc = _is_service(sheet)
        if not is_sup and not is_svc:
            is_svc = True  # standalone = services only
        key = _proj_key(sheet)
        if key not in projects:
            projects[key] = {'has_services': False, 'has_support': False, 'sheets': {}}
        if is_sup:
            projects[key]['has_support'] = True
            projects[key]['sheets']['support'] = sheet.strip()
        if is_svc:
            projects[key]['has_services'] = True
            projects[key]['sheets']['services'] = sheet.strip()

    # Save global config
    db.set_override('project_config', 'global', 'phase_configs', projects)

    return jsonify({'ok': True, 'projects_found': len(projects),
                    'projects': projects})


@app.route('/api/project-config/all', methods=['GET'])
def api_get_all_project_configs():
    """Get all saved project configs."""
    configs = db.get_override('project_config', 'global', 'phase_configs') or {}
    return jsonify({'ok': True, 'configs': configs})


@app.route('/api/project-config/comment', methods=['GET', 'POST'])
def api_project_config_comment():
    """Get or save a comment for a project+phase."""
    if request.method == 'GET':
        proj  = request.args.get('project', '')
        phase = request.args.get('phase', '')
        saved = db.get_override('project_config_comments', proj, phase) or {}
        comment = saved.get('text', '') if isinstance(saved, dict) else str(saved or '')
        return jsonify({'ok': True, 'comment': comment})
    else:
        body  = request.json or {}
        proj  = body.get('project', '')
        phase = body.get('phase', '')
        val   = body.get('comment', '')
        db.set_override('project_config_comments', proj, phase, {'text': val})
        return jsonify({'ok': True})


@app.route('/api/estimated-summary', methods=['GET'])
def api_get_estimated_summary_v2():
    """Get saved estimated cost summary — supports optional project_name param."""
    phase = request.args.get('phase', 'development')
    proj_name = request.args.get('project_name', '')
    if proj_name:
        # Look up project_id by name match
        configs = db.get_override('project_config', 'global', 'phase_configs') or {}
        matched_id = None
        for k in configs:
            if k.lower().strip() in proj_name.lower() or proj_name.lower() in k.lower():
                matched_id = k
                break
        if matched_id:
            summary = db.get_override('estimated_summary', matched_id, phase)
            return jsonify({'ok': True, 'summary': summary or {}})
    proj_id = session.get('project_id') or ''
    summary = db.get_override('estimated_summary', str(proj_id), phase)
    return jsonify({'ok': True, 'summary': summary or {}})


@app.route('/api/project-config', methods=['GET'])
def api_get_project_config():
    """Get phase config for current project from saved Excel config."""
    _proj_name = active_project_name()
    configs = db.get_override('project_config', 'global', 'phase_configs') or {}

    # Find best match for current project name
    def _match(proj_name, configs):
        pn = proj_name.lower().strip()
        # Exact
        for k, v in configs.items():
            if k.lower().strip() == pn:
                return v
        # Substring both ways
        for k, v in configs.items():
            kl = k.lower().strip()
            if kl in pn or pn in kl:
                return v
        return None

    cfg = _match(_proj_name, configs)
    if cfg:
        return jsonify({'ok': True, 'project': _proj_name, 'config': cfg,
                        'has_services': cfg.get('has_services', True),
                        'has_support': cfg.get('has_support', False)})
    return jsonify({'ok': True, 'project': _proj_name, 'config': None,
                    'has_services': True, 'has_support': False})


@app.route('/api/effort/<phase_key>/unassigned-hours')
def api_effort_unassigned_hours(phase_key):
    """Find timesheets logged on this project but NOT in any known phase.
    Returns total hours, employee breakdown, and sample tasks.
    """
    try:
        if not odoo.uid: odoo.connect()
        _proj_name = active_project_name()
        _proj_id_ua = session.get('project_id')
        _is_bog_ua  = not _proj_id_ua or str(_proj_id_ua) == '228'

        # Find project
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', _proj_name)]],
            {'fields': ['id', 'name'], 'limit': 3}
        )
        if not projects:
            return jsonify({'ok': True, 'total_hours': 0, 'employees': [], 'tasks': []})
        project_id = projects[0]['id']

        # Get ALL tasks in project
        all_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('project_id', '=', project_id)]],
            {'fields': ['id', 'name', 'phase_id', 'stage_id'], 'limit': 10000}
        )

        # Get tasks that ARE in known phases (already counted in effort tab)
        phase_task_ids = set()
        if _is_bog_ua:
            phase_names = get_phase_mapping().get(phase_key, [])
            if phase_names:
                phases = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'project.phase', 'search_read',
                    [[('name', 'in', phase_names), ('project_id', '=', project_id)]],
                    {'fields': ['id'], 'limit': 50}
                )
                phase_ids = [p['id'] for p in phases]
                for t in all_tasks:
                    ph = t.get('phase_id')
                    if ph and isinstance(ph, list) and ph[0] in phase_ids:
                        phase_task_ids.add(t['id'])
        else:
            # Non-BOG: find tasks in ALL known phases
            all_proj_phases = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.phase', 'search_read',
                [[('project_id', '=', project_id)]],
                {'fields': ['id', 'name'], 'limit': 100}
            )
            all_phase_ids = [p['id'] for p in all_proj_phases]
            for t in all_tasks:
                ph = t.get('phase_id')
                if ph and isinstance(ph, list) and ph[0] in all_phase_ids:
                    phase_task_ids.add(t['id'])

        # Unassigned = tasks with NO phase_id
        unassigned_task_ids = [t['id'] for t in all_tasks if not t.get('phase_id')]

        if not unassigned_task_ids:
            return jsonify({'ok': True, 'total_hours': 0, 'employees': [], 'tasks': [],
                            'message': 'All hours are assigned to phases'})

        # Get timesheets for unassigned tasks
        timesheets = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.analytic.line', 'search_read',
            [[('task_id', 'in', unassigned_task_ids)]],
            {'fields': ['id', 'employee_id', 'unit_amount', 'date', 'task_id', 'name'],
             'limit': 5000}
        )

        total_hours = sum(t.get('unit_amount', 0) for t in timesheets)

        # Group by employee
        emp_hours = {}
        for ts in timesheets:
            emp = ts.get('employee_id')
            if emp and isinstance(emp, list):
                eid, ename = emp[0], emp[1]
                if eid not in emp_hours:
                    emp_hours[eid] = {'name': ename, 'hours': 0}
                emp_hours[eid]['hours'] += ts.get('unit_amount', 0)

        # Sample tasks
        task_map = {t['id']: t['name'] for t in all_tasks}
        task_hours = {}
        for ts in timesheets:
            tid = ts.get('task_id')
            if tid and isinstance(tid, list):
                tid0 = tid[0]
                if tid0 not in task_hours:
                    task_hours[tid0] = {'name': task_map.get(tid0, '?'), 'hours': 0}
                task_hours[tid0]['hours'] += ts.get('unit_amount', 0)

        return jsonify({
            'ok': True,
            'total_hours': round(total_hours, 1),
            'total_mds': round(total_hours / 8, 1),
            'employees': sorted(emp_hours.values(), key=lambda x: -x['hours'])[:20],
            'tasks': sorted(task_hours.values(), key=lambda x: -x['hours'])[:10],
        })
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/debug/phases-for-project')
def debug_phases_for_project():
    """Debug: show all phases and task.types for current project."""
    try:
        if not odoo.uid: odoo.connect()
        _proj_name = active_project_name()
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', _proj_name)]],
            {'fields': ['id', 'name'], 'limit': 5}
        )
        if not projects:
            return jsonify({'error': f'Project not found: {_proj_name}'})
        pid = projects[0]['id']

        # Try project.phase with different domain options
        phases_by_project = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('project_id', '=', pid)]],
            {'fields': ['id', 'name', 'project_id'], 'limit': 50}
        )
        phases_by_project_in = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('project_id', 'in', [pid])]],
            {'fields': ['id', 'name', 'project_id'], 'limit': 50}
        )
        # Get all phases (no filter) to see what exists
        all_phases_sample = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[]],
            {'fields': ['id', 'name', 'project_id'], 'limit': 20}
        )
        # Task types (stages)
        task_types = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task.type', 'search_read',
            [[('project_ids', 'in', [pid])]],
            {'fields': ['id', 'name'], 'limit': 50}
        )
        # Sample tasks with phase_id
        tasks_sample = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('project_id', '=', pid)]],
            {'fields': ['id', 'name', 'phase_id', 'stage_id'], 'limit': 10}
        )
        return jsonify({
            'project': projects[0],
            'phases_by_project_eq': phases_by_project,
            'phases_by_project_in': phases_by_project_in,
            'all_phases_sample': all_phases_sample,
            'task_types_stages': task_types,
            'tasks_sample': tasks_sample,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()})


@app.route('/debug/projects-raw')
def debug_projects_raw():
    """Debug: show raw Odoo project data to verify fields."""
    try:
        if not odoo.uid: odoo.connect()
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read', [[]],
            {'fields': ['id', 'name', 'stage_id', 'active', 'user_id', 'date_start', 'end_date'],
             'limit': 20, 'order': 'name asc'}
        )
        return jsonify({'count': len(projects), 'projects': projects})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/projects/list')
def api_projects_list():
    """Fetch all projects from Odoo with stage, PM, dates, value."""
    try:
        if not odoo.uid:
            if not odoo.connect():
                return jsonify({'error': 'Odoo not connected', 'projects': []}), 503
        # Fetch projects — use minimal safe fields first, then try optional ones
        safe_fields = ['id', 'name', 'user_id', 'date_start', 'end_date', 'stage_id']
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[]],   # all projects (active filter may differ by Odoo version)
            {'fields': safe_fields, 'limit': 300, 'order': 'name asc'}
        )

        # Try to get extra fields separately (optional — may not exist in v14)
        extra_map = {}
        try:
            extras = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[]],
                {'fields': ['id', 'value', 'coordinator_id', 'last_update_status'], 'limit': 300}
            )
            extra_map = {e['id']: e for e in extras}
        except Exception as _ex:
            logger.warning(f"Optional project fields unavailable: {_ex}")

        result = []
        for p in projects:
            pm    = p.get('user_id')
            ex    = extra_map.get(p['id'], {})
            coord = ex.get('coordinator_id') or []
            stage = p.get('stage_id')
            stage_name = (stage[1] if isinstance(stage, list) and len(stage) > 1 else '') or ''

            # Map stage name to status color
            sn_lower = stage_name.lower()
            if any(k in sn_lower for k in ['progress', 'active', 'ongoing', 'in progress']):
                status = 'on_track'
            elif any(k in sn_lower for k in ['closing', 'receive', 'confirmation', 'release']):
                status = 'at_risk'
            elif any(k in sn_lower for k in ['completed', 'closed', 'done', 'finish']):
                status = 'completed'
            elif any(k in sn_lower for k in ['draft', 'new', 'pending']):
                status = 'draft'
            else:
                status = ex.get('last_update_status') or 'on_track'

            result.append({
                'id':          p['id'],
                'name':        p['name'],
                'pm':          pm[1] if isinstance(pm, list) and len(pm) > 1 else '',
                'coordinator': coord[1] if isinstance(coord, list) and len(coord) > 1 else '',
                'date_start':  (p.get('date_start') or '')[:10],
                'end_date':    (p.get('end_date')   or '')[:10],
                'status':      status,
                'stage_name':  stage_name,
                'value':       float(ex.get('value') or 0),
            })
        return jsonify({'projects': result})
    except Exception as e:
        logger.error(f"Projects list error: {e}", exc_info=True)
        return jsonify({'error': str(e), 'projects': []}), 500

@app.route('/api/overview')
def api_overview():
    """KPIs from Roadmap + Odoo project details"""
    _proj_name = active_project_name()
    _proj_id   = session.get('project_id')
    _is_bog    = not _proj_id or str(_proj_id) == '228'

    # Roadmap data only for BOG — other projects get empty roadmap
    roadmap_services = (ROADMAP or []) if _is_bog else []
    total_wd_roadmap = sum(s.get('wd') or 0 for s in roadmap_services)

    teams = set()
    for s in roadmap_services:
        if s.get('team'):
            teams.add(s['team'])

    milestones_count = len(MILESTONES) if (_is_bog and MILESTONES) else 0

    # ── Fetch live project data from Odoo ──────────────────────────────
    odoo_project = {}
    try:
        if not odoo.uid:
            odoo.connect()
        if odoo.uid:
            projects = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[('name', 'ilike', _proj_name)]],
                {'fields': ['id', 'name', 'start_date', 'end_date',
                            'user_id',        # Project Manager (many2one res.users)
                            'coordinator_id', # Coordinator (many2one res.users)
                            'value',          # Project value (float)
                            ], 'limit': 3}
            )
            if projects:
                # Prefer exact match
                proj = next((p for p in projects if (p.get('name') or '').strip().lower() == (_proj_name or '').strip().lower()), projects[0])
                odoo_project = proj
    except Exception as _e:
        logger.warning(f"Odoo project fetch for overview failed: {_e}")

    # Project start/end: Odoo wins, fallback to roadmap_data
    def _odoo_date(field):
        v = odoo_project.get(field)
        if isinstance(v, str) and len(v) >= 10:
            return v[:10]
        return None

    proj_start = _odoo_date('start_date') or PROJECT_INFO.get('start_date')
    proj_end   = _odoo_date('end_date')   or PROJECT_INFO.get('end_date')

    # Duration in months
    duration_months = PROJECT_INFO.get('duration_months')
    try:
        if proj_start and proj_end:
            ps = datetime.strptime(proj_start, '%Y-%m-%d').date()
            pe = datetime.strptime(proj_end,   '%Y-%m-%d').date()
            duration_months = round((pe - ps).days / 30.44, 0)
    except Exception:
        pass

    # Project Manager & Coordinator from Odoo
    pm_name = ''
    coord_name = ''
    try:
        uid_field = odoo_project.get('user_id')
        if isinstance(uid_field, list) and len(uid_field) > 1:
            pm_name = uid_field[1]
        # Coordinator from coordinator_id (many2one)
        cf = odoo_project.get('coordinator_id')
        if isinstance(cf, list) and len(cf) > 1:
            coord_name = cf[1]
    except Exception:
        pass

    # Project value from Odoo — field name is 'value' (float)
    project_value_sar = 0
    v = odoo_project.get('value')
    if v:
        try:
            project_value_sar = float(v)
        except (TypeError, ValueError):
            pass

    # Timeline
    time_progress_pct = days_elapsed = days_remaining = 0
    try:
        if proj_start and proj_end:
            ps = datetime.strptime(proj_start, '%Y-%m-%d').date()
            pe = datetime.strptime(proj_end,   '%Y-%m-%d').date()
            today = date.today()
            total = (pe - ps).days
            elapsed = max(0, (today - ps).days)
            days_elapsed = elapsed
            days_remaining = max(0, (pe - today).days)
            if total > 0:
                time_progress_pct = round(min(100, elapsed / total * 100), 1)
    except Exception:
        pass

    # Project Progress + Remaining MDs from Variance overrides
    project_progress = project_remaining = 0
    dev_eac = con_eac = 0  # EAC per phase from plan overrides
    try:
        overrides = load_plan_overrides()
        plan_overrides = overrides.get('plan_overrides', {}) or {}
        for phase_key in ['development', 'consultation', 'support']:
            phase_data = plan_overrides.get(phase_key, {}) or {}
            if not phase_data:
                continue
            sorted_months = sorted(phase_data.keys(), reverse=True)
            ph_comp = ph_rem = 0
            for mk in sorted_months:
                md = phase_data.get(mk, {}) or {}
                if 'completion' in md and not ph_comp:
                    ph_comp = float(md['completion'])
                if 'remaining' in md and not ph_rem:
                    ph_rem = float(md['remaining'])
                if ph_comp and ph_rem:
                    break
            if phase_key == 'development':
                project_progress  = ph_comp
                project_remaining = ph_rem
                dev_eac = ph_rem  # simplified: remaining MDs = EAC delta
            elif phase_key == 'consultation':
                con_eac = ph_rem
    except Exception as e:
        logger.warning(f"Could not read variance overrides: {e}")

    # Current Cost per phase — from effort API cached data (best effort)
    # We don't re-fetch here to keep this fast; JS will fill in from AppState
    # Get consultation progress separately
    con_progress = con_remaining = 0.0
    try:
        overrides2 = load_plan_overrides()
        con_data = (overrides2.get('plan_overrides') or {}).get('consultation', {}) or {}
        for mk in sorted(con_data.keys(), reverse=True):
            md = con_data.get(mk, {}) or {}
            if 'completion' in md and not con_progress:
                con_progress = float(md['completion'])
            if 'remaining' in md and not con_remaining:
                con_remaining = float(md['remaining'])
            if con_progress and con_remaining:
                break
    except Exception:
        pass

    return jsonify({
        'project_name':       PROJECT_NAME,
        'phase':              PROJECT_INFO.get('phase'),
        'roadmap_start':      proj_start,
        'roadmap_end':        proj_end,
        'duration_months':    int(duration_months) if duration_months else None,
        'total_mandays':      total_wd_roadmap,
        'teams_count':        len(teams),
        'teams':              sorted(teams),
        'milestones_count':   milestones_count,
        # Odoo live
        'project_manager':    pm_name,
        'coordinator':        coord_name,
        'project_value_sar':  project_value_sar,
        # Development variance
        'progress_pct':       round(project_progress, 1),
        'remaining_mds':      round(project_remaining, 1),
        'dev_eac_mds':        round(dev_eac, 1),
        # Consultation variance
        'con_progress_pct':   round(con_progress,   1),
        'con_remaining_mds':  round(con_remaining,  1),
        'con_eac_mds':        round(con_eac,        1),
        # Timeline
        'time_progress_pct':  time_progress_pct,
        'days_elapsed':       days_elapsed,
        'days_remaining':     days_remaining,
        # Project identity
        'project_id':         session.get('project_id'),
        'is_bog':             _is_bog,
    })


@app.route('/api/overview/tags-analysis')
def api_overview_tags_analysis():
    """Group tasks by their Odoo tags. Filtered by phase_group + phases."""
    _proj_name = active_project_name()
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'tags': [], 'connected': False, 'error': 'Odoo unreachable'})

    # Phase filter
    _proj_id_tags = session.get('project_id')
    _is_bog_tags  = not _proj_id_tags or str(_proj_id_tags) == '228'
    phase_group = request.args.get('phase_group', 'development')
    phases_param = request.args.get('phases')
    if phases_param:
        phase_names = [p.strip() for p in phases_param.split(',') if p.strip()]
    else:
        phase_names = get_phase_mapping().get(phase_group, [])

    try:
        # For non-BOG: auto-detect phases from project.phase by keywords
        _proj_odoo_id_tags = None
        if not phase_names and not _is_bog_tags:
            projs = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[('name', 'ilike', _proj_name)]],
                {'fields': ['id'], 'limit': 3}
            )
            if projs:
                _proj_odoo_id_tags = projs[0]['id']
                all_proj_phases = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'project.phase', 'search_read',
                    [[('project_id', '=', _proj_odoo_id_tags)]],
                    {'fields': ['id', 'name'], 'limit': 100}
                )
                if phase_group == 'support':
                    phase_names = [p['name'] for p in all_proj_phases
                                   if any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
                    if not phase_names:
                        return jsonify({'tags': [], 'connected': True, 'no_phases': True,
                                        'phases_available': [], 'summary': {},
                                        'note': f'No support/operation phases found in this project'})
                else:
                    phase_names = [p['name'] for p in all_proj_phases
                                   if not any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
                if not phase_names and all_proj_phases and phase_group != 'support':
                    phase_names = [p['name'] for p in all_proj_phases]  # fallback: all

        # Resolve phase IDs
        phase_id_to_name = {}
        if phase_names:
            phases = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.phase', 'search_read',
                [[('name', 'in', phase_names)]],
                {'fields': ['id', 'name'], 'limit': 50}
            )
            phase_id_to_name = {p['id']: p['name'] for p in phases}

        # Build base domain - tasks in this project
        project_domain = []
        if _proj_name:
            project_domain.append(('project_id.name', 'ilike', _proj_name))

        # Check if 'No Phase' was requested in phases filter
        _include_no_phase = 'No Phase' in (phases_param or '')

        # Get parent tasks under requested phases
        parent_domain = list(project_domain)
        if phase_id_to_name:
            if _include_no_phase:
                # Include tasks WITH phase AND tasks with NO phase
                parent_domain.append('|')
                parent_domain.append(('phase_id', 'in', list(phase_id_to_name.keys())))
                parent_domain.append(('phase_id', '=', False))
            else:
                parent_domain.append(('phase_id', 'in', list(phase_id_to_name.keys())))
        elif _include_no_phase:
            # Only no-phase tasks requested
            parent_domain.append(('phase_id', '=', False))

        parent_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [parent_domain],
            {'fields': ['id', 'name', 'phase_id', 'project_id'], 'limit': 5000}
        )
        parent_task_ids = {t['id'] for t in parent_tasks}
        parent_phase_map = {}  # parent_id -> phase name
        for pt in parent_tasks:
            if pt.get('phase_id'):
                parent_phase_map[pt['id']] = pt['phase_id'][1] if isinstance(pt['phase_id'], list) else ''

        # Check if there are tasks with NO phase — add 'No Phase' to available list
        _unphased_domain = list(project_domain) + [('phase_id', '=', False)]
        _unphased_count = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_count',
            [_unphased_domain]
        )
        _has_unphased = _unphased_count > 0

        # If 'No Phase' requested, add unphased tasks to parent_tasks
        if _include_no_phase and _has_unphased:
            _unphased_tasks = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.task', 'search_read',
                [_unphased_domain],
                {'fields': ['id', 'name', 'phase_id', 'project_id'], 'limit': 2000}
            )
            parent_tasks = parent_tasks + [t for t in _unphased_tasks if t['id'] not in parent_task_ids]
            parent_task_ids = {t['id'] for t in parent_tasks}

        # Get ALL project tasks (so we can walk parent chain)
        project_ids = set()
        for pt in parent_tasks:
            if pt.get('project_id') and isinstance(pt['project_id'], list):
                project_ids.add(pt['project_id'][0])

        all_domain = list(project_domain)
        if project_ids:
            all_domain = [('project_id', 'in', list(project_ids))]

        all_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [all_domain],
            {'fields': ['id', 'name', 'planned_hours', 'effective_hours',
                        'progress', 'parent_id', 'tag_ids', 'stage_id',
                        'phase_id', 'project_id'],
             'limit': 10000}
        )

        # Build lookup
        task_by_id = {t['id']: t for t in all_tasks}

        # Filter tasks: must descend from a parent_task (under our phases)
        filtered_tasks = []
        for t in all_tasks:
            cur = t
            visited = set()
            depth = 0
            found_root = None
            while cur and cur['id'] not in visited and depth < 10:
                visited.add(cur['id'])
                depth += 1
                if cur['id'] in parent_task_ids:
                    found_root = cur['id']
                    break
                if not cur.get('parent_id'):
                    break
                pid = cur['parent_id'][0] if isinstance(cur['parent_id'], list) else cur['parent_id']
                cur = task_by_id.get(pid)
            if found_root:
                t['_root_id'] = found_root
                filtered_tasks.append(t)

        # Get all tag definitions
        all_tag_ids = set()
        for t in filtered_tasks:
            for tid in t.get('tag_ids', []) or []:
                all_tag_ids.add(tid)

        tag_id_to_info = {}
        if all_tag_ids:
            tags = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.tags', 'search_read',
                [[('id', 'in', list(all_tag_ids))]],
                {'fields': ['id', 'name', 'color']}
            )
            tag_id_to_info = {t['id']: {'name': t['name'], 'color': t.get('color', 0)} for t in tags}

        # Get timesheets for these tasks
        task_ids = [t['id'] for t in filtered_tasks]
        timesheets = []
        if task_ids:
            timesheets = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'account.analytic.line', 'search_read',
                [[('task_id', 'in', task_ids)]],
                {'fields': ['task_id', 'employee_id', 'unit_amount'], 'limit': 50000}
            )

        # Aggregate hours per task
        task_hours = {}
        task_employees = {}
        for ts in timesheets:
            tid = ts['task_id'][0] if ts.get('task_id') else None
            if not tid:
                continue
            h = float(ts.get('unit_amount', 0) or 0)
            task_hours[tid] = task_hours.get(tid, 0) + h
            emp = ts.get('employee_id', [None, 'Unknown'])[1] if ts.get('employee_id') else 'Unknown'
            if tid not in task_employees:
                task_employees[tid] = {}
            task_employees[tid][emp] = task_employees[tid].get(emp, 0) + h

        # Aggregate by tag (with per-task breakdown)
        tag_data = {}
        UNTAGGED = -1

        for t in filtered_tasks:
            tid = t['id']
            planned = float(t.get('planned_hours') or 0)
            actual = task_hours.get(tid, 0)
            task_emps = task_employees.get(tid, {})

            # Skip tasks with no tags AND no work AND no planning (noise)
            tag_ids_for_task = t.get('tag_ids', []) or []
            if not tag_ids_for_task:
                if planned == 0 and actual == 0:
                    continue
                tag_ids_for_task = [UNTAGGED]

            for tag_id in tag_ids_for_task:
                if tag_id not in tag_data:
                    tag_data[tag_id] = {
                        'tag_id': tag_id,
                        'name': tag_id_to_info.get(tag_id, {}).get('name', 'Untagged') if tag_id != UNTAGGED else 'Untagged',
                        'color': tag_id_to_info.get(tag_id, {}).get('color', 0) if tag_id != UNTAGGED else 0,
                        'planned': 0,
                        'actual': 0,
                        'tasks_count': 0,
                        'employees': {},
                        'tasks': [],  # per-task breakdown
                    }
                tag_data[tag_id]['planned'] += planned
                tag_data[tag_id]['actual'] += actual
                tag_data[tag_id]['tasks_count'] += 1
                tag_data[tag_id]['tasks'].append({
                    'id': tid,
                    'name': t.get('name', ''),
                    'planned_hours': round(planned, 1),
                    'actual_hours': round(actual, 1),
                    'employees': sorted([{'name': e[0], 'hours': round(e[1], 1)} for e in task_emps.items()],
                                        key=lambda x: -x['hours'])[:5],
                })
                for emp, eh in task_emps.items():
                    tag_data[tag_id]['employees'][emp] = tag_data[tag_id]['employees'].get(emp, 0) + eh

        # Build result
        result = []
        for tag_id, td in tag_data.items():
            planned = td['planned']
            actual = td['actual']
            remaining = max(0, planned - actual)
            progress = round(min(100, actual / planned * 100), 1) if planned > 0 else 0
            sorted_emps = sorted(td['employees'].items(), key=lambda x: -x[1])
            # Sort tasks by actual hours desc
            td['tasks'].sort(key=lambda x: -x['actual_hours'])
            result.append({
                'tag_id': td['tag_id'],
                'name': td['name'],
                'color': td['color'],
                'planned_hours': round(planned, 1),
                'actual_hours': round(actual, 1),
                'remaining_hours': round(remaining, 1),
                'planned_days': round(planned / WORK_HOURS_PER_DAY, 1),
                'actual_days': round(actual / WORK_HOURS_PER_DAY, 1),
                'progress_pct': progress,
                'tasks_count': td['tasks_count'],
                'top_employees': [{'name': e[0], 'hours': round(e[1], 1)} for e in sorted_emps[:5]],
                'employees_count': len(td['employees']),
                'tasks': td['tasks'],
            })

        result.sort(key=lambda x: -x['actual_hours'])

        return jsonify({
            'tags': result,
            'connected': True,
            'phases_active': phase_names,
            'phases_available': [v for v in phase_id_to_name.values() if v] + (['No Phase'] if _has_unphased else []),
            'phase_group': phase_group,
            'summary': {
                'total_tags': len(result),
                'total_planned': round(sum(t['planned_hours'] for t in result), 1),
                'total_actual': round(sum(t['actual_hours'] for t in result), 1),
                'total_remaining': round(sum(t['remaining_hours'] for t in result), 1),
            }
        })
    except Exception as e:
        logger.error(f"api_overview_tags_analysis: {e}\n{traceback.format_exc()}")
        return jsonify({'tags': [], 'connected': False, 'error': str(e)})


@app.route('/api/overview/analysis/<phase_group>')
def api_overview_analysis(phase_group):
    """Per-task analysis from Odoo with multi-phase + employee filter support.
    phase_group: 'development' or 'consultation' — defines default phase set.
    Query params:
      - phases: comma-separated list of phase names (overrides default)
      - employees: comma-separated employee names to filter by
    """
    _proj_name = active_project_name()
    _proj_id = session.get('project_id')
    _is_bog  = not _proj_id or str(_proj_id) == '228'

    # Parse query params
    phases_param = request.args.get('phases')
    if phases_param:
        phase_names = [p.strip() for p in phases_param.split(',') if p.strip()]
    else:
        phase_names = get_phase_mapping().get(phase_group, [])

    employees_param = request.args.get('employees')
    employees_filter = []
    if employees_param:
        employees_filter = [e.strip() for e in employees_param.split(',') if e.strip()]

    # For non-BOG: get project phases from Odoo project.phase filtered by services/support keywords
    _proj_odoo_id = None
    if not phase_names and not _is_bog:
        if not odoo.uid: odoo.connect()
        try:
            projects = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[('name', 'ilike', _proj_name)]],
                {'fields': ['id'], 'limit': 3}
            )
            if projects:
                _proj_odoo_id = projects[0]['id']
                # Get ALL project.phase records for this project
                all_proj_phases = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'project.phase', 'search_read',
                    [[('project_id', '=', _proj_odoo_id)]],
                    {'fields': ['id', 'name'], 'limit': 100}
                )
                if all_proj_phases:
                    # Split by support keywords
                    if phase_group == 'support':
                        phase_names = [p['name'] for p in all_proj_phases
                                       if any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
                    else:  # services = all except support
                        phase_names = [p['name'] for p in all_proj_phases
                                       if not any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
                else:
                    # No project.phase — use ALL tasks (no phase filter)
                    phase_names = ['__all__']
        except Exception as _e:
            logger.warning(f"Phase detect failed: {_e}")

    if not phase_names and not _is_bog:
        return jsonify({'tasks': [], 'connected': True, 'phases_active': [],
                        'phases_available': [], 'employees_available': [], 'stages_used': [],
                        'no_phases': True,
                        'note': f'No {phase_group} phases found — tasks in this project are not assigned to a phase named with {phase_group} keywords'})

    if not phase_names:
        return jsonify({'tasks': [], 'error': f'No phases for {phase_group}'})

    # Connect to Odoo
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'tasks': [], 'connected': False, 'error': 'Odoo unreachable'})

    try:
        _has_unphased = False  # default — updated after project domain is known
        _include_no_phase = 'No Phase' in (phases_param or '')

        # Get phase IDs — try project.phase first (BOG), then task.type (non-BOG)
        phases = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('name', 'in', phase_names)]],
            {'fields': ['id', 'name'], 'limit': 50}
        )
        phase_ids = [p['id'] for p in phases]
        phase_id_to_name = {p['id']: p['name'] for p in phases}

        # Handle __all__ sentinel (no project.phase records — use all project tasks)
        use_all_tasks = '__all__' in phase_names
        if use_all_tasks:
            phase_ids = []
            phase_id_to_name = {}
            if not _proj_odoo_id:
                _proj_odoo_id = get_project_odoo_id(_proj_name)
        
        if not phase_ids and not _is_bog and not use_all_tasks:
            if not _proj_odoo_id:
                _proj_odoo_id = get_project_odoo_id(_proj_name)

        # Step 1: Get parent tasks — by phase_id or all project tasks
        if use_all_tasks and _proj_odoo_id:
            project_domain = [('project_id', '=', _proj_odoo_id),
                               ('parent_id', '=', False)]
        elif phase_ids:
            project_domain = [('phase_id', 'in', phase_ids)]
            if _proj_name:
                project_domain.append(('project_id.name', 'ilike', _proj_name))
        elif _proj_odoo_id:
            project_domain = [('project_id', '=', _proj_odoo_id),
                               ('parent_id', '=', False)]
        else:
            return jsonify({'tasks': [], 'error': f'No phases for {phase_group}'})

        # Try to detect available assignment fields in this Odoo version.
        # Common multi-assignee field names:
        #   user_ids       (Odoo standard - newer)
        #   project_user_ids (Custom v14 - this codebase)
        #   users_list     (string with comma-separated names - this codebase)
        TASK_FIELDS_BASE = ['id', 'name', 'planned_hours', 'effective_hours',
                            'progress', 'parent_id', 'user_id', 'project_id',
                            'date_deadline', 'stage_id', 'kanban_state',
                            'phase_id', 'date_start', 'date_end', 'child_ids',
                            'project_user_ids', 'users_list']
        # Set to true if extra fields were accepted
        use_multi_assignee = True
        try:
            parent_tasks = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.task', 'search_read',
                [project_domain],
                {'fields': TASK_FIELDS_BASE, 'limit': 5000}
            )
        except Exception as e:
            if 'Invalid field' in str(e):
                # Drop the multi-assignee fields if not available
                FALLBACK_FIELDS = ['id', 'name', 'planned_hours', 'effective_hours',
                                   'progress', 'parent_id', 'user_id', 'project_id',
                                   'date_deadline', 'stage_id', 'kanban_state',
                                   'phase_id', 'date_start', 'date_end', 'child_ids']
                use_multi_assignee = False
                logger.info(f"Multi-assignee fields not available: {e}, falling back to user_id only")
                parent_tasks = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'project.task', 'search_read',
                    [project_domain],
                    {'fields': FALLBACK_FIELDS, 'limit': 5000}
                )
                TASK_FIELDS_BASE = FALLBACK_FIELDS
            else:
                raise

        if not parent_tasks:
            return jsonify({'tasks': [], 'connected': True, 'phases_available': phase_names})

        # Diagnostic: log parent tasks with their child_ids count
        for pt in parent_tasks:
            logger.info(f"  Parent task: id={pt['id']} name='{pt.get('name')}' child_ids={len(pt.get('child_ids') or [])}")

        # Step 2: Fetch ALL tasks belonging to the same project as parent tasks
        # (use project_id from parent_tasks, not name match — more reliable)
        project_ids = set()
        for pt in parent_tasks:
            if pt.get('project_id') and isinstance(pt['project_id'], list) and len(pt['project_id']) > 0:
                project_ids.add(pt['project_id'][0])

        if project_ids:
            all_project_domain = [('project_id', 'in', list(project_ids))]
        elif PROJECT_NAME:
            all_project_domain = [('project_id.name', 'ilike', _proj_name)]
        else:
            all_project_domain = []

        # Check for unphased tasks
        _unphased_domain = list(all_project_domain) + [('phase_id', '=', False)]
        try:
            _unphased_count = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.task', 'search_count', [_unphased_domain]
            )
            _has_unphased = _unphased_count > 0
        except Exception:
            _has_unphased = False

        # If No Phase selected, add unphased tasks to parent_tasks
        if _include_no_phase and _has_unphased:
            _unphased_tasks = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.task', 'search_read',
                [_unphased_domain],
                {'fields': ['id', 'name', 'phase_id', 'project_id'], 'limit': 2000}
            )
            _parent_ids = {t['id'] for t in parent_tasks}
            parent_tasks = parent_tasks + [t for t in _unphased_tasks if t['id'] not in _parent_ids]

        all_project_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [all_project_domain],
            {'fields': TASK_FIELDS_BASE, 'limit': 10000}
        )
        logger.info(f"Total project tasks fetched: {len(all_project_tasks)} (multi_assignee={use_multi_assignee})")

        # Build lookup: id -> task
        task_by_id = {t['id']: t for t in all_project_tasks}
        parent_task_ids = {t['id'] for t in parent_tasks}

        # For each task, walk up parent chain to see if it descends from a parent_task
        all_tasks_by_id = {}
        for t in all_project_tasks:
            cur = t
            visited = set()
            depth = 0
            while cur and cur['id'] not in visited and depth < 10:
                visited.add(cur['id'])
                depth += 1
                if cur['id'] in parent_task_ids:
                    # this task descends from one of our parents (or IS a parent)
                    # add the original task t (not cur)
                    all_tasks_by_id[t['id']] = t
                    break
                # Move up
                if not cur.get('parent_id'):
                    break
                parent_id = cur['parent_id'][0] if isinstance(cur['parent_id'], list) else cur['parent_id']
                cur = task_by_id.get(parent_id)
                if not cur:
                    break

        tasks = list(all_tasks_by_id.values())
        logger.info(f"Tasks in scope: {len(parent_tasks)} parents → {len(tasks)} total (after walking parent chains)")

        # Get all timesheet entries for these tasks
        task_ids = [t['id'] for t in tasks]
        timesheets = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.analytic.line', 'search_read',
            [[('task_id', 'in', task_ids)]],
            {'fields': ['task_id', 'employee_id', 'unit_amount', 'date'], 'limit': 50000}
        )

        # Group timesheets by task → employee
        ts_by_task = {}
        all_employees_set = set()
        for ts in timesheets:
            tid = ts['task_id'][0] if ts.get('task_id') else None
            if not tid:
                continue
            emp = ts.get('employee_id', [None, 'Unknown'])[1] if ts.get('employee_id') else 'Unknown'
            all_employees_set.add(emp)
            h = float(ts.get('unit_amount', 0) or 0)
            d = ts.get('date')
            if tid not in ts_by_task:
                ts_by_task[tid] = {'employees': {}, 'first_date': None, 'last_date': None}
            ts_by_task[tid]['employees'][emp] = ts_by_task[tid]['employees'].get(emp, 0) + h
            if d:
                if not ts_by_task[tid]['first_date'] or d < ts_by_task[tid]['first_date']:
                    ts_by_task[tid]['first_date'] = d
                if not ts_by_task[tid]['last_date'] or d > ts_by_task[tid]['last_date']:
                    ts_by_task[tid]['last_date'] = d

        # Collect all stages used (for visualization)
        stages_used = {}  # id -> name
        for t in tasks:
            if t.get('stage_id') and isinstance(t['stage_id'], list) and len(t['stage_id']) > 1:
                stages_used[t['stage_id'][0]] = t['stage_id'][1]

        # Pre-batch: collect ALL project_user_ids from all tasks → resolve in single query
        all_user_ids_to_resolve = set()
        tasks_with_assignees = 0
        if use_multi_assignee:
            for t in tasks:
                # project_user_ids is the multi-assignee field in this Odoo
                if t.get('project_user_ids') and isinstance(t['project_user_ids'], list) and t['project_user_ids']:
                    tasks_with_assignees += 1
                    for uid in t['project_user_ids']:
                        all_user_ids_to_resolve.add(uid)

        logger.info(f"Tasks with assignees: {tasks_with_assignees}, total user_ids to resolve: {len(all_user_ids_to_resolve)}")

        user_id_to_name = {}
        if all_user_ids_to_resolve:
            try:
                users_data = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'res.users', 'read',
                    [list(all_user_ids_to_resolve)], {'fields': ['name']}
                )
                user_id_to_name = {u['id']: u['name'] for u in users_data if u.get('name')}
                logger.info(f"Resolved {len(user_id_to_name)} assignees")
            except Exception as ue:
                logger.warning(f"Batch user resolution failed: {ue}")

        # Build task list
        task_id_to_obj = {}
        for t in tasks:
            tid = t['id']
            ts_info = ts_by_task.get(tid, {'employees': {}, 'first_date': None, 'last_date': None})
            actual_hours = sum(ts_info['employees'].values())
            planned_hours = float(t.get('planned_hours') or 0)
            progress_pct = float(t.get('progress') or 0)
            if progress_pct == 0 and planned_hours > 0:
                progress_pct = min(150, round(actual_hours / planned_hours * 100, 1))

            sorted_emps = sorted(ts_info['employees'].items(), key=lambda x: -x[1])
            allocation = [{'name': e[0], 'hours': round(e[1], 1)} for e in sorted_emps]

            # Collect ALL assignee names
            # Priority: project_user_ids (resolved IDs) > users_list (string) > user_id (single)
            assignee_names = []

            # 1. Multi-user field (project_user_ids) - list of integer IDs
            if use_multi_assignee and t.get('project_user_ids') and isinstance(t['project_user_ids'], list):
                for uid in t['project_user_ids']:
                    name = user_id_to_name.get(uid)
                    if name and name not in assignee_names:
                        assignee_names.append(name)

            # 2. users_list (comma-separated string of names) - fallback if IDs didn't resolve
            if not assignee_names and use_multi_assignee and t.get('users_list'):
                ulist_str = str(t['users_list']).strip()
                if ulist_str:
                    # Split by common separators
                    for n in ulist_str.replace(';', ',').split(','):
                        n = n.strip()
                        if n and n not in assignee_names:
                            assignee_names.append(n)

            # 3. Single user field (user_id) - returns [id, name] tuple
            assignee_name = None
            if t.get('user_id') and isinstance(t['user_id'], list) and len(t['user_id']) > 1:
                assignee_name = t['user_id'][1]
                if assignee_name and assignee_name not in assignee_names:
                    assignee_names.append(assignee_name)

            # Add all assignees to all_employees_set + allocation
            for aname in assignee_names:
                all_employees_set.add(aname)
                if aname not in ts_info['employees']:
                    allocation.append({'name': aname, 'hours': 0})

            stage = ''
            stage_id = None
            if t.get('stage_id') and isinstance(t['stage_id'], list) and len(t['stage_id']) > 1:
                stage_id = t['stage_id'][0]
                stage = t['stage_id'][1]

            _ph = t.get('phase_id')
            if _ph and isinstance(_ph, list) and _ph[0]:
                phase_label = phase_id_to_name.get(_ph[0]) or (_ph[1] if len(_ph)>1 else 'No Phase')
            else:
                phase_label = 'No Phase'

            parent_id = None
            parent_name = ''
            if t.get('parent_id') and isinstance(t['parent_id'], list) and len(t['parent_id']) > 1:
                parent_id = t['parent_id'][0]
                parent_name = t['parent_id'][1]

            child_count = len(t.get('child_ids') or [])

            task_id_to_obj[tid] = {
                'id': tid,
                'name': t.get('name'),
                'parent_id': parent_id,
                'parent_name': parent_name,
                'is_parent': not bool(parent_id),
                'child_count': child_count,
                'phase': phase_label,
                'stage': stage,
                'stage_id': stage_id,
                'kanban_state': t.get('kanban_state') or 'normal',
                'assignee': assignee_name,
                'planned_hours': round(planned_hours, 1),
                'actual_hours': round(actual_hours, 1),
                'planned_days': round(planned_hours / WORK_HOURS_PER_DAY, 1) if planned_hours else 0,
                'actual_days': round(actual_hours / WORK_HOURS_PER_DAY, 1),
                'progress_pct': progress_pct,
                'deadline': t.get('date_deadline'),
                'date_start': t.get('date_start'),
                'date_end': t.get('date_end'),
                'allocation': allocation,
                'inherited_allocation': [],  # filled below if task has no direct allocation
                'first_log': ts_info['first_date'],
                'last_log': ts_info['last_date'],
                # Roll-up totals (filled below)
                'subtask_planned_hours': 0,
                'subtask_actual_hours': 0,
                'subtask_count_total': 0,  # all descendants count
                # Allocations from descendants (for parent display)
                'rollup_allocation': {},
            }

        # Build parent->children index for roll-up
        children_by_parent = {}
        for tid, obj in task_id_to_obj.items():
            if obj['parent_id'] and obj['parent_id'] in task_id_to_obj:
                children_by_parent.setdefault(obj['parent_id'], []).append(tid)

        def rollup(tid):
            """Recursively sum planned + actual from all descendants."""
            obj = task_id_to_obj.get(tid)
            if not obj:
                return 0, 0, 0, {}
            child_ids = children_by_parent.get(tid, [])
            sub_planned = 0
            sub_actual = 0
            sub_count = 0
            sub_alloc = {}  # emp -> hours
            # Add this task's own allocation to rollup
            for a in obj.get('allocation', []):
                sub_alloc[a['name']] = sub_alloc.get(a['name'], 0) + a['hours']
            for cid in child_ids:
                # Recurse
                cp, ca, cn, calloc = rollup(cid)
                # Direct child contributes itself + its descendants
                child_obj = task_id_to_obj[cid]
                sub_planned += child_obj['planned_hours'] + cp
                sub_actual += child_obj['actual_hours'] + ca
                sub_count += 1 + cn
                for emp, h in calloc.items():
                    sub_alloc[emp] = sub_alloc.get(emp, 0) + h
            obj['subtask_planned_hours'] = round(sub_planned, 1)
            obj['subtask_actual_hours'] = round(sub_actual, 1)
            obj['subtask_count_total'] = sub_count
            # For parent display: total includes own + sub
            obj['total_planned_hours'] = round(obj['planned_hours'] + sub_planned, 1)
            obj['total_actual_hours'] = round(obj['actual_hours'] + sub_actual, 1)
            # Compute roll-up progress (uses total)
            if obj['total_planned_hours'] > 0:
                obj['rollup_progress_pct'] = min(150, round(obj['total_actual_hours'] / obj['total_planned_hours'] * 100, 1))
            else:
                obj['rollup_progress_pct'] = obj['progress_pct']
            # Remaining = total_planned - total_actual
            obj['total_remaining_hours'] = max(0, round(obj['total_planned_hours'] - obj['total_actual_hours'], 1))
            # Per-task remaining = planned - actual (own only)
            obj['remaining_hours'] = max(0, round(obj['planned_hours'] - obj['actual_hours'], 1))
            obj['rollup_allocation'] = sub_alloc
            return sub_planned, sub_actual, sub_count, sub_alloc

        # Run rollup for every task
        for tid in list(task_id_to_obj.keys()):
            rollup(tid)

        # Inheritance: if a task has no allocation, look at parent's allocation
        # This handles new tasks where assignee hasn't been set yet (e.g. new resources)
        for tid, t in task_id_to_obj.items():
            if not t['allocation'] and t.get('parent_id'):
                # Walk up to find first ancestor with allocation
                cur_pid = t['parent_id']
                visited = set()
                while cur_pid and cur_pid not in visited:
                    visited.add(cur_pid)
                    parent = task_id_to_obj.get(cur_pid)
                    if not parent:
                        break
                    if parent.get('allocation'):
                        # Inherit allocation as "context" (lower hours = 0 to indicate inherited)
                        t['inherited_allocation'] = parent['allocation'][:3]  # top 3 from parent
                        t['allocation_source'] = f"inherited from: {parent['name']}"
                        break
                    cur_pid = parent.get('parent_id')

        # Apply employee filter (matches allocation OR assignee)
        # Walk UP entire parent chain to keep ancestors visible (for context)
        if employees_filter:
            matching_task_ids = set()
            for tid, t in task_id_to_obj.items():
                emp_names = {a['name'] for a in t.get('allocation', [])}
                if t.get('assignee'):
                    emp_names.add(t['assignee'])
                if any(e in emp_names for e in employees_filter):
                    matching_task_ids.add(tid)
                    cur = t
                    visited = {tid}
                    while cur and cur.get('parent_id') and cur['parent_id'] not in visited:
                        pid = cur['parent_id']
                        if pid in task_id_to_obj:
                            matching_task_ids.add(pid)
                            visited.add(pid)
                            cur = task_id_to_obj[pid]
                        else:
                            break
            task_id_to_obj = {k: v for k, v in task_id_to_obj.items() if k in matching_task_ids}

        result = list(task_id_to_obj.values())

        # Sort: parents first by name, then their children grouped
        # Build hierarchical order
        parents = [t for t in result if t['is_parent']]
        parents.sort(key=lambda x: (x['name'] or '').lower())

        ordered = []
        seen = set()
        for p in parents:
            ordered.append(p)
            seen.add(p['id'])
            # Append children of this parent
            children = [t for t in result if t['parent_id'] == p['id']]
            children.sort(key=lambda x: (-x['progress_pct'], x['name'] or ''))
            for c in children:
                ordered.append(c)
                seen.add(c['id'])
        # Append orphans (children without parent in current set)
        for t in result:
            if t['id'] not in seen:
                ordered.append(t)

        # Summary
        total_planned = sum(t['planned_hours'] for t in ordered)
        total_actual = sum(t['actual_hours'] for t in ordered)
        overall_progress = round(min(150, total_actual / total_planned * 100), 1) if total_planned > 0 else 0

        # Add 'No Phase' to available if unphased tasks exist
        phases_avail = [p for p in phase_names if p]
        if _has_unphased and 'No Phase' not in phases_avail:
            phases_avail.append('No Phase')

        return jsonify({
            'tasks': ordered,
            'connected': True,
            'phases_available': phases_avail,
            'phases_active': phases_avail,
            'employees_available': sorted(all_employees_set),
            'employees_filter': employees_filter,
            'stages_used': [{'id': sid, 'name': name} for sid, name in stages_used.items()],
            'summary': {
                'total_tasks': len(ordered),
                'parent_tasks': sum(1 for t in ordered if t['is_parent']),
                'sub_tasks': sum(1 for t in ordered if not t['is_parent']),
                'total_planned_hours': round(total_planned, 1),
                'total_actual_hours': round(total_actual, 1),
                'total_planned_days': round(total_planned / WORK_HOURS_PER_DAY, 1),
                'total_actual_days': round(total_actual / WORK_HOURS_PER_DAY, 1),
                'overall_progress_pct': overall_progress,
                'tasks_with_planning': sum(1 for t in ordered if t['planned_hours'] > 0),
                'tasks_in_progress': sum(1 for t in ordered if 0 < t['progress_pct'] < 100),
                'tasks_completed': sum(1 for t in ordered if t['progress_pct'] >= 100),
            },
            'phase_group': phase_group,
        })
    except Exception as e:
        logger.error(f"api_overview_analysis: {e}\n{traceback.format_exc()}")
        return jsonify({'tasks': [], 'connected': False, 'error': str(e)})



SERVICES_OVERRIDES_FILE = os.path.join(PERSIST_DIR, 'services_overrides.json')
RISKS_FILE = os.path.join(PERSIST_DIR, 'risks_issues.json')

# Google Sheet for team members
TEAM_SHEET_ID = os.environ.get('TEAM_SHEET_ID', '1MtpNyBKnoayhdgpDbe2WacjZgJOLCSeMjhDuOhabo0Y')
TEAM_SHEET_GID = os.environ.get('TEAM_SHEET_GID', '0')

def fetch_team_members_from_sheet():
    """Fetch team members from public Google Sheet (CSV export).
    Returns: {success: bool, members: [...], error: str, columns: [...]}
    """
    if not TEAM_SHEET_ID:
        return {'success': False, 'error': 'TEAM_SHEET_ID not configured', 'members': []}

    csv_url = f'https://docs.google.com/spreadsheets/d/{TEAM_SHEET_ID}/export?format=csv&gid={TEAM_SHEET_GID}'
    try:
        import urllib.request
        req = urllib.request.Request(csv_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            content = resp.read().decode('utf-8', errors='replace')

        import csv
        from io import StringIO
        reader = csv.reader(StringIO(content))
        rows = list(reader)
        if not rows:
            return {'success': False, 'error': 'Sheet is empty', 'members': []}

        # First row = headers
        headers = [h.strip() for h in rows[0]]

        # Detect department column (first column, usually called "Department")
        dept_col_idx = 0
        for i, h in enumerate(headers):
            if h.lower() in ('department', 'team', 'group'):
                dept_col_idx = i
                break

        members = []
        last_dept = ''  # track last seen department for merged-cell forward-fill
        for r in rows[1:]:
            if not r:
                continue

            # Forward-fill department column (handles merged cells in sheet)
            dept_value = (r[dept_col_idx] if dept_col_idx < len(r) else '').strip()
            if dept_value:
                last_dept = dept_value

            # Build row dict using current values
            row_dict = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = (r[i] if i < len(r) else '').strip()
                row_dict[h] = val

            # Override department with forward-filled value
            if dept_col_idx < len(headers):
                row_dict[headers[dept_col_idx]] = last_dept

            # Skip empty/separator rows: must have at least a member name
            # Find the "Members" column (usually 2nd col)
            member_value = ''
            for h_key in row_dict:
                if h_key.lower() in ('members', 'member', 'name', 'full name'):
                    member_value = row_dict[h_key]
                    break
            # Fallback: use 2nd column if no Members header found
            if not member_value and len(r) > 1:
                member_value = (r[1] if 1 < len(r) else '').strip()

            if not member_value:
                continue  # skip separator rows

            members.append(row_dict)

        return {
            'success': True,
            'members': members,
            'columns': headers,
            'total': len(members),
            'sheet_url': f'https://docs.google.com/spreadsheets/d/{TEAM_SHEET_ID}/edit'
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'Could not fetch sheet: {e}. Make sure sheet is shared "Anyone with link can view".',
            'members': [],
            'sheet_url': f'https://docs.google.com/spreadsheets/d/{TEAM_SHEET_ID}/edit'
        }


@app.route('/api/team/summary')
def api_team_summary():
    """Returns count of ACTIVE members (Active=Yes) for the overview KPI card.
    Inactive members and unassigned slots are excluded from this count."""
    result = fetch_team_members_from_sheet()
    if not result['success']:
        return jsonify({'success': False, 'total': 0, 'error': result.get('error')})

    members = result.get('members', [])
    columns = result.get('columns', [])
    cols_lower = {c.lower(): c for c in columns}

    # Find columns
    active_col = None
    member_col = None
    site_col = None
    for c_lower, c_orig in cols_lower.items():
        if 'active' in c_lower and not active_col:
            active_col = c_orig
        if c_lower in ('members', 'member', 'name', 'full name') and not member_col:
            member_col = c_orig
        if ('onsite' in c_lower or 'offshore' in c_lower or 'on-site' in c_lower) and not site_col:
            site_col = c_orig

    active_count = 0
    onsite_count = 0
    offshore_count = 0
    for m in members:
        # Skip unassigned
        name = (m.get(member_col, '') if member_col else '').strip().lower()
        if not name or 'not assigned' in name or 'unassigned' in name:
            continue
        # Check active flag
        if active_col:
            active_val = (m.get(active_col, '') or '').strip().lower()
            if active_val in ('no', 'n', 'false', '0', 'inactive'):
                continue
        active_count += 1

        # Read onsite/offshore from dedicated column in sheet
        if site_col:
            sv = (m.get(site_col, '') or '').strip().lower()
            if sv in ('onsite', 'on-site', 'on site'):
                onsite_count += 1
            elif sv in ('offshore', 'off-shore', 'off shore', 'remote'):
                offshore_count += 1

    return jsonify({
        'success': True,
        'total': active_count,
        'total_raw': len(members),
        'onsite': onsite_count,
        'offshore': offshore_count,
    })


@app.route('/api/team/members')
def api_team_members():
    """Returns full team list with auto-detected hierarchy"""
    result = fetch_team_members_from_sheet()
    if not result['success']:
        return jsonify(result)

    members = result['members']
    columns = result['columns']

    # Auto-detect column intent (case-insensitive matching)
    col_map = {}
    cols_lower = {c.lower(): c for c in columns}
    for keyword, key in [
        ('member', 'name'), ('full name', 'name'), ('name', 'name'),
        ('title', 'title'),  # job title (PM, Lead BA, etc.)
        ('position', 'position'),  # full position string (KSA - Project Manager)
        ('postion', 'position'),  # typo handling
        ('role', 'role'),
        ('department', 'department'), ('team', 'team'), ('squad', 'team'), ('group', 'team'),
        ('manager', 'manager'), ('reports to', 'manager'), ('reporting', 'manager'),
        ('hour rate', 'hour_rate'), ('rate', 'hour_rate'),
        ('overtime', 'overtime_rate'),
        ('active', 'active'),
        ('allocation', 'allocation'),
        ('onsite/offshore', 'onsite_offshore'),  # explicit column from sheet
        ('on-site', 'onsite_offshore'),
        ('onsite', 'onsite_offshore'),
        ('offshore', 'onsite_offshore'),
        ('site', 'onsite_offshore'),
        ('email', 'email'), ('mail', 'email'),
        ('phone', 'phone'), ('mobile', 'phone'),
        ('country', 'country'), ('location', 'country'),
        ('start date', 'start_date'), ('joined', 'start_date'),
        ('status', 'status'),
    ]:
        for c_lower, c_orig in cols_lower.items():
            if keyword in c_lower and key not in col_map:
                col_map[key] = c_orig
                break

    # Normalize each member
    normalized = []
    for m in members:
        # Get name
        name = m.get(col_map.get('name', ''), '') or m.get(columns[0] if columns else '', '')
        if not name or name.lower().startswith('not assigned'):
            # Mark unassigned slots
            name = name or '(Unassigned)'

        # Get position string
        position_str = m.get(col_map.get('position', ''), '')

        # Read onsite/offshore directly from sheet column (case-insensitive)
        onsite_offshore = ''
        site_val = (m.get(col_map.get('onsite_offshore', ''), '') or '').strip().lower()
        if site_val in ('onsite', 'on-site', 'on site'):
            onsite_offshore = 'Onsite'
        elif site_val in ('offshore', 'off-shore', 'off shore', 'remote'):
            onsite_offshore = 'Offshore'

        # Active flag - default Yes if column missing. Unassigned slots are NOT active.
        active_val = (m.get(col_map.get('active', ''), '') or '').strip().lower()
        is_unassigned = name == '(Unassigned)' or 'not assigned' in (name or '').lower()
        is_active = (active_val not in ('no', 'n', 'false', '0', 'inactive')) and not is_unassigned

        # Title is the simple title (PM, Lead BA), use as role if no role column
        title = m.get(col_map.get('title', ''), '')
        role = m.get(col_map.get('role', ''), '') or title

        normalized.append({
            'name': name,
            'role': role,
            'title': title,
            'position': position_str,
            'team': m.get(col_map.get('team', ''), ''),
            'manager': m.get(col_map.get('manager', ''), ''),
            'department': m.get(col_map.get('department', ''), ''),
            'email': m.get(col_map.get('email', ''), ''),
            'phone': m.get(col_map.get('phone', ''), ''),
            'country': m.get(col_map.get('country', ''), ''),
            'status': m.get(col_map.get('status', ''), ''),
            'hour_rate': m.get(col_map.get('hour_rate', ''), ''),
            'allocation': m.get(col_map.get('allocation', ''), ''),
            'active': is_active,
            'onsite_offshore': onsite_offshore,
            'is_unassigned': is_unassigned,
            'raw': m,
        })

    # Build hierarchy: Use Department column from sheet directly
    # Position rank from Title field (Manager > Lead > Senior > Junior > Unassigned)

    def position_rank(title, role):
        """Return rank: 0 = highest, higher = lower"""
        text = ((title or '') + ' ' + (role or '')).lower()
        # Manager / Director / PM
        if any(k in text for k in ['director', 'head of', 'chief', 'cto', 'cio', 'ceo']):
            return 0
        if any(k in text for k in ['manager', 'pm', 'project coordinator', 'coordinator']):
            return 1
        if 'pmo' in text:
            return 2
        # Lead / Principal
        if any(k in text for k in ['lead', 'principal']):
            return 3
        # Senior
        if any(k in text for k in ['senior', 'sr.', 'sr ', 'lead ba', 'manager ba']):
            return 4
        # Mid (default for engineers without prefix)
        if any(k in text for k in ['engineer', 'analyst', 'designer', 'developer', 'consultant', 'specialist',
                                    'sw', 'ba ', 'ux ', 'qc']):
            if 'junior' in text or 'jr' in text:
                return 6
            return 5
        # Junior
        if any(k in text for k in ['junior', 'jr', 'intern', 'trainee']):
            return 6
        return 7

    # Group by Department directly
    # RULE: Hide inactive members from hierarchy
    # EXCEPTION: Mariam (PMO) is always shown in Management even if inactive
    dept_groups = {}
    for m in normalized:
        # Skip inactive members EXCEPT Mariam (PMO is always shown)
        if not m.get('active'):
            name_lower = (m.get('name') or '').lower()
            title_lower = (m.get('title') or '').lower()
            is_mariam_pmo = 'mariam' in name_lower and 'pmo' in title_lower
            if not is_mariam_pmo:
                continue  # skip this inactive member

        # Skip unassigned slots
        if m.get('is_unassigned'):
            continue

        dept = (m.get('department') or '').strip() or 'Other'
        if dept not in dept_groups:
            dept_groups[dept] = []
        dept_groups[dept].append(m)

    # Sort within each department by position rank, then by name
    def sort_key(m):
        return (
            position_rank(m.get('title'), m.get('role')),
            (m.get('name') or '').lower()
        )

    for d in dept_groups:
        dept_groups[d].sort(key=sort_key)

    # Build final ordered groups list
    grouped_list = []

    # Define preferred order for departments
    DEPT_ORDER = [
        'Management',
        'Business Team', 'Bussines Team',  # support both spellings
        'Development Team', 'Development',
        'AI Team', 'AI',
        'UX Team', 'UI/UX', 'UX',
        'QC Team', 'QC',
        'UAT Team', 'UAT',
        'Infrastructure Team', 'InfraStructure Team', 'InfraStrcture Team', 'Infrastructure',
        'DevOps',
    ]

    def dept_order_key(name):
        # Find the first matching pattern in DEPT_ORDER
        n = name.lower().strip()
        for i, dn in enumerate(DEPT_ORDER):
            if dn.lower() == n:
                return i
        # Partial match
        for i, dn in enumerate(DEPT_ORDER):
            if dn.lower() in n or n in dn.lower():
                return i
        return 999

    sorted_depts = sorted(dept_groups.items(), key=lambda x: (dept_order_key(x[0]), x[0]))

    for dept_name, dept_members in sorted_depts:
        # Add icon for Management
        display_name = dept_name
        is_mgmt = 'management' in dept_name.lower()
        if is_mgmt and not display_name.startswith('👑'):
            display_name = '👑 ' + display_name

        # Counts (only members shown in this group - inactive already excluded above)
        # Mariam (PMO) is the only inactive shown - count her separately
        shown_count = len(dept_members)
        onsite_count = sum(1 for m in dept_members if m.get('onsite_offshore') == 'Onsite')
        offshore_count = sum(1 for m in dept_members if m.get('onsite_offshore') == 'Offshore')

        grouped_list.append({
            'name': display_name,
            'members': dept_members,
            'count': shown_count,
            'active_count': shown_count,  # all shown members are "in scope"
            'onsite_count': onsite_count,
            'offshore_count': offshore_count,
            'is_management': is_mgmt,
        })

    # Compute totals from ALL normalized (active count = excludes inactive + unassigned)
    total_active = sum(1 for m in normalized if m.get('active'))
    total_onsite = sum(1 for m in normalized if m.get('onsite_offshore') == 'Onsite' and m.get('active'))
    total_offshore = sum(1 for m in normalized if m.get('onsite_offshore') == 'Offshore' and m.get('active'))

    return jsonify({
        'success': True,
        'total': len(normalized),
        'total_active': total_active,
        'total_onsite': total_onsite,
        'total_offshore': total_offshore,
        'columns': columns,
        'col_map': col_map,
        'hierarchy_key': 'department',
        'groups': grouped_list,
        'all_members': normalized,
        'sheet_url': result['sheet_url'],
    })


def load_risks():
    """Load risks from DB (replaces JSON file)."""
    return db.list_risks()


def save_risks(data):
    """Backward-compat: bulk save a list of risks. Used by restore endpoint."""
    if isinstance(data, list):
        for r in data:
            if isinstance(r, dict) and r.get('id'):
                db.upsert_risk(r['id'], r)


@app.route('/api/risks')
def api_risks_list():
    """List all risks/issues, optionally filtered by phase_group"""
    phase_group = request.args.get('phase_group')
    risks = load_risks()
    if phase_group:
        risks = [r for r in risks if r.get('phase_group') == phase_group]
    risks.sort(key=lambda x: (
        {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}.get(x.get('severity'), 4),
        -int(str(x.get('updated_at', '')).replace('-', '').replace(':', '').replace('T', '').replace('.', '')[:14] or 0)
    ))
    return jsonify({'risks': risks, 'count': len(risks)})


@app.route('/api/risks/<rid>', methods=['DELETE'])
def api_risks_delete(rid):
    db.delete_risk(rid)
    return jsonify({'ok': True})

def load_services_overrides():
    """Load services overrides from DB.
    Returns: {service_key: {dept: {field: value}}}
    """
    # Stored as namespace='services', phase=<service_key>, key='<dept>.<field>'
    raw = db.get_namespace_overrides('services')
    out = {}
    for service_key, items in raw.items():
        if not service_key:
            continue
        out[service_key] = {}
        for combined, val in items.items():
            if '.' in combined:
                dept, field = combined.rsplit('.', 1)
                if dept not in out[service_key]:
                    out[service_key][dept] = {}
                out[service_key][dept][field] = val
    return out


def save_services_overrides(data):
    """Bulk save (used by restore endpoint).
    Format: {service_key: {dept: {field: value}}}
    """
    for service_key, dept_data in data.items():
        if not isinstance(dept_data, dict):
            continue
        for dept, fields in dept_data.items():
            if isinstance(fields, dict):
                for field, value in fields.items():
                    db.set_override('services', service_key, f"{dept}.{field}", value)


def get_odoo_parent_task_metadata(service_name):
    """For a given service name, find the matching parent task in Odoo and return:
    - earliest entry date (actual_start)
    - assignee (from task user_ids/user_id)
    - top contributor (most hours)
    """
    if not odoo.uid:
        if not odoo.connect():
            return {}
    try:
        # Search for task by name
        tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('name', '=', service_name), ('parent_id', '=', False)]],
            {'fields': ['id', 'name', 'user_id', 'date_start', 'date_end',
                        'date_deadline', 'stage_id', 'kanban_state'], 'limit': 5}
        )
        if not tasks:
            # Try ilike
            tasks = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.task', 'search_read',
                [[('name', 'ilike', service_name), ('parent_id', '=', False)]],
                {'fields': ['id', 'name', 'user_id', 'date_start', 'date_end',
                            'date_deadline', 'stage_id', 'kanban_state'], 'limit': 5}
            )
        if not tasks:
            return {}
        task = tasks[0]
        meta = {
            'task_id': task['id'],
            'task_name': task['name'],
        }
        if task.get('user_id') and isinstance(task['user_id'], list) and len(task['user_id']) > 1:
            meta['assignee'] = task['user_id'][1]
        if task.get('date_start'):
            meta['actual_start'] = str(task['date_start'])[:10]
        if task.get('date_end'):
            meta['actual_end'] = str(task['date_end'])[:10]
        if task.get('stage_id') and isinstance(task['stage_id'], list) and len(task['stage_id']) > 1:
            meta['odoo_stage'] = task['stage_id'][1]
        return meta
    except Exception as e:
        logger.warning(f"Odoo task lookup for {service_name}: {e}")
        return {}


@app.route('/api/services')
def api_services():
    """Services with department-level breakdown: baseline / planned / actuals / status per dept.
    Returns one row per (service, department) combination.
    """
    services = load_services()
    overrides = load_services_overrides()

    # Compute actuals from Odoo timesheets grouped by service (parent_task)
    data, _ = get_all_timesheets()
    actuals_by_service = {}
    for entry in data:
        s = entry.get('service', '')
        if not s:
            continue
        emp = entry.get('employee', 'Unknown')
        d = entry.get('date', '')
        h = entry.get('hours', 0)
        if s not in actuals_by_service:
            actuals_by_service[s] = {
                'hours': 0,
                'first_date': None,
                'last_date': None,
                'employees_hours': {},  # emp -> total hours
                'employees_dates': {},  # emp -> {first_date, last_date}
            }
        actuals_by_service[s]['hours'] += h
        if d:
            if not actuals_by_service[s]['first_date'] or d < actuals_by_service[s]['first_date']:
                actuals_by_service[s]['first_date'] = d
            if not actuals_by_service[s]['last_date'] or d > actuals_by_service[s]['last_date']:
                actuals_by_service[s]['last_date'] = d
        actuals_by_service[s]['employees_hours'][emp] = actuals_by_service[s]['employees_hours'].get(emp, 0) + h
        # Track each employee's first and last log date for this service
        if emp not in actuals_by_service[s]['employees_dates']:
            actuals_by_service[s]['employees_dates'][emp] = {'first': d, 'last': d}
        else:
            ed = actuals_by_service[s]['employees_dates'][emp]
            if d and (not ed['first'] or d < ed['first']):
                ed['first'] = d
            if d and (not ed['last'] or d > ed['last']):
                ed['last'] = d

    today = date.today().isoformat()

    # Build response: list of services with all dept data merged + odoo metadata
    result = []
    for s in services:
        name = s.get('اسم الخدمة المستقبلي', '') or ''
        if not name:
            continue

        # Match service name to actuals (exact or fuzzy)
        actual_entry = actuals_by_service.get(name)
        if not actual_entry and name:
            for k, v in actuals_by_service.items():
                if k and (name in k or k in name):
                    actual_entry = v
                    break

        actual_hours = actual_entry['hours'] if actual_entry else 0
        actual_days = round(actual_hours / WORK_HOURS_PER_DAY, 1)

        # Top contributor (fallback)
        top_contributor = None
        if actual_entry and actual_entry.get('employees_hours'):
            top_contributor = max(actual_entry['employees_hours'].items(), key=lambda x: x[1])[0]

        # Date-based assignation: list of people who logged time within planned date range
        # (or all contributors if dates are missing). Sorted by total hours descending.
        assignation_list = []
        if actual_entry and actual_entry.get('employees_hours'):
            planned_start = s.get('planned_start')
            planned_end = s.get('planned_end')
            emps_in_range = {}
            emps_outside = {}
            for emp, ed in actual_entry.get('employees_dates', {}).items():
                emp_first = ed.get('first')
                emp_last = ed.get('last')
                hours = actual_entry['employees_hours'].get(emp, 0)
                # Check if their work overlaps with planned window
                in_range = True
                if planned_start and emp_last and emp_last < planned_start:
                    in_range = False
                if planned_end and emp_first and emp_first > planned_end:
                    in_range = False
                if in_range:
                    emps_in_range[emp] = hours
                else:
                    emps_outside[emp] = hours
            # Prefer in-range, sorted by hours desc; fall back to outside
            sorted_in = sorted(emps_in_range.items(), key=lambda x: -x[1])
            sorted_out = sorted(emps_outside.items(), key=lambda x: -x[1])
            assignation_list = [e[0] for e in sorted_in] + [e[0] for e in sorted_out]

        auto_assignation = ', '.join(assignation_list[:3]) if assignation_list else ''

        # Try Odoo task metadata (assignee, dates, stage)
        odoo_meta = {}  # disabled by default — only fetch if needed (slow)
        # Uncomment to enable: odoo_meta = get_odoo_parent_task_metadata(name)

        # Override key — service ID is the Arabic name (stable)
        override_key = name
        sov = overrides.get(override_key, {})

        # Build per-dept rows
        baseline_by_dept = s.get('baseline_by_dept', {}) or {}

        # Prepare base service data
        service_obj = {
            'name': name,
            'planned_team': s.get('planned_team'),
            'planned_start': s.get('planned_start'),
            'planned_end': s.get('planned_end'),
            'planned_wd_roadmap': s.get('planned_wd_roadmap'),
            # Odoo-derived
            'actual_start': (actual_entry or {}).get('first_date') or odoo_meta.get('actual_start'),
            'actual_end_from_logs': (actual_entry or {}).get('last_date'),
            'odoo_assignee': odoo_meta.get('assignee'),
            'odoo_stage': odoo_meta.get('odoo_stage'),
            'top_contributor': top_contributor,
            'actuals_total_hours': actual_hours,
            'actuals_total_days': actual_days,
            # Departments breakdown
            'departments': {},
        }

        # For each department with non-zero baseline, build a row
        for dept_label in DEPT_LABELS.values():
            base_v = baseline_by_dept.get(dept_label)
            # New DB format: sov is already {dept: {field: val}}
            dept_override = sov.get(dept_label, {}) if isinstance(sov, dict) else {}

            # Read overrides
            planned = dept_override.get('planned')
            remaining = dept_override.get('remaining')
            status = dept_override.get('status')
            assignation = dept_override.get('assignation')

            # Auto-defaults if not overridden
            if status is None:
                # Compute auto status using actuals
                if base_v and actual_days >= base_v:
                    status = 'Done'
                elif s.get('planned_end') and s['planned_end'] < today and base_v and actual_days < base_v:
                    status = 'Overdue'
                elif s.get('planned_start') and s['planned_start'] <= today and base_v:
                    status = 'In Progress' if actual_days > 0 else 'Not Started'
                else:
                    status = 'Not Started'

            # Auto-default assignation: date-aware list (in-range contributors first)
            if not assignation:
                assignation = odoo_meta.get('assignee') or auto_assignation or top_contributor or ''

            # Auto-default remaining
            if remaining is None and base_v is not None:
                remaining = max(0, base_v - actual_days)

            service_obj['departments'][dept_label] = {
                'baseline': base_v,
                'planned': planned,
                'actuals_days': actual_days if base_v else None,
                'assignation': assignation,
                'remaining': remaining,
                'status': status,
                'has_baseline': base_v is not None and base_v > 0,
            }

        # Actual end: latest log date if all departments are 'Done'
        all_done = all(
            d['status'] == 'Done'
            for d in service_obj['departments'].values()
            if d['has_baseline']
        )
        service_obj['actual_end'] = service_obj['actual_end_from_logs'] if all_done else None

        result.append(service_obj)

    # Default sort by planned_start (roadmap order). Frontend will re-sort to push Done to bottom per-dept.
    result.sort(key=lambda x: (x.get('planned_start') or '9999-12-31', x.get('name') or ''))

    return jsonify({
        'services': result,
        'departments': list(DEPT_LABELS.values()),
        'today': today,
    })


@app.route('/api/services/override', methods=['POST'])
def api_services_override():
    """Save manual override for a (service, dept, field).
    Body: { service_name, department, field, value }
    Fields: planned, remaining, status, assignation
    """
    body = request.json or {}
    service_name = body.get('service_name')
    dept = body.get('department')
    field = body.get('field')
    value = body.get('value')

    if not service_name or not dept or not field:
        return jsonify({'error': 'service_name, department, and field required'}), 400

    valid_fields = {'planned', 'remaining', 'status', 'assignation'}
    if field not in valid_fields:
        return jsonify({'error': f'field must be one of {valid_fields}'}), 400

    # Cast value to correct type
    if value is not None and value != '':
        if field in ('planned', 'remaining'):
            try:
                value = float(value)
            except (ValueError, TypeError):
                value = None
        else:
            value = str(value)
    else:
        value = None

    # Save to DB: namespace='services', phase=<service_name>, key='<dept>.<field>'
    db.set_override('services', service_name, f"{dept}.{field}", value)
    return jsonify({'ok': True})


@app.route('/api/services/overrides', methods=['GET'])
def api_services_overrides_get():
    # Return in old format with 'departments' wrapper for backward compat
    raw = load_services_overrides()
    out = {}
    for service_key, dept_data in raw.items():
        out[service_key] = {'departments': dept_data}
    return jsonify(out)

@app.route('/api/phases')
def api_phases():
    """List of phases for the project (from Odoo or fallback)"""
    _proj_name = active_project_name()
    _proj_id   = session.get('project_id')
    _is_bog    = not _proj_id or str(_proj_id) == '228'

    if _is_bog:
        # BOG: use project.phase model
        phases = odoo.get_phases(project_name=_proj_name)
        if phases is None:
            return jsonify({'connected': False, 'is_bog': True, 'phases': [
                {'name': 'Consultation phase - Initiation'},
                {'name': 'Consultation phase - Analysis'},
                {'name': 'Consultation phase - General'},
                {'name': 'Consultation phase - UX'},
                {'name': 'Development Phase'},
            ], 'default': 'Development Phase'})
        return jsonify({'connected': True, 'is_bog': True,
                        'phases': [{'id': p['id'], 'name': p['name']} for p in phases],
                        'default': 'Development Phase'})
    else:
        # Non-BOG: use project.task.type (Kanban stages) — more reliable
        try:
            if not odoo.uid: odoo.connect()
            projects = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[('name', 'ilike', _proj_name)]],
                {'fields': ['id', 'name'], 'limit': 3}
            )
            if not projects:
                return jsonify({'connected': True, 'is_bog': False, 'phases': [], 'default': ''})
            pid = projects[0]['id']
            task_types = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.task.type', 'search_read',
                [[('project_ids', 'in', [pid])]],
                {'fields': ['id', 'name'], 'limit': 50}
            )
            default = task_types[0]['name'] if task_types else ''
            return jsonify({'connected': True, 'is_bog': False,
                            'phases': [{'id': t['id'], 'name': t['name']} for t in task_types],
                            'default': default})
        except Exception as e:
            return jsonify({'connected': False, 'is_bog': False, 'phases': [], 'default': '', 'error': str(e)})

def parse_phases_param(args):
    """Parse phases query param. Accepts 'phases=A,B,C' or 'phases=A&phases=B'."""
    phases_csv = args.get('phases')
    if phases_csv:
        return [p.strip() for p in phases_csv.split(',') if p.strip()]
    # Multi-value
    multi = args.getlist('phases')
    return [p for p in multi if p] if multi else None

@app.route('/api/timesheets/employees')
def api_ts_employees():
    """Aggregated by employee with optional phase filter"""
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    phases = parse_phases_param(request.args)

    data, connected = get_all_timesheets(date_from=date_from, date_to=date_to, phases=phases)

    df = pd.DataFrame(data) if data else pd.DataFrame()
    if df.empty:
        return jsonify({'connected': connected, 'employees': [], 'total_hours': 0,
                        'phases_filter': phases, 'date_from': date_from, 'date_to': date_to})

    by_emp = df.groupby('employee').agg(
        total_hours=('hours', 'sum'),
        days_logged=('date', 'nunique'),
        entries=('hours', 'count')
    ).reset_index().sort_values('total_hours', ascending=False)

    return jsonify({
        'connected': connected,
        'phases_filter': phases,
        'employees': [{
            'name': r['employee'],
            'total_hours': float(r['total_hours']),
            'days_logged': int(r['days_logged']),
            'entries': int(r['entries']),
        } for _, r in by_emp.iterrows()],
        'total_hours': float(df['hours'].sum()),
        'date_from': date_from,
        'date_to': date_to,
    })

@app.route('/api/timesheets/employee/<name>')
def api_ts_employee_detail(name):
    """Drill-down: days only by default, tasks via expand"""
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    phases = parse_phases_param(request.args)

    data, _ = get_all_timesheets(date_from=date_from, date_to=date_to, phases=phases)
    data = [d for d in data if d.get('employee') == name]

    by_date = {}
    for entry in data:
        d = entry.get('date')
        if d not in by_date:
            by_date[d] = {'date': d, 'total_hours': 0, 'tasks': []}
        by_date[d]['total_hours'] += entry.get('hours', 0)
        by_date[d]['tasks'].append({
            'task': entry.get('task'),
            'description': entry.get('description'),
            'hours': entry.get('hours'),
            'service': entry.get('service', ''),
            'phase': entry.get('phase', ''),
        })
    days = sorted(by_date.values(), key=lambda x: x['date'], reverse=True)

    return jsonify({
        'employee': name,
        'total_hours': sum(e.get('hours', 0) for e in data),
        'total_days': len(by_date),
        'days': days,
    })

@app.route('/api/timesheets/export')
def api_ts_export():
    """CSV export of timesheets"""
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    phases = parse_phases_param(request.args)

    data, _ = get_all_timesheets(date_from=date_from, date_to=date_to, phases=phases)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Employee', 'Project', 'Phase', 'Service', 'Task', 'Description', 'Hours'])
    for d in data:
        writer.writerow([
            d.get('date', ''), d.get('employee', ''), d.get('project', ''),
            d.get('phase', ''), d.get('service', ''), d.get('task', ''),
            d.get('description', ''), d.get('hours', 0)
        ])

    csv_content = output.getvalue()
    output.close()
    filename = f"timesheets_{date.today().isoformat()}.csv"
    return Response(
        '\ufeff' + csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.route('/api/missing-hours')
def api_missing_hours():
    """Missing hours: last 30 days default, skip if logged on other projects"""
    date_to = request.args.get('to') or date.today().isoformat()
    date_from = request.args.get('from')

    if not date_from:
        # Default: 30 days back
        from_date = (date.today() - timedelta(days=30))
        date_from = from_date.isoformat()

    # Get THIS project entries
    project_data, connected = get_all_timesheets(date_from=date_from, date_to=date_to, all_projects=False)
    # Get ALL projects entries (for cross-project check)
    all_data, _ = get_all_timesheets(date_from=date_from, date_to=date_to, all_projects=True)

    if not project_data and not all_data:
        return jsonify({'connected': connected, 'employees': [], 'date_from': date_from, 'date_to': date_to})

    # Hours per employee per date — across ALL projects
    all_hours_by_emp_date = {}  # {emp: {date: hours}}
    for e in all_data:
        emp = e.get('employee')
        d = e.get('date')
        all_hours_by_emp_date.setdefault(emp, {})
        all_hours_by_emp_date[emp][d] = all_hours_by_emp_date[emp].get(d, 0) + e.get('hours', 0)

    # Hours per employee per date — THIS project only
    proj_hours_by_emp_date = {}
    for e in project_data:
        emp = e.get('employee')
        d = e.get('date')
        proj_hours_by_emp_date.setdefault(emp, {})
        proj_hours_by_emp_date[emp][d] = proj_hours_by_emp_date[emp].get(d, 0) + e.get('hours', 0)

    # Working days in range
    working_days = get_working_days_between(date_from, date_to)

    employees_summary = []
    # Only employees who appeared in THIS project at least once
    employees_in_project = set(proj_hours_by_emp_date.keys())

    for emp_name in employees_in_project:
        proj_hours = proj_hours_by_emp_date.get(emp_name, {})
        all_hours = all_hours_by_emp_date.get(emp_name, {})

        missing_dates_detail = []
        underlogged_dates_detail = []
        total_missing = 0

        for d in working_days:
            proj_h = proj_hours.get(d, 0)
            all_h = all_hours.get(d, 0)

            if all_h >= WORK_HOURS_PER_DAY:
                # full day logged somewhere — not missing
                continue
            elif proj_h == 0 and all_h == 0:
                # nothing logged at all
                missing_dates_detail.append({'date': d, 'logged_hrs': 0, 'missing_hrs': WORK_HOURS_PER_DAY,
                                             'reason': 'No entries'})
                total_missing += WORK_HOURS_PER_DAY
            elif all_h < WORK_HOURS_PER_DAY:
                # partial day
                missing = WORK_HOURS_PER_DAY - all_h
                underlogged_dates_detail.append({'date': d, 'logged_hrs': all_h, 'missing_hrs': missing,
                                                 'reason': f'Only {all_h}h logged total'})
                total_missing += missing

        total_logged = sum(proj_hours.values())
        total_logged_all = sum(all_hours.values())
        expected_total = len(working_days) * WORK_HOURS_PER_DAY

        employees_summary.append({
            'name': emp_name,
            'expected_days': len(working_days),
            'logged_days_project': len(proj_hours),
            'logged_days_total': len(all_hours),
            'missing_days_count': len(missing_dates_detail),
            'underlogged_days_count': len(underlogged_dates_detail),
            'logged_hours_project': float(total_logged),
            'logged_hours_total': float(total_logged_all),
            'expected_hours': float(expected_total),
            'missing_hours': float(total_missing),
            'compliance_pct': float((expected_total - total_missing) / expected_total * 100) if expected_total else 100,
            'missing_dates_detail': missing_dates_detail,
            'underlogged_dates_detail': underlogged_dates_detail,
        })

    employees_summary.sort(key=lambda x: x['missing_hours'], reverse=True)

    return jsonify({
        'connected': connected,
        'date_from': date_from,
        'date_to': date_to,
        'work_hours_per_day': WORK_HOURS_PER_DAY,
        'employees': employees_summary,
    })

@app.route('/api/roadmap')
def api_roadmap():
    team1 = [s for s in ROADMAP if s['team'] == 'الفريق 1']
    team2 = [s for s in ROADMAP if s['team'] == 'الفريق 2']
    admin = [s for s in ROADMAP if s['team'] == 'الإدارة']
    return jsonify({
        'project_info': PROJECT_INFO,
        'milestones': MILESTONES,
        'services': ROADMAP,
        'team_breakdown': {
            'team_1': {'name': 'الفريق الأول', 'count': len(team1), 'total_wd': sum(s.get('wd') or 0 for s in team1)},
            'team_2': {'name': 'الفريق الثاني', 'count': len(team2), 'total_wd': sum(s.get('wd') or 0 for s in team2)},
            'admin': {'name': 'نقل البيانات', 'count': len(admin)},
        }
    })

@app.route('/api/budget')
def api_budget():
    return jsonify({
        'project_name': PROJECT_NAME,
        'pm': 'Abdelrahman Doghish',
        'support_start': '2027-05-18',
        'support_end': '2028-05-17',
        'total_mandays': 6556,
        'approved': {'cost_usd': 2735436.00, 'cost_sar': 10257885.00, 'revenue_sar': 20570344.00,
                     'profit_sar': 10312459.00, 'profit_pct': 50},
        'final': {'cost_sar': 10257885.00, 'revenue_sar': 20150344.00, 'profit_sar': 9892459.00, 'profit_pct': 49},
        'changes': [{'reason': 'Third party license', 'plan_id': '', 'changes_cost': 0, 'changes_revenue': -420000.00}],
        'total_change_cost': 0,
        'total_change_revenue': -420000.00
    })

# ============================================================
# VARIANCE — reads variance.xlsx
# ============================================================
VARIANCE_FILE = os.path.join(BASE_DIR, 'data', 'variance.xlsx')  # Read-only Excel from repo
TRAVEL_FILE = os.path.join(PERSIST_DIR, 'travel.json')
BUDGET_OVERRIDES_FILE = os.path.join(PERSIST_DIR, 'budget_overrides.json')

def load_budget_overrides():
    """Load all budget overrides from DB.
    Returns: {phase: {field_path: value}}
    """
    return db.get_namespace_overrides('budget')


def save_budget_override(phase, path, value):
    """Save (or delete if value is None) a single budget override."""
    proj_set_override('budget', phase, path, value)


def apply_budget_overrides(budget_info, phase_key):
    """Apply saved overrides on top of the parsed budget data."""
    pfx = active_db_prefix()
    full_ns = f'{pfx}_budget' if pfx else 'budget'
    overrides = db.get_namespace_overrides(full_ns, phase_key)
    if not overrides:
        return budget_info
    for path, value in overrides.items():
        # Path format: "section.field" e.g. "approved.cost_sar"
        parts = path.split('.')
        if len(parts) != 2:
            continue
        section, field = parts
        if section in budget_info and isinstance(budget_info[section], dict):
            budget_info[section][field] = value
        elif section == 'contract' and 'contract' in budget_info:
            for k in budget_info['contract']:
                if k == field:
                    if isinstance(budget_info['contract'][k], dict):
                        budget_info['contract'][k]['value'] = value
                    else:
                        budget_info['contract'][k] = value
                    break
    budget_info['_has_overrides'] = bool(overrides)
    budget_info['_override_keys'] = list(overrides.keys())
    return budget_info


@app.route('/api/variance/budget-override', methods=['POST'])
def api_budget_override_save():
    """Save a single budget field override.
    Body: { phase: 'development', path: 'approved.cost_sar', value: 12000000 }
    """
    body = request.json or {}
    phase = body.get('phase')
    path = body.get('path')
    value = body.get('value')

    if not phase or not path:
        return jsonify({'error': 'phase and path required'}), 400

    # Try to convert to float for numeric values
    if value is not None and value != '':
        try:
            value = float(value)
        except (ValueError, TypeError):
            pass  # keep as string

    save_budget_override(phase, path, value)
    return jsonify({'ok': True, 'phase': phase, 'path': path, 'value': value})


@app.route('/api/variance/budget-override/<phase>')
def api_budget_overrides_get(phase):
    """Get all overrides for a phase."""
    overrides = proj_get_namespace('budget', phase)
    return jsonify({'phase': phase, 'overrides': overrides})


@app.route('/api/overview/financials')
def api_overview_financials():
    """Returns Final Budget revenue per phase for the Overview summary.
    Revenue = approved.revenue_sar + sum of delta_rev from budget_changes.
    Uses same logic as budgetAutoCalc in variance.js.
    """
    _proj_id_fin = session.get('project_id')
    _is_bog_fin  = not _proj_id_fin or str(_proj_id_fin) == '228'
    phases = ['development', 'consultation', 'support'] if _is_bog_fin else ['services', 'support']
    result = {}
    total_rev = 0.0
    for phase in phases:
        try:
            # Strategy 1: direct override key
            overrides = proj_get_namespace('budget', phase) or {}
            rev_val = 0.0

            # approved.revenue_sar may be stored as a dotted key
            raw = overrides.get('approved.revenue_sar')
            if raw is not None:
                try: rev_val = float(raw)
                except (TypeError, ValueError): pass

            # Strategy 2: nested dict {'approved': {'revenue_sar': ...}}
            if not rev_val:
                approved_block = overrides.get('approved') or {}
                if isinstance(approved_block, dict):
                    r2 = approved_block.get('revenue_sar')
                    if r2 is not None:
                        try: rev_val = float(r2)
                        except: pass

            # Strategy 3: read directly with get_override
            if not rev_val:
                direct = proj_get_override('budget', phase, 'approved.revenue_sar')
                if direct is not None:
                    try: rev_val = float(direct)
                    except: pass

            # Strategy 4: fallback — if no revenue saved yet, note it for frontend
            # (frontend will show '—' which is correct — revenue not yet entered)

            # Add Δ Revenue from budget changes (Final Budget = Approved + Changes)
            changes_raw = proj_get_override('budget_changes', '', phase)
            if isinstance(changes_raw, list):
                for ch in changes_raw:
                    try: rev_val += float(ch.get('delta_rev') or 0)
                    except: pass

            result[phase] = round(rev_val, 2)
            total_rev += rev_val
            logger.info(f"Financials [{phase}]: rev={rev_val} overrides_keys={list(overrides.keys())[:5]}")
        except Exception as _e:
            logger.error(f"Financials [{phase}] error: {_e}")
            result[phase] = 0.0
    result['total'] = round(total_rev, 2)
    return jsonify(result)


@app.route('/debug/financials')
def debug_financials():
    """Debug: show raw budget override data for each phase."""
    out = {}
    for phase in ['development', 'consultation', 'support']:
        overrides = db.get_namespace_overrides('budget', phase) or {}
        direct    = db.get_override('budget', phase, 'approved.revenue_sar')
        changes   = db.get_override('budget_changes', '', phase) or []
        out[phase] = {
            'namespace_overrides': overrides,
            'direct_get': direct,
            'budget_changes_count': len(changes) if isinstance(changes, list) else 0,
            'budget_changes': changes if isinstance(changes, list) else [],
        }
    return jsonify(out)


@app.route('/api/variance/budget-override/<phase>/<path:path>', methods=['DELETE'])
def api_budget_override_delete(phase, path):
    """Delete a specific override."""
    proj_set_override('budget', phase, path, None)
    return jsonify({'ok': True})

# Sub-tab definitions: which sheets feed which tab
VARIANCE_TABS = {
    'development': {
        'label': 'Development',
        'sections': [
            {'key': 'budget', 'label': 'Budget', 'sheet': 'Budget - Development', 'parser': 'budget'},
            {'key': 'profitability', 'label': 'Profitability', 'sheet': 'Profitability - Development', 'parser': 'profitability'},
            {'key': 'effort', 'label': 'Current Effort', 'sheet': 'Current Effort - Development', 'parser': 'effort'},
            {'key': 'estimated', 'label': 'Estimated Cost', 'sheet': 'Estimated Cost - Development', 'parser': 'estimated'},
        ]
    },
    'consultation': {
        'label': 'Consultation',
        'sections': [
            {'key': 'budget', 'label': 'Budget', 'sheet': 'Budget - Consultation', 'parser': 'budget'},
            {'key': 'profitability', 'label': 'Profitability', 'sheet': 'Profitability - Consultation', 'parser': 'profitability'},
            {'key': 'effort', 'label': 'Current Effort', 'sheet': 'Current Effort - Consultation', 'parser': 'effort'},
            {'key': 'estimated', 'label': 'Estimated Cost', 'sheet': 'Estimated Cost - Consultation', 'parser': 'estimated'},
        ]
    },
    'support': {
        'label': 'Support',
        'sections': [
            {'key': 'budget', 'label': 'Budget', 'sheet': 'Budget - Support', 'parser': 'budget'},
            {'key': 'estimated', 'label': 'Estimated Cost', 'sheet': 'Estimated Cost - Support', 'parser': 'estimated'},
        ]
    },
    'travel': {
        'label': 'Travel & Onsite',
        'sections': [
            {'key': 'travel', 'label': 'Travel Records', 'parser': 'travel'},
        ]
    },
}

def safe_val(v):
    """Convert pandas value to JSON-safe value"""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.isoformat()
    if isinstance(v, (int, float)):
        return v
    return str(v)

def parse_budget_sheet(df):
    """Parse a Budget sheet (Development/Consultation/Support)"""
    info = {}
    # Column 2 = label, column 3-4 = values
    label_col, val_col, val_col2 = 2, 3, 4
    rows = df.values.tolist()
    info['contract'] = {}
    contract_keys = ['Project Name', 'Client', 'Contract start date', 'Contract end date',
                     'Contract duration', 'Contract Type', 'Scope', 'Support start and end dates', 'Progress']
    for r in rows:
        if len(r) > val_col and pd.notna(r[label_col]) and r[label_col] in contract_keys:
            v1 = safe_val(r[val_col]) if len(r) > val_col else None
            v2 = safe_val(r[val_col2]) if len(r) > val_col2 else None
            info['contract'][r[label_col]] = {'value': v1, 'value2': v2}

    # Approved budget block (rows 11-18)
    info['approved'] = {}
    info['final'] = {}
    for r in rows:
        if not r or len(r) < 5:
            continue
        lbl = r[label_col] if pd.notna(r[label_col]) else None
        if lbl == 'Total Mandays':
            info['approved']['total_mandays'] = safe_val(r[4])
        elif lbl == 'Total Estimated Cost ($)':
            info['approved']['cost_usd'] = safe_val(r[4])
        elif lbl == 'Total Estimated Cost (SAR)':
            info['approved']['cost_sar'] = safe_val(r[4])
        elif lbl == 'Total Revenue':
            info['approved']['revenue_sar'] = safe_val(r[4])
        elif lbl == 'Budget Profit (SAR)' and 'profit_sar' not in info['approved']:
            info['approved']['profit_sar'] = safe_val(r[4])
        elif lbl == 'Budget Profit (%)' and 'profit_pct' not in info['approved']:
            info['approved']['profit_pct'] = safe_val(r[4])
        # Final budget on right side (col 6 label, col 8 value)
        if len(r) > 8 and pd.notna(r[6]):
            rlabel = r[6]
            rval = safe_val(r[8])
            if rlabel == 'Total Cost after changes':
                info['final']['cost_sar'] = rval
            elif rlabel == 'Total Revenue after changes':
                info['final']['revenue_sar'] = rval
            elif rlabel == 'Budget Profit (SAR)':
                info['final']['profit_sar'] = rval
            elif rlabel == 'Budget Profit (%)':
                info['final']['profit_pct'] = rval
            elif rlabel == 'Total Changes on Budget':
                info['final']['total_change_cost'] = safe_val(r[8])
                if len(r) > 9:
                    info['final']['total_change_revenue'] = safe_val(r[9])

    # Changes log (rows 3-10ish, cols 6,7,8,9)
    changes = []
    for r in rows[3:11]:
        if len(r) > 9 and pd.notna(r[6]) and r[6] != 'Reason':
            changes.append({
                'reason': safe_val(r[6]),
                'plan_id': safe_val(r[7]) if len(r) > 7 else None,
                'changes_cost': safe_val(r[8]) if len(r) > 8 else None,
                'changes_revenue': safe_val(r[9]) if len(r) > 9 else None,
            })
    info['changes'] = changes

    return info

def parse_profitability_sheet(df):
    """Parse Profitability sheet — month-by-month variance metrics"""
    rows = df.values.tolist()
    # Header at row 5 (0-indexed)
    if len(rows) < 6:
        return {'months': [], 'columns': []}

    headers = []
    for v in rows[5]:
        h = str(v) if pd.notna(v) else ''
        headers.append(h.strip())

    months = []
    for r in rows[7:]:  # data starts at row 7
        if not r or len(r) == 0:
            continue
        if pd.isna(r[0]):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if h and i < len(r):
                row_data[h] = safe_val(r[i])
        months.append(row_data)

    return {'columns': headers, 'months': months}

def parse_effort_sheet(df):
    """Parse Current Effort sheet — team monthly hours"""
    rows = df.values.tolist()
    if len(rows) < 5:
        return {'team': [], 'months': [], 'totals': {}}

    # Row 3: month names; Row 4: column headers
    months_row = rows[3] if len(rows) > 3 else []
    headers = rows[4] if len(rows) > 4 else []

    # Find month columns (skip first 6 cols which are #, Name, Position, Hour Rate, Overtime Rate)
    month_blocks = []
    cur_month = None
    for i, m in enumerate(months_row):
        if pd.notna(m):
            cur_month = str(m).strip()
        if cur_month and i >= 6:
            month_blocks.append({'month': cur_month, 'col': i})

    # Group by month (every 3 cols = Regular, Ramadan, Overtime)
    seen_months = []
    last_month = None
    for b in month_blocks:
        if b['month'] != last_month:
            seen_months.append(b['month'])
            last_month = b['month']

    team = []
    for r in rows[5:]:
        if len(r) < 4 or pd.isna(r[1]):
            continue
        # Stop at totals row
        if isinstance(r[1], str) and 'total' in str(r[1]).lower():
            continue
        if pd.isna(r[0]):
            # could be totals/summary row
            if len(r) > 5 and pd.notna(r[5]) and isinstance(r[5], str) and 'cost' in r[5].lower():
                continue
            if pd.isna(r[1]):
                continue
        member = {
            'num': safe_val(r[0]) if len(r) > 0 else None,
            'name': safe_val(r[1]) if len(r) > 1 else None,
            'position': safe_val(r[3]) if len(r) > 3 else None,
            'hour_rate': safe_val(r[4]) if len(r) > 4 else None,
            'overtime_rate': safe_val(r[5]) if len(r) > 5 else None,
            'monthly': [],
            'total_cost': safe_val(r[39]) if len(r) > 39 else None,
            'current_mds': safe_val(r[40]) if len(r) > 40 else None,
        }
        # 11 months × 3 cols starting at col 6
        for m_idx, month in enumerate(seen_months):
            base = 6 + m_idx * 3
            if base + 2 < len(r):
                member['monthly'].append({
                    'month': month,
                    'regular': safe_val(r[base]),
                    'ramadan': safe_val(r[base + 1]),
                    'overtime': safe_val(r[base + 2]),
                })
        team.append(member)

    return {'team': team, 'months': seen_months}

def parse_estimated_sheet(df):
    """Parse Estimated Cost sheet"""
    rows = df.values.tolist()
    # Header typically at row 5
    if len(rows) < 6:
        return {'positions': [], 'columns': []}

    # Find header row
    header_row = None
    for i, r in enumerate(rows[:10]):
        if r and pd.notna(r[0]) and 'Position' in str(r[0]):
            header_row = i
            break
    if header_row is None:
        header_row = 5

    headers = [str(h) if pd.notna(h) else f'col_{i}' for i, h in enumerate(rows[header_row])]
    positions = []
    for r in rows[header_row + 1:]:
        if not r or pd.isna(r[0]):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i < len(r):
                row_data[h] = safe_val(r[i])
        if row_data.get(headers[0]):
            positions.append(row_data)
    return {'columns': headers, 'positions': positions}

# ============================================================
# COMPUTED EFFORT FROM ODOO TIMESHEETS
# ============================================================

# Phase mapping: variance tab → list of Odoo phases
# Default PHASE_MAPPING for BOG (id=228)
# For other projects, phases are auto-detected from Odoo or set via /manage
PHASE_MAPPING_DEFAULT = {
    'development':  ['Development Phase'],
    'consultation': ['Consultation phase - Initiation', 'Consultation phase -  Analysis',
                     'Consultation phase - General', 'Consultation phase -  UX'],
    'support':      [],
}

def get_phase_mapping():
    """Get phase mapping for active project.
    Priority:
    1. Per-project override saved in DB
    2. BOG default if project_id == 228 or no project set
    3. Empty (auto-detect per API call)
    """
    from flask import session as _sess
    proj_id = _sess.get('project_id')
    override = db.get_override('phase_mapping', str(proj_id or ''), 'mapping')
    if override and isinstance(override, dict):
        return override
    if not proj_id or str(proj_id) == '228':
        return PHASE_MAPPING_DEFAULT
    return {'services': [], 'support': []}


SUPPORT_KWS = ['support', 'operation', 'maintenance', 'hypercare', 'production', 'production activities', 'دعم', 'الدعم', 'دعم فني', 'تشغيل']

def auto_detect_phases_for_project(project_id, phase_key):
    """For non-BOG projects: tasks have no phase_id, only stage_id (Kanban).
    The Services/Support split comes from the Excel config, NOT from Odoo.
    So we return ALL stages for both keys — the split is handled at the project level.
    Returns list of stage_id values (integers) for direct filtering.
    """
    try:
        if not odoo.uid: odoo.connect()
        all_stages = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task.type', 'search_read',
            [[('project_ids', 'in', [project_id])]],
            {'fields': ['id', 'name'], 'limit': 50}
        )
        # Return all stage names — Services/Support split is Excel-config-based
        return [s['name'] for s in all_stages]
    except Exception as _e:
        logger.warning(f"auto_detect_phases: {_e}")
        return []


def get_all_stage_ids_for_project(project_id):
    """Get all task.type IDs for a project."""
    try:
        if not odoo.uid: odoo.connect()
        stages = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task.type', 'search_read',
            [[('project_ids', 'in', [project_id])]],
            {'fields': ['id', 'name'], 'limit': 50}
        )
        return stages
    except Exception:
        return []


def get_project_odoo_id(proj_name):
    """Get Odoo project ID by name."""
    try:
        if not odoo.uid: odoo.connect()
        projs = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', proj_name)]],
            {'fields': ['id', 'name'], 'limit': 3}
        )
        return projs[0]['id'] if projs else None
    except Exception:
        return None

PHASE_MAPPING = PHASE_MAPPING_DEFAULT  # keep for backward compat

# Ramadan dates (current year). Update annually.
RAMADAN_RANGES = {
    'KSA': {'start': '2026-02-18', 'end': '2026-03-19'},  # Saudi
    'EGY': {'start': '2026-02-19', 'end': '2026-03-20'},  # Egypt
}

# Tunisia weekend exception
TUNIS_WEEKEND = [5, 6]  # Saturday, Sunday (Mon=0)
DEFAULT_WEEKEND = [4, 5]  # Friday, Saturday
RAMADAN_HOURS = 6
NORMAL_HOURS = 8

# ── Public Holidays (overtime if logged on these days) ──
# Update annually or via env/DB in the future
PUBLIC_HOLIDAYS = {
    'EGY': {
        "2026-01-07", "2026-01-29", "2026-03-19", "2026-03-20", "2026-03-21",
        "2026-03-22", "2026-03-23", "2026-04-13", "2026-04-25", "2026-05-01",
        "2026-06-30", "2026-07-23", "2026-10-06",
    },
    'KSA': {
        "2026-02-22", "2026-03-19", "2026-03-20", "2026-03-21", "2026-03-22",
        "2026-03-23", "2026-03-24", "2026-05-26", "2026-05-27", "2026-05-28",
        "2026-05-29", "2026-09-23",
    },
    'TUN': {
        "2026-01-01", "2026-03-20", "2026-03-21", "2026-03-22", "2026-05-01",
        "2026-05-27", "2026-05-28", "2026-06-16", "2026-07-25", "2026-08-13",
        "2026-08-25", "2026-10-15", "2026-12-17",
    },
}

def is_public_holiday(date_str, country):
    """Returns True if the date is a public holiday for the given country."""
    return date_str in PUBLIC_HOLIDAYS.get(country, set())

def get_country_from_employee_name(name):
    """Detect country from Odoo employee code prefix.
    Format: '[E109] Basem Mohamed' → EGY
            '[R323] Talal Abdulwahed' → KSA
            '[T...] Name' → TUN (verify by name as well)
    """
    if not name:
        return 'EGY'
    import re
    # Match [X###] or [X## ] at the start (allow optional trailing space inside brackets)
    m = re.match(r'^\s*\[([A-Z])(\d+)\s*\]', str(name).strip())
    if m:
        letter = m.group(1).upper()
        if letter == 'E':
            return 'EGY'
        elif letter == 'R':
            return 'KSA'
        elif letter == 'T':
            return 'TUN'
    # Fallback: check name for hints
    n = str(name).upper()
    if 'TUNIS' in n or 'TUN ' in n:
        return 'TUN'
    if 'SAUDI' in n or 'KSA' in n:
        return 'KSA'
    return 'EGY'  # default

def get_country_from_position(position_name):
    """Backward compat: detect country from position string"""
    if not position_name:
        return 'EGY'
    pn = str(position_name).upper()
    if 'KSA' in pn or 'SAUDI' in pn:
        return 'KSA'
    if 'TUN' in pn or 'TUNIS' in pn:
        return 'TUN'
    return 'EGY'

def is_in_ramadan(date_str, country):
    """Check if a date falls within Ramadan for the given country"""
    rng = RAMADAN_RANGES.get(country, RAMADAN_RANGES['EGY'])
    return rng['start'] <= date_str <= rng['end']

def get_weekend_for_country(country):
    """Tunis: Sat+Sun. Others: Fri+Sat"""
    if country == 'TUN':
        return TUNIS_WEEKEND
    return DEFAULT_WEEKEND

def _names_match(a, b):
    """Fuzzy name match — handles [E259] prefix, hyphens, spacing differences."""
    import re as _re3
    def clean(s):
        s = (_re3.sub(r'\[[A-Z]\d+\]\s*', '', s or '')).strip()
        return s.lower().replace('-',' ').replace('_',' ').replace("'",'')
    a = clean(a); b = clean(b)
    if not a or not b: return False
    if a == b: return True
    if a.replace(' ','') == b.replace(' ',''): return True
    aw = a.split(); bw = b.split()
    if len(aw) >= 2 and len(bw) >= 2 and aw[0]==bw[0] and aw[-1]==bw[-1]: return True
    a_ns = a.replace(' ',''); b_ns = b.replace(' ','')
    if a_ns in b_ns or b_ns in a_ns: return True
    return False


def _extract_emp_code(name):
    """Extract employee code like E259, R210, T123 from name string."""
    import re as _rec
    m = _rec.search(r'\[([ERT]\d+)\]', name or '')
    return m.group(1).upper() if m else None


def get_travel_periods_for_employee(name, odoo_employee_id=None):
    """Returns list of (start_date, end_date_or_None) tuples.
    Match priority:
      1. Odoo employee ID (most reliable — set by auto-link or manual link)
      2. Employee code match (E259 in name vs odoo_employee_code in travel record)
      3. Fuzzy name match (fallback)
    """
    records = load_travel()
    periods = []
    emp_code = _extract_emp_code(name)  # extract from "[E259] Sara Samir"

    for r in records:
        matched = False

        # 1. Odoo ID match
        if odoo_employee_id and r.get('odoo_employee_id'):
            try:
                matched = int(r['odoo_employee_id']) == int(odoo_employee_id)
            except (ValueError, TypeError):
                pass

        # 2. Employee code match (E259 == E259)
        if not matched and emp_code and r.get('odoo_employee_code'):
            matched = emp_code.upper() == str(r['odoo_employee_code']).upper().strip()

        # 3. Fuzzy name match
        if not matched:
            # Strip [E259] prefix from timesheet name before comparing
            clean_name = _names_match.__code__ and name  # just use name
            import re as _re4
            clean_name = _re4.sub(r'\[[ERT]\d+\]\s*', '', name or '').strip()
            matched = _names_match(r.get('name', ''), clean_name)

        if matched:
            periods.append((r.get('start_date'), r.get('end_date')))
    return periods

def is_onsite_on_date(name, date_str, odoo_employee_id=None):
    """Check if employee was onsite on a given date"""
    periods = get_travel_periods_for_employee(name, odoo_employee_id=odoo_employee_id)
    for start, end in periods:
        if not start:
            continue
        if date_str < start:
            continue
        if end is None or date_str <= end:
            return True
    return False

def get_position_rates(positions_list):
    """Build dict {position_name: {hour_rate, md_rate}} from positions list"""
    rate_map = {}
    for p in positions_list:
        if p.get('name'):
            rate_map[p['name'].strip()] = {
                'hour_rate': p.get('hour_rate'),
                'md_rate': p.get('md_rate'),
            }
    return rate_map

def compute_effort_from_odoo(phase_key, year, month, position_lookup=None):
    """Compute Regular / Ramadan / Overtime hours per person for a specific month.
    Now also tracks onsite days based on travel records and computes effective rate.
    """
    phases = get_phase_mapping().get(phase_key, [])
    if not phases and phase_key != 'support':
        return {'team': [], 'months': [], 'error': f'No phases mapped for {phase_key}'}

    month_start = date(year, month, 1).isoformat()
    if month == 12:
        next_month_start = date(year + 1, 1, 1)
    else:
        next_month_start = date(year, month + 1, 1)
    month_end = (next_month_start - timedelta(days=1)).isoformat()

    raw = odoo.get_timesheets(
        date_from=month_start,
        date_to=month_end,
        phase_filter=phases if phases else None
    )
    if raw is None:
        return {'team': [], 'months': [date(year, month, 1).strftime('%B')], 'error': 'Odoo unreachable'}

    entries = [normalize_timesheet(e) for e in raw]

    # Load positions + rates from Excel
    positions_list = get_positions_from_excel()
    rate_map = get_position_rates(positions_list)
    tunis_rates = get_tunis_rates_from_excel()

    # Group by employee, then by date
    by_emp = {}
    emp_odoo_ids = {}  # emp_name -> odoo_employee_id
    for entry in entries:
        emp = entry['employee']
        d = entry['date']
        h = entry['hours']
        if emp not in by_emp:
            by_emp[emp] = {}
        if d not in by_emp[emp]:
            by_emp[emp][d] = 0
        by_emp[emp][d] += h
        if entry.get('odoo_employee_id') and emp not in emp_odoo_ids:
            emp_odoo_ids[emp] = entry['odoo_employee_id']

    team = []
    for emp_name, day_hours in by_emp.items():
        # Country from Odoo employee code prefix [E###], [R###], [T###]
        country = get_country_from_employee_name(emp_name)
        weekend_days = get_weekend_for_country(country)

        # Base role from Odoo position lookup (e.g. "Senior Business Analyst")
        # We strip any country prefix or onsite suffix to get the bare role
        odoo_role = (position_lookup or {}).get(emp_name) or ''

        regular_mh = 0
        ramadan_mh = 0
        overtime_mh = 0
        onsite_hours = 0
        non_onsite_hours = 0
        onsite_days = 0
        total_days = 0

        for day_str, total_h in day_hours.items():
            try:
                d_obj = datetime.strptime(day_str, '%Y-%m-%d').date()
            except Exception:
                continue
            wd = d_obj.weekday()
            is_weekend = wd in weekend_days
            in_ramadan = is_in_ramadan(day_str, country)
            is_onsite = is_onsite_on_date(emp_name, day_str, odoo_employee_id=emp_odoo_ids.get(emp_name))
            is_holiday = is_public_holiday(day_str, country)

            total_days += 1
            if is_onsite:
                onsite_days += 1
                onsite_hours += total_h
            else:
                non_onsite_hours += total_h

            if is_weekend or is_holiday:
                overtime_mh += total_h
            else:
                expected = RAMADAN_HOURS if in_ramadan else NORMAL_HOURS
                if total_h <= expected:
                    if in_ramadan:
                        ramadan_mh += total_h
                    else:
                        regular_mh += total_h
                else:
                    if in_ramadan:
                        ramadan_mh += expected
                    else:
                        regular_mh += expected
                    overtime_mh += (total_h - expected)

        total_h = regular_mh + ramadan_mh + overtime_mh
        total_md = round(total_h / NORMAL_HOURS, 2)

        # Position resolution depends on country:
        # - For TUN: rate is per-name from "Tunis Rates" sheet (no position concept)
        # - For EGY/KSA: position = "{COUNTRY} - {role}" or "+ - onsite" if any onsite hours
        if country == 'TUN':
            tunis_match = find_tunis_rate(emp_name, tunis_rates)
            if tunis_match:
                base_position = f"TUN - {odoo_role}".strip(' -') if odoo_role else 'TUN'
                base_rate_info = {
                    'hour_rate': tunis_match['hour_rate'],
                    'md_rate': tunis_match['md_rate'],
                }
                onsite_rate_info = None  # Tunis has no onsite variants
                country_position = base_position
                onsite_position = ''
            else:
                base_position = f"TUN - {odoo_role}".strip(' -') if odoo_role else 'TUN'
                base_rate_info = None
                onsite_rate_info = None
                country_position = base_position
                onsite_position = ''
        else:
            country_position = f"{country} - {odoo_role}".strip(' -') if odoo_role else ''
            onsite_position = f"{country} - {odoo_role} - onsite".strip(' -') if odoo_role else ''

            def _find_position_match(target):
                if not target:
                    return None
                if target in rate_map:
                    return target, rate_map[target]
                target_lower = target.lower()
                for k, v in rate_map.items():
                    if k.lower() == target_lower:
                        return k, v
                target_clean = target_lower.replace('.', '').replace('-', ' ').replace('  ', ' ').strip()
                for k, v in rate_map.items():
                    k_clean = k.lower().replace('.', '').replace('-', ' ').replace('  ', ' ').strip()
                    if k_clean == target_clean:
                        return k, v
                return None

            base_match = _find_position_match(country_position)
            onsite_match = _find_position_match(onsite_position)

            base_position = base_match[0] if base_match else country_position
            base_rate_info = base_match[1] if base_match else None
            onsite_rate_info = onsite_match[1] if onsite_match else None

        effective_hour_rate = None
        if total_h > 0:
            base_hr = (base_rate_info or {}).get('hour_rate') or 0
            onsite_hr = (onsite_rate_info or {}).get('hour_rate') or base_hr
            if base_hr or onsite_hr:
                effective_hour_rate = round(
                    (onsite_hours * (onsite_hr or 0) + non_onsite_hours * (base_hr or 0)) / total_h, 2
                ) if total_h else 0

        team.append({
            'name': emp_name,
            'odoo_role': odoo_role,
            'position': base_position,
            'has_base_rate': bool(base_rate_info),
            'has_onsite_rate': bool(onsite_rate_info),
            'country': country,
            'regular_mh': round(regular_mh, 2),
            'ramadan_mh': round(ramadan_mh, 2),
            'overtime_mh': round(overtime_mh, 2),
            'total_hours': round(total_h, 2),
            'mds': total_md,
            'onsite_days': onsite_days,
            'onsite_hours': round(onsite_hours, 2),
            'base_hour_rate': (base_rate_info or {}).get('hour_rate'),
            'onsite_hour_rate': (onsite_rate_info or {}).get('hour_rate'),
            'effective_hour_rate': effective_hour_rate,
        })

    # Sort alphabetically
    team.sort(key=lambda x: (x['name'] or '').lower())

    return {
        'team': team,
        'month_label': date(year, month, 1).strftime('%B %Y'),
        'year': year,
        'month': month,
        'date_from': month_start,
        'date_to': month_end,
        'phases_used': phases,
    }


def get_positions_from_excel():
    """Read Positions sheet and return list of {position, hour_rate, md_rate}"""
    if not os.path.exists(VARIANCE_FILE):
        return []
    try:
        df = pd.read_excel(VARIANCE_FILE, sheet_name='Positions', header=None)
        rows = df.values.tolist()
        positions = []
        for r in rows[1:]:  # skip header row
            if not r or pd.isna(r[0]):
                continue
            positions.append({
                'name': str(r[0]).strip(),
                'hour_rate': safe_val(r[1]) if len(r) > 1 else None,
                'md_rate': safe_val(r[2]) if len(r) > 2 else None,
            })
        return positions
    except Exception as e:
        logger.error(f"Positions parse: {e}")
        return []


def get_tunis_rates_from_excel():
    """Read 'Tunis Rates' sheet — Tunis people are paid per name, not per position"""
    if not os.path.exists(VARIANCE_FILE):
        return {}
    try:
        df = pd.read_excel(VARIANCE_FILE, sheet_name='Tunis Rates', header=None)
        rows = df.values.tolist()
        tunis = {}
        for r in rows[1:]:
            if not r or pd.isna(r[0]):
                continue
            name = str(r[0]).strip()
            hr = safe_val(r[1]) if len(r) > 1 else None
            if hr is not None:
                # Store by lowercase name for case-insensitive lookup
                tunis[name.lower()] = {
                    'name': name,
                    'hour_rate': float(hr),
                    'md_rate': float(hr) * 8,  # 8 hours per MD
                }
        return tunis
    except Exception as e:
        logger.warning(f"Tunis Rates parse: {e}")
        return {}


def find_tunis_rate(emp_name, tunis_rates):
    """Match Tunis employee by partial name (handles case + extra prefix like '[T###]')"""
    if not emp_name or not tunis_rates:
        return None
    # Strip [code] prefix
    import re
    clean = re.sub(r'^\s*\[[A-Z]\d+\s*\]\s*', '', str(emp_name)).strip().lower()
    # Direct match
    if clean in tunis_rates:
        return tunis_rates[clean]
    # Partial — match if all words from tunis name appear in clean (or vice versa)
    for tname, tinfo in tunis_rates.items():
        tparts = set(tname.split())
        cparts = set(clean.split())
        if tparts.issubset(cparts) or cparts.issubset(tparts):
            return tinfo
        # Also try if any 2 significant words match
        common = tparts & cparts
        if len(common) >= 2:
            return tinfo
    return None


def get_odoo_position_for_employee(employee_name):
    """Try to fetch position from Odoo hr.employee model for an employee by name.
    Strips [code] prefix since Odoo stores names without it.
    Returns position name or None.
    """
    if not employee_name:
        return None
    if not odoo.uid:
        if not odoo.connect():
            return None

    import re
    clean_name = re.sub(r'^\s*\[[A-Z]\d+\s*\]\s*', '', employee_name).strip()
    candidates = [clean_name, employee_name]

    try:
        emp = None
        for q in candidates:
            if not q:
                continue
            emps = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'hr.employee', 'search_read',
                [[('name', '=', q)]],
                {'fields': ['name', 'job_title', 'job_id'], 'limit': 1}
            )
            if emps:
                emp = emps[0]
                break
            emps = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'hr.employee', 'search_read',
                [[('name', 'ilike', q)]],
                {'fields': ['name', 'job_title', 'job_id'], 'limit': 1}
            )
            if emps:
                emp = emps[0]
                break

        if not emp:
            return None
        if emp.get('job_id'):
            return emp['job_id'][1]
        return emp.get('job_title') or None
    except Exception as e:
        logger.warning(f"Odoo position lookup for '{employee_name}' (cleaned: '{clean_name}'): {e}")
        return None


# SAR → USD conversion rate
SAR_TO_USD = 3.75
OVERTIME_MULTIPLIER = 1.5


def get_odoo_rate_for_employee(employee_name):
    """Fetch hourly rate from Odoo hr.employee (timesheet_cost in SAR).
    Converts to USD using SAR_TO_USD.
    Returns: { 'hour_rate': float, 'overtime_rate': float, 'source': 'odoo' } or None.
    """
    if not employee_name:
        return None
    if not odoo.uid:
        if not odoo.connect():
            return None

    # IMPORTANT: Odoo stores names WITHOUT the [E123] prefix that appears in timesheets!
    # E.g. timesheet shows "[E102] AbdelRahman Doghish" but hr.employee.name = "AbdelRahman Doghish"
    import re
    clean_name = re.sub(r'^\s*\[[A-Z]\d+\s*\]\s*', '', employee_name).strip()
    # Try both - exact match on cleaned, then ilike on cleaned, then ilike on raw
    candidates = [clean_name, employee_name]

    try:
        emp = None
        for q in candidates:
            if not q:
                continue
            # Try exact first
            emps = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'hr.employee', 'search_read',
                [[('name', '=', q)]],
                {'fields': ['name', 'timesheet_cost'], 'limit': 1}
            )
            if emps:
                emp = emps[0]
                break
            # Try ilike (fuzzy)
            emps = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'hr.employee', 'search_read',
                [[('name', 'ilike', q)]],
                {'fields': ['name', 'timesheet_cost'], 'limit': 1}
            )
            if emps:
                emp = emps[0]
                break

        if not emp:
            return None

        sar_rate = emp.get('timesheet_cost')
        if not sar_rate or sar_rate <= 0:
            return None

        usd_rate = round(sar_rate / SAR_TO_USD, 2)
        return {
            'hour_rate': usd_rate,
            'overtime_rate': round(usd_rate * OVERTIME_MULTIPLIER, 2),
            'source': 'odoo',
            'sar_rate': sar_rate,
            'matched_name': emp.get('name'),
        }
    except Exception as e:
        logger.warning(f"Odoo rate lookup for '{employee_name}' (cleaned: '{clean_name}'): {e}")
        return None


def get_employee_rate(employee_name, position_name=None, is_onsite=False):
    """Universal rate lookup with priority:
    1. Tunis employee (by name) → use tunis_rates from DB
    2. Try Odoo timesheet_cost → convert SAR→USD
    3. Fallback: position_name lookup in DB positions catalog
    Returns: { hour_rate, overtime_rate, md_rate?, source, position? } or None
    """
    if not employee_name:
        return None

    # 1. Tunis check (prefix [T...])
    country = get_country_from_employee_name(employee_name)
    if country == 'TUN':
        tunis = get_tunis_rate_by_name(db, employee_name)
        if tunis and tunis.get('hour_rate'):
            return {
                'hour_rate': tunis['hour_rate'],
                'overtime_rate': round(tunis['hour_rate'] * OVERTIME_MULTIPLIER, 2),
                'source': 'tunis_rates_db',
                'position': f'TUN - {tunis["name"]}',
            }
        # If no Tunis rate found, fall through to other lookups

    # 2. Try Odoo first (preferred - real-time data)
    odoo_rate = get_odoo_rate_for_employee(employee_name)
    if odoo_rate:
        # For onsite (Egyptian traveling), Odoo rate is the BASE
        # The onsite premium is applied via DB position lookup below
        if not is_onsite:
            return odoo_rate
        # If onsite, fall through to use DB onsite position rate

    # 3. Fallback to DB positions catalog
    if position_name:
        # Apply onsite suffix if needed
        full_pos = position_name
        if is_onsite and country == 'EGY' and not position_name.endswith('- onsite'):
            full_pos = position_name + ' - onsite'

        pos_info = get_position_by_name(db, full_pos)
        if pos_info:
            hr = pos_info.get('hour_rate')
            if hr:
                return {
                    'hour_rate': hr,
                    'overtime_rate': round(hr * OVERTIME_MULTIPLIER, 2),
                    'md_rate': pos_info.get('md_rate'),
                    'source': 'positions_db' + ('_onsite' if is_onsite else ''),
                    'position': full_pos,
                }

    # If Odoo had a rate but we wanted onsite + no DB match, return Odoo as fallback
    if odoo_rate:
        return odoo_rate

    return None


@app.route('/api/employee-rate')
def api_employee_rate():
    """Test endpoint: get rate for an employee.
    Query: ?name=...&position=...&onsite=true
    """
    name = request.args.get('name', '')
    position = request.args.get('position')
    onsite = request.args.get('onsite', '').lower() in ('true', '1', 'yes')

    if not name:
        return jsonify({'error': 'name required'}), 400

    rate = get_employee_rate(name, position, onsite)
    return jsonify({
        'employee_name': name,
        'requested_position': position,
        'is_onsite': onsite,
        'rate': rate,
    })


def batch_fetch_employees_from_odoo(employee_names):
    """Fetch position + timesheet_cost for a batch of employees in ONE Odoo query.
    Strips [E123] prefix from names since Odoo stores names without it.
    Returns: { original_name: {'odoo_name', 'position', 'sar_rate'} }
    """
    if not employee_names:
        return {}
    if not odoo.uid:
        if not odoo.connect():
            return {}

    import re
    cleaned_to_originals = {}
    for n in employee_names:
        if not n:
            continue
        clean = re.sub(r'^\s*\[[A-Z]\d+\s*\]\s*', '', n).strip()
        if clean:
            cleaned_to_originals.setdefault(clean, []).append(n)

    if not cleaned_to_originals:
        return {}

    try:
        emps = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'search_read',
            [[('name', 'in', list(cleaned_to_originals.keys()))]],
            {'fields': ['id', 'name', 'job_title', 'job_id', 'timesheet_cost']}
        )

        result = {}
        for emp in emps:
            emp_name = emp.get('name', '')
            position = None
            if emp.get('job_id'):
                position = emp['job_id'][1] if isinstance(emp['job_id'], list) else None
            position = position or emp.get('job_title')
            sar_rate = emp.get('timesheet_cost') or 0

            entry = {
                'odoo_name': emp_name,
                'position': position,
                'sar_rate': sar_rate,
            }

            for original in cleaned_to_originals.get(emp_name, []):
                result[original] = entry

        logger.info(f"Batch fetched {len(emps)} of {len(cleaned_to_originals)} employees from Odoo")
        return result
    except Exception as e:
        logger.warning(f"Batch employee fetch failed: {e}")
        return {}


@app.route('/debug/employee-fields')
def debug_employee_fields():
    """Returns ALL fields of all hr.employee records to discover the correct rate field name."""
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'error': 'Odoo unreachable'}), 500
    try:
        # First, get available fields of hr.employee model
        fields_info = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'hr.employee', 'fields_get',
            [],
            {'attributes': ['type', 'string']}
        )
        # Filter to likely rate-related fields
        rate_like = {}
        for fname, finfo in fields_info.items():
            label = (finfo.get('string') or '').lower()
            if any(k in fname.lower() for k in ['cost', 'rate', 'hour', 'salary', 'wage']):
                rate_like[fname] = finfo
            elif any(k in label for k in ['cost', 'rate', 'hour', 'salary', 'wage']):
                rate_like[fname] = finfo

        # Try to fetch sample employee with all likely fields
        sample = None
        emp_name = request.args.get('name')
        if emp_name:
            try:
                emps = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'hr.employee', 'search_read',
                    [[('name', 'ilike', emp_name)]],
                    {'fields': list(rate_like.keys()) + ['name'], 'limit': 5}
                )
                sample = emps
            except Exception as e:
                sample = {'error': str(e)}

        return jsonify({
            'rate_like_fields_count': len(rate_like),
            'rate_like_fields': rate_like,
            'sample_employee': sample,
            'usage': 'Add ?name=Doghish to see actual values for a specific employee',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
# PLAN OVERRIDES (stored in JSON)
# ============================================================
PLAN_FILE = os.path.join(PERSIST_DIR, 'plan_overrides.json')  # legacy path (kept for backup endpoint)


def load_plan_overrides():
    """Load plan overrides from DB with project prefix.
    Returns structure compatible with old JSON format:
    {
      'plan_overrides': {phase: {month_key: {field: value}}},
      'position_overrides': {emp_name: position_name}
    }
    """
    pfx = active_db_prefix()
    plan_ns = f'{pfx}_plan' if pfx else 'plan'

    # Plan overrides: namespace='plan', phase=<phase>, key='<month_key>.<field>'
    plan_data = db.get_namespace_overrides(plan_ns)
    plan_overrides = {}
    for phase, items in plan_data.items():
        if not phase:
            continue
        plan_overrides[phase] = {}
        for combined_key, val in items.items():
            # combined_key = '<month_key>.<field>'
            if '.' in combined_key:
                month_key, field = combined_key.rsplit('.', 1)
                if month_key not in plan_overrides[phase]:
                    plan_overrides[phase][month_key] = {}
                plan_overrides[phase][month_key][field] = val

    # Position overrides: namespace='position', phase='', key=<emp_name>
    position_overrides = db.get_namespace_overrides('position', '')

    return {
        'plan_overrides': plan_overrides,
        'position_overrides': position_overrides,
    }


def save_plan_overrides(data):
    """Save plan_overrides back to DB with project prefix."""
    pfx = active_db_prefix()
    plan_ns = f'{pfx}_plan' if pfx else 'plan'

    plan = data.get('plan_overrides', {}) or {}
    for phase, months in plan.items():
        if not isinstance(months, dict):
            continue
        for month_key, fields in months.items():
            if not isinstance(fields, dict):
                continue
            for field, value in fields.items():
                db.set_override(plan_ns, phase, f"{month_key}.{field}", value)

    pos = data.get('position_overrides', {}) or {}
    for emp_name, pos_val in pos.items():
        db.set_override('position', '', emp_name, pos_val)


@app.route('/api/positions')
def api_positions():
    """Get full positions catalog from DB (replaces Excel-based)."""
    positions = get_all_positions(db)
    tunis = get_all_tunis_rates(db)
    return jsonify({
        'positions': positions,
        'count': len(positions),
        'tunis_rates': tunis,
        'tunis_count': len(tunis),
    })


@app.route('/api/positions/reseed', methods=['POST'])
def api_positions_reseed():
    """Force re-seed positions catalog with updated rates (overwrites existing)."""
    from positions_catalog import POSITIONS_SEED, TUNIS_RATES_SEED
    count = 0
    for p in POSITIONS_SEED:
        upsert_position(
            db,
            position_name=p['position'],
            hour_rate=p['hour_rate'],
            md_rate=p.get('md_rate'),
            country=p.get('country'),
            is_onsite=p.get('is_onsite', False),
        )
        count += 1
    for t in TUNIS_RATES_SEED:
        upsert_tunis_rate(db, t['name'], t['hour_rate'])
        count += 1
    return jsonify({'ok': True, 'updated': count})


def api_positions_save():
    """Add or update a position. Body: { position, hour_rate, md_rate, country, is_onsite }"""
    body = request.json or {}
    position = body.get('position', '').strip()
    if not position:
        return jsonify({'error': 'position name required'}), 400

    def to_float(v):
        if v is None or v == '':
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    result = upsert_position(
        db,
        position_name=position,
        hour_rate=to_float(body.get('hour_rate')),
        md_rate=to_float(body.get('md_rate')),
        country=body.get('country'),
        is_onsite=body.get('is_onsite'),
    )
    return jsonify({'ok': True, 'position': result})


@app.route('/api/positions/<path:position_name>', methods=['DELETE'])
def api_positions_delete(position_name):
    delete_position(db, position_name)
    return jsonify({'ok': True})


@app.route('/api/tunis-rates', methods=['POST'])
def api_tunis_rates_save():
    """Add or update a Tunis person rate."""
    body = request.json or {}
    name = body.get('name', '').strip()
    hour_rate = body.get('hour_rate')
    if not name:
        return jsonify({'error': 'name required'}), 400
    try:
        hour_rate = float(hour_rate) if hour_rate else None
    except Exception:
        hour_rate = None
    upsert_tunis_rate(db, name, hour_rate)
    return jsonify({'ok': True, 'name': name, 'hour_rate': hour_rate})


@app.route('/api/tunis-rates/<path:name>', methods=['DELETE'])
def api_tunis_rates_delete(name):
    delete_tunis_rate(db, name)
    return jsonify({'ok': True})


@app.route('/api/effort/<phase_key>')
def api_effort(phase_key):
    """Computed effort for a phase + month from Odoo (single month)"""
    year = int(request.args.get('year') or date.today().year)
    month = int(request.args.get('month') or date.today().month)

    employees_in_data = set()
    raw = odoo.get_timesheets(
        date_from=date(year, month, 1).isoformat(),
        date_to=(date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)).isoformat(),
    )
    if raw:
        for e in raw:
            emp = e.get('employee_id')
            if emp and emp[1]:
                employees_in_data.add(emp[1])

    position_lookup = {}
    for emp_name in employees_in_data:
        pos = get_odoo_position_for_employee(emp_name)
        if pos:
            position_lookup[emp_name] = pos

    overrides = load_plan_overrides().get('position_overrides', {})
    position_lookup.update(overrides)

    result = compute_effort_from_odoo(phase_key, year, month, position_lookup)
    return jsonify(result)


@app.route('/api/effort/<phase_key>/all-months')
def api_effort_all_months(phase_key):
    """Excel-style Current Effort table:
       - Rows: each employee (one row per person)
       - Cols: # / Name / Position / Hour Rate / Overtime Rate / [month1: Reg/Ram/OT] / [month2: Reg/Ram/OT] ...
       - Months start from FIRST month with any log in the project (across the phase)
       - Each cell is total hours classified as Regular / Ramadan / Overtime per country rules.
    """
    _proj_name = active_project_name()
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'error': 'Odoo unreachable', 'employees': [], 'months': []}), 503

    _is_bog_eff = not session.get('project_id') or str(session.get('project_id')) == '228'

    # Step 1: Determine the relevant tasks
    phase_names = get_phase_mapping().get(phase_key, []) if _is_bog_eff else []

    try:
        # Find project
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', _proj_name)]],
            {'fields': ['id', 'name'], 'limit': 5}
        )
        if not projects:
            return jsonify({'employees': [], 'months': []})
        project_id = projects[0]['id']

        # BOG: resolve phase names from mapping
        if _is_bog_eff and not phase_names:
            phase_names = auto_detect_phases_for_project(project_id, phase_key)

        if _is_bog_eff and not phase_names:
            return jsonify({'employees': [], 'months': [], 'total_employees': 0,
                            'note': f'No phases found for {phase_key} in this project'})

        # Non-BOG: get project.phase records split by support keywords
        use_all_tasks_eff = False
        if not _is_bog_eff:
            all_proj_phases = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.phase', 'search_read',
                [[('project_id', '=', project_id)]],
                {'fields': ['id', 'name'], 'limit': 100}
            )
            if all_proj_phases:
                if phase_key == 'support':
                    phase_names = [p['name'] for p in all_proj_phases
                                   if any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
                else:
                    phase_names = [p['name'] for p in all_proj_phases
                                   if not any(kw in p['name'].lower() for kw in SUPPORT_KWS)]
            else:
                use_all_tasks_eff = True  # no phases → all tasks

        # Non-BOG support with no phases → return empty with clear message
        if not _is_bog_eff and phase_key == 'support' and not phase_names:
            return jsonify({'employees': [], 'months': [], 'total_employees': 0,
                           'has_phases': False,
                           'note': f'No support/operation phases found in this project'})

        # Find phases: try project.phase first (BOG), then project.task.type (non-BOG stages)
        phases = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('name', 'in', phase_names)]],
            {'fields': ['id', 'name']}
        )
        found_names_lower = {p['name'].strip().lower() for p in phases}
        for pname in phase_names:
            if pname.strip().lower() not in found_names_lower:
                extra = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'project.phase', 'search_read',
                    [[('name', 'ilike', pname.strip()), ('project_id', '=', project_id)]],
                    {'fields': ['id', 'name'], 'limit': 1}
                )
                if extra:
                    phases.extend(extra)
                    found_names_lower.add(extra[0]['name'].strip().lower())
        phase_ids = [p['id'] for p in phases]

        # All project tasks
        all_proj_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('project_id', '=', project_id)]],
            {'fields': ['id', 'parent_id', 'phase_id', 'stage_id'], 'limit': 10000}
        )
        task_map = {t['id']: t for t in all_proj_tasks}

        if _is_bog_eff and phase_ids:
            # BOG: filter by project.phase (phase_id on task)
            relevant_ids = set()
            for t in all_proj_tasks:
                ph = t.get('phase_id')
                if ph and isinstance(ph, list) and ph[0] in phase_ids:
                    relevant_ids.add(t['id'])
        elif not _is_bog_eff and phase_ids:
            # Non-BOG with phases: filter by phase_id
            relevant_ids = set()
            for t in all_proj_tasks:
                ph = t.get('phase_id')
                if ph and isinstance(ph, list) and ph[0] in phase_ids:
                    relevant_ids.add(t['id'])
            logger.info(f"Non-BOG effort {phase_key}: {len(relevant_ids)} tasks in {len(phase_ids)} phases")
        else:
            # Non-BOG with no phases: all project tasks
            relevant_ids = {t['id'] for t in all_proj_tasks}
            logger.info(f"Non-BOG effort {phase_key}: using all {len(relevant_ids)} tasks (no phases)")

        # Check if user chose to include unassigned hours in this phase
        _incl_unassigned = proj_get_override('plan', phase_key, 'unassigned.include_unassigned')
        if _incl_unassigned:
            unassigned_ids = {t['id'] for t in all_proj_tasks if not t.get('phase_id')}
            relevant_ids = relevant_ids | unassigned_ids
            logger.info(f"Including {len(unassigned_ids)} unassigned tasks in {phase_key}")

        # Walk parent chains to include sub-tasks of phase tasks
        for t in all_proj_tasks:
            if t['id'] in relevant_ids:
                continue
            cur = t
            visited = set()
            while cur and cur['id'] not in visited:
                visited.add(cur['id'])
                if cur['id'] in relevant_ids:
                    relevant_ids.add(t['id'])
                    break
                if not cur.get('parent_id'):
                    break
                par = cur['parent_id'][0] if isinstance(cur['parent_id'], list) else cur['parent_id']
                cur = task_map.get(par)

        # Fetch all timesheets for these tasks (no date filter - want full history)
        timesheets = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.analytic.line', 'search_read',
            [[('task_id', 'in', list(relevant_ids))]],
            {'fields': ['employee_id', 'date', 'unit_amount'], 'limit': 100000}
        )
    except Exception as e:
        import traceback
        logger.error(f"all-months effort fetch failed: {e}\n{traceback.format_exc()}")
        return jsonify({'error': str(e), 'trace': traceback.format_exc(), 'employees': [], 'months': []}), 500

    if not timesheets:
        return jsonify({'employees': [], 'months': [], 'project_start_month': None})

    # Step 2: Determine the first month with any log
    earliest = None
    latest = None
    for ts in timesheets:
        d = ts.get('date')
        if not d:
            continue
        try:
            dt = datetime.strptime(d, '%Y-%m-%d').date()
            if earliest is None or dt < earliest:
                earliest = dt
            if latest is None or dt > latest:
                latest = dt
        except Exception:
            continue

    if not earliest:
        return jsonify({'employees': [], 'months': []})

    # Build list of months from first log to current month (or last log, whichever later)
    today_d = date.today()
    end_month = max(latest, today_d) if latest else today_d

    months = []
    cur = date(earliest.year, earliest.month, 1)
    end_first = date(end_month.year, end_month.month, 1)
    while cur <= end_first:
        months.append({
            'year': cur.year,
            'month': cur.month,
            'label': cur.strftime('%B %Y'),     # "April 2026"
            'short': cur.strftime('%b'),         # "Apr"
            'key': cur.strftime('%Y-%m'),
        })
        # Next month
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)

    # Step 3: Aggregate per employee per month
    # person_data[emp_name][month_key] = {regular, ramadan, overtime}
    from collections import defaultdict
    person_data = defaultdict(lambda: defaultdict(lambda: {'regular': 0, 'ramadan': 0, 'overtime': 0}))

    # Travel records for onsite detection
    travel_records = load_travel()

    for ts in timesheets:
        emp = ts.get('employee_id', [None, ''])
        emp_name = emp[1] if isinstance(emp, list) and len(emp) > 1 else ''
        if not emp_name:
            continue
        d_str = ts.get('date')
        if not d_str:
            continue
        try:
            d = datetime.strptime(d_str, '%Y-%m-%d').date()
        except Exception:
            continue
        h = float(ts.get('unit_amount') or 0)
        if h <= 0:
            continue
        month_key = d.strftime('%Y-%m')

        country = get_country_from_employee_name(emp_name)

        # Check Ramadan
        is_ramadan = False
        ramadan = RAMADAN_RANGES.get(country)
        if ramadan:
            try:
                r_start = datetime.strptime(ramadan['start'], '%Y-%m-%d').date()
                r_end = datetime.strptime(ramadan['end'], '%Y-%m-%d').date()
                if r_start <= d <= r_end:
                    is_ramadan = True
            except Exception:
                pass

        # Weekend check → overtime
        weekend = TUNIS_WEEKEND if country == 'TUN' else WEEKEND_DAYS
        is_weekend = d.weekday() in weekend
        is_holiday = is_public_holiday(d_str, country)

        if is_weekend or is_holiday:
            person_data[emp_name][month_key]['overtime'] += h
        elif is_ramadan:
            person_data[emp_name][month_key]['ramadan'] += h
        else:
            person_data[emp_name][month_key]['regular'] += h

    # Step 4: BATCH fetch all employees from Odoo in ONE query (performance!)
    all_emp_names = list(person_data.keys())
    odoo_data = batch_fetch_employees_from_odoo(all_emp_names)
    logger.info(f"Resolved {len(odoo_data)} of {len(all_emp_names)} employees from Odoo")

    # Build emp_odoo_ids map: emp_name → Odoo employee ID (for travel matching)
    emp_odoo_ids = {}
    for ts in timesheets:
        emp = ts.get('employee_id', [None, ''])
        emp_name_ts = emp[1] if isinstance(emp, list) and len(emp) > 1 else ''
        emp_id_ts   = emp[0] if isinstance(emp, list) and len(emp) > 0 else None
        if emp_name_ts and emp_id_ts and emp_name_ts not in emp_odoo_ids:
            emp_odoo_ids[emp_name_ts] = emp_id_ts

    # ── imports for onsite split logic ──
    import re as _re_effort
    import calendar as _cal
    from datetime import date as _date_cls2

    # ── keyword-based onsite position resolver ──
    _EGY_ONSITE_MAP = {
        ('lead',   'business analyst'):  'EGY - Lead Business Analyst - onsite',
        ('senior', 'business analyst'):  'EGY - Sr Business Analyst - onsite',
        ('mid',    'business analyst'):  'EGY - Business Analyst - onsite',
        ('lead',   'software engineer'): 'EGY - Lead Software Engineer - onsite',
        ('senior', 'software engineer'): 'EGY - Sr. Software Engineer - onsite',
        ('mid',    'software engineer'): 'EGY - Software Engineer - onsite',
        ('lead',   'quality engineer'):  'EGY - Lead Quality Engineer - onsite',
        ('senior', 'quality engineer'):  'EGY - Sr Quality Engineer - onsite',
        ('mid',    'quality engineer'):  'EGY - Quality Engineer - onsite',
        ('lead',   'ux designer'):       'EGY - Lead UX Designer - onsite',
        ('senior', 'ux designer'):       'EGY - Sr UX Designer - onsite',
        ('mid',    'ux designer'):       'EGY - UX Designer - onsite',
        ('lead',   'project manager'):   'EGY - Project Manager - onsite',
        ('senior', 'project manager'):   'EGY - Project Manager - onsite',
        ('mid',    'project manager'):   'EGY - Project Manager - onsite',
    }

    def _detect_level(s):
        p = s.lower()
        if any(x in p for x in ['lead', 'team lead', 'principal', 'head']):
            return 'lead'
        if any(x in p for x in ['senior', 'sr.', ' sr ']):
            return 'senior'
        return 'mid'

    def _detect_role(s):
        p = s.lower()
        if any(x in p for x in ['project manager', 'pm ']):
            return 'project manager'
        if any(x in p for x in ['business anal', 'ba ', ' ba']):
            return 'business analyst'
        if any(x in p for x in ['software', 'developer', ' sw ']):
            return 'software engineer'
        if any(x in p for x in ['quality', 'qc', 'qa', 'test']):
            return 'quality engineer'
        if any(x in p for x in ['ux', 'ui/ux', 'design']):
            return 'ux designer'
        if 'engineer' in p:
            return 'software engineer'
        if 'analyst' in p:
            return 'business analyst'
        return None

    def _resolve_onsite_position(odoo_pos_str, country):
        if not odoo_pos_str or country != 'EGY':
            return None
        level = _detect_level(odoo_pos_str)
        role = _detect_role(odoo_pos_str)
        if role:
            return _EGY_ONSITE_MAP.get((level, role))
        return None

    def _calc_mds(regular_h, ramadan_h, overtime_h):
        """Correct MD formula:
        MDs = (Regular + Overtime) / 8  +  Ramadan / 6
        """
        return round((regular_h + overtime_h) / 8.0 + ramadan_h / 6.0, 2)

    def _calc_cost(regular_h, ramadan_h, overtime_h, hour_rate, overtime_rate):
        """Correct cost formula:
        Cost = (Regular + Ramadan) * hour_rate  +  Overtime * overtime_rate
        """
        base = (regular_h + ramadan_h) * (hour_rate or 0)
        ot = overtime_h * (overtime_rate or 0)
        return round(base + ot, 2)

    _all_positions = get_all_positions(db)
    all_promotions = load_promotions()

    # For each employee: segment hours by (position, is_onsite) per day
    employees_out = []

    for emp_name, months_data in sorted(person_data.items()):
        country = get_country_from_employee_name(emp_name)

        emp_code_match = _re_effort.match(r'^\[([A-Z]\d+)\s*\]', emp_name)
        emp_code = emp_code_match.group(1) if emp_code_match else ''
        emp_display = _re_effort.sub(r'^\[[A-Z]\d+\s*\]\s*', '', emp_name).strip()

        odoo_info = odoo_data.get(emp_name, {})
        odoo_pos = odoo_info.get('position')
        sar_rate = odoo_info.get('sar_rate', 0)

        # ── Travel periods (match by Odoo ID, code, or name) ──
        emp_travel_periods = []
        en_clean = _re_effort.sub(r'\[[A-Z]\d+\s*\]\s*', '', emp_name).strip().lower()
        # Extract employee code from timesheet name [E259] Sara Samir
        _code_m = _re_effort.match(r'^\[([A-Z]\d+)\s*\]', emp_name)
        emp_code = _code_m.group(1).upper() if _code_m else ''
        # Get Odoo employee ID from timesheet data
        _emp_odoo_id = emp_odoo_ids.get(emp_name)
        for tr in travel_records:
            if not tr.get('name') or not tr.get('start_date'):
                continue
            matched_tr = False
            # 1. Odoo ID match
            if _emp_odoo_id and tr.get('odoo_employee_id'):
                try:
                    matched_tr = int(tr['odoo_employee_id']) == int(_emp_odoo_id)
                except: pass
            # 2. Employee code match (emp_code = E259 from "[E259] Sara Samir")
            if not matched_tr and emp_code and tr.get('odoo_employee_code'):
                matched_tr = emp_code.upper() == str(tr['odoo_employee_code']).upper().strip()
            # 3. Name match
            if not matched_tr:
                tn = tr['name'].strip().lower()
                matched_tr = (tn == en_clean or tn in en_clean or en_clean in tn)
            if matched_tr:
                emp_travel_periods.append({'start': tr['start_date'], 'end': tr.get('end_date')})

        # ── Promotion records (match by code or name) ──
        emp_promos = []
        for pr in all_promotions:
            matched_pr = False
            # 1. Odoo employee ID match (most reliable)
            if _emp_odoo_id and pr.get('odoo_employee_id'):
                try:
                    matched_pr = int(pr['odoo_employee_id']) == int(_emp_odoo_id)
                except: pass
            # 2. Code match (E259 etc)
            if not matched_pr and emp_code and pr.get('odoo_employee_code'):
                matched_pr = emp_code.upper() == str(pr['odoo_employee_code']).upper().strip()
            # 3. Name match (fuzzy)
            if not matched_pr:
                pn = (pr.get('name') or '').strip().lower()
                matched_pr = (pn == en_clean or pn in en_clean or en_clean in pn)
            if matched_pr:
                emp_promos.append(pr)
        emp_promos.sort(key=lambda x: x.get('promotion_date') or x.get('effective_date') or '')

        def _date_is_onsite(date_str):
            for period in emp_travel_periods:
                if date_str < period['start']: continue
                if period['end'] is None or date_str <= period['end']: return True
            return False

        def _strip_country(pos_str):
            """Strip country prefix: EGY - Senior BA → Senior BA"""
            import re as _re_strip
            return _re_strip.sub(r'^(EGY|KSA|TUN|UAE|KWT|BHR|OMN|QAT|JOR|LBN|SAU)\s*-\s*', '', (pos_str or '').strip())

        def _downgrade_position(pos_str):
            """Infer the position BEFORE a promotion by reversing the level.
            Lead X      → Sr X
            Sr X / Sr. X → X  (drop Sr prefix)
            Senior X    → X
            If no level prefix found, returns None (can't infer).
            """
            import re as _re_dg
            s = (pos_str or '').strip()
            # Lead → Sr
            m = _re_dg.match(r'^Lead\s+(.+)$', s, _re_dg.IGNORECASE)
            if m:
                return 'Sr. ' + m.group(1).strip()
            # Sr. / Sr / Senior → base
            m = _re_dg.match(r'^(?:Sr\.?|Senior)\s+(.+)$', s, _re_dg.IGNORECASE)
            if m:
                return m.group(1).strip()
            return None  # can't infer (already at base level or unknown pattern)

        def _base_position_on_date(date_str):
            """Return raw position for employee on given date, considering promotions.
            Priority:
            1. Before promotion date:
               a. old_title if explicitly set → use it
               b. old_title empty → infer from new_title by reversing level (Lead→Sr, Sr→base)
               c. can't infer → use Odoo current position (fallback, log warning)
            2. On or after promotion date → new_title from applicable promo record
            3. No promotions → odoo current position
            """
            base = _strip_country(odoo_pos or '')

            if not emp_promos:
                return base

            sorted_promos = sorted(emp_promos,
                key=lambda x: x.get('promotion_date') or x.get('effective_date') or '')

            # Find the last promotion whose date <= date_str
            applicable = None
            for promo in sorted_promos:
                promo_date = promo.get('promotion_date') or promo.get('effective_date') or ''
                if promo_date and date_str >= promo_date:
                    applicable = promo

            if applicable is None:
                # date_str is BEFORE any promotion — use old_title of earliest promo
                first = sorted_promos[0]
                old = (first.get('old_position') or first.get('old_title') or '').strip()
                if old:
                    return _strip_country(old)
                # old_title is empty → try to infer from new_title
                new_for_infer = (first.get('new_position') or first.get('new_title') or '').strip()
                inferred = _downgrade_position(_strip_country(new_for_infer))
                if inferred:
                    logger.debug(f"[promo] {emp_name}: old_title missing, inferred '{inferred}' from '{new_for_infer}' for {date_str}")
                    return inferred
                # Can't infer → log warning and use Odoo base (least bad option)
                logger.warning(f"[promo] {emp_name}: old_title missing & can't infer for {date_str}, using Odoo pos '{base}'")
                return base
            else:
                # On or after this promotion date — use new_title
                new_pos = (applicable.get('new_position') or applicable.get('new_title') or '').strip()
                return _strip_country(new_pos) if new_pos else base

        # KSA position aliases: Odoo returns various formats, normalize to DB keys
        _KSA_ALIASES = {
            'ksa-business analyst':              'KSA - Business Analyst',
            'ksa-senior business analyst':       'KSA - Sr Business Analyst',
            'ksa-sr business analyst':           'KSA - Sr Business Analyst',
            'ksa - senior business analyst':     'KSA - Sr Business Analyst',
            'ksa - ksa-business analyst':        'KSA - Business Analyst',
            'ksa - ksa-senior business analyst': 'KSA - Sr Business Analyst',
            'ksa - ksa-sr business analyst':     'KSA - Sr Business Analyst',
            'ksa - senior project manager':      'KSA - Project Manager',
            'ksa - technical project manager':   'KSA - Project Manager',
            'ksa - business analysis team lead': 'KSA - Lead Business Analyst',
            'ksa - sr. manager, business consulting': 'KSA - Solution Architect / Manager',
            'ksa - associated business analyst': 'KSA - Business Analyst',
        }

        def _lookup_pos_fuzzy(pos_str):
            """Try exact DB lookup, then normalize spaces/dashes, then alias map."""
            if not pos_str:
                return None
            # 0. Check alias map first
            alias = _KSA_ALIASES.get(pos_str.strip().lower())
            if alias:
                pi = get_position_by_name(db, alias)
                if pi and pi.get('hour_rate'):
                    return pi
            # 1. Exact match
            pi = get_position_by_name(db, pos_str)
            if pi and pi.get('hour_rate'):
                return pi
            # 2. Normalize: collapse spaces around dashes
            import re as _re_pos
            def _norm(s):
                s = _re_pos.sub(r'\s*-\s*', ' - ', s)
                s = _re_pos.sub(r'\s+', ' ', s).strip()
                return s.lower()
            target = _norm(pos_str)
            all_pos = get_all_positions(db)
            for p in all_pos:
                if _norm(p.get('position', '') or p.get('name', '')) == target and p.get('hour_rate'):
                    return p
            # 3. Also try normalizing Sr. / Senior / Sr variants
            def _core(s):
                s = _re_pos.sub(r'\bsenior\b', 'sr', s, flags=_re_pos.IGNORECASE)
                s = _re_pos.sub(r'\bsr\.\b', 'sr', s, flags=_re_pos.IGNORECASE)
                return _norm(s)
            t_core = _core(pos_str)
            for p in all_pos:
                pname = p.get('position', '') or p.get('name', '')
                if _core(pname) == t_core and p.get('hour_rate'):
                    return p
            return None

        def _get_rate_for_pos(base_pos_with_country, is_onsite):
            hour_rate = None; rate_source = None
            if country == 'TUN':
                tunis = get_tunis_rate_by_name(db, emp_name)
                if tunis:
                    hour_rate = tunis['hour_rate']; rate_source = 'tunis_rates_db'
            elif is_onsite:
                raw_pos = (base_pos_with_country or '').replace(f'{country} - ', '')
                onsite_pos = _resolve_onsite_position(raw_pos, country)
                if onsite_pos:
                    pi = _lookup_pos_fuzzy(onsite_pos)
                    if pi:
                        hour_rate = pi['hour_rate']; rate_source = 'positions_db_onsite'
                if hour_rate is None and sar_rate and sar_rate > 0:
                    hour_rate = round(sar_rate / SAR_TO_USD, 2); rate_source = 'odoo_fallback'
            else:
                # Always try DB first (handles promotions — each segment has different position)
                if base_pos_with_country:
                    pi = _lookup_pos_fuzzy(base_pos_with_country)
                    if pi:
                        hour_rate = pi['hour_rate']; rate_source = 'positions_db'
                # Fallback to Odoo SAR rate ONLY if no DB match
                if hour_rate is None and sar_rate and sar_rate > 0:
                    hour_rate = round(sar_rate / SAR_TO_USD, 2); rate_source = 'odoo'
            overtime_rate = round(hour_rate * OVERTIME_MULTIPLIER, 2) if hour_rate else None
            return hour_rate, overtime_rate, rate_source

        # ── Per-month segmentation ──
        from collections import defaultdict as _dd2
        segment_data = _dd2(lambda: _dd2(lambda: {'regular': 0.0, 'ramadan': 0.0, 'overtime': 0.0}))
        weekend = TUNIS_WEEKEND if country == 'TUN' else WEEKEND_DAYS

        for m in months:
            mkey = m['key']
            cell = months_data.get(mkey, {'regular': 0, 'ramadan': 0, 'overtime': 0})
            total_cell = cell['regular'] + cell['ramadan'] + cell['overtime']
            if total_cell == 0:
                continue

            seg_counts = _dd2(int)
            days_in_month = _cal.monthrange(m['year'], m['month'])[1]
            for day in range(1, days_in_month + 1):
                try: d_obj = _date_cls2(m['year'], m['month'], day)
                except: continue
                if d_obj.weekday() in weekend: continue
                d_str = f"{m['year']:04d}-{m['month']:02d}-{day:02d}"
                if is_public_holiday(d_str, country): continue  # holidays → overtime, skip from working day count
                raw_pos = _base_position_on_date(d_str)
                base_p = f"{country} - {raw_pos}" if raw_pos else None
                is_on = _date_is_onsite(d_str)
                seg_counts[(base_p, is_on)] += 1

            total_working = sum(seg_counts.values())
            if total_working == 0:
                raw_pos = _base_position_on_date(mkey + '-15')
                base_p = f"{country} - {raw_pos}" if raw_pos else None
                seg_counts[(base_p, False)] = 1
                total_working = 1

            for (base_p, is_on), cnt in seg_counts.items():
                ratio = cnt / total_working
                segment_data[(base_p, is_on)][mkey]['regular']  += cell['regular']  * ratio
                segment_data[(base_p, is_on)][mkey]['ramadan']  += cell['ramadan']  * ratio
                segment_data[(base_p, is_on)][mkey]['overtime'] += cell['overtime'] * ratio

        # ── Build one row per segment ──
        if not segment_data:
            base_p = f"{country} - {odoo_pos}" if odoo_pos else '—'
            hr, ot_r, src = _get_rate_for_pos(base_p, False)
            employees_out.append({
                'name': emp_display, 'full_name': emp_name, 'code': emp_code, 'country': country,
                'position': base_p, 'hour_rate': hr, 'overtime_rate': ot_r, 'rate_source': src,
                'sar_rate': sar_rate, 'is_onsite': False,
                'months': {m['key']: {'regular':0,'ramadan':0,'overtime':0,'total':0} for m in months},
                'total_hours': 0, 'total_cost_usd': 0, 'current_mds': 0,
            })
            continue

        for (base_p, is_on), mkey_data in sorted(segment_data.items(), key=lambda x: (x[0][0] or '', x[0][1])):
            full_months = {}
            for m in months:
                c = mkey_data.get(m['key'], {'regular': 0, 'ramadan': 0, 'overtime': 0})
                full_months[m['key']] = {
                    'regular':  round(c['regular'], 2),
                    'ramadan':  round(c['ramadan'], 2),
                    'overtime': round(c['overtime'], 2),
                    'total':    round(c['regular'] + c['ramadan'] + c['overtime'], 2),
                }
            seg_reg = sum(v['regular']  for v in full_months.values())
            seg_ram = sum(v['ramadan']  for v in full_months.values())
            seg_ot  = sum(v['overtime'] for v in full_months.values())
            if seg_reg + seg_ram + seg_ot < 0.01:
                continue

            hr, ot_r, src = _get_rate_for_pos(base_p, is_on)
            if is_on:
                raw_pos = (base_p or '').replace(f'{country} - ', '')
                display_pos = _resolve_onsite_position(raw_pos, country) or ((base_p or '') + ' - onsite')
            else:
                display_pos = base_p or '—'

            employees_out.append({
                'name': emp_display + (' — Onsite' if is_on else ''),
                'full_name': emp_name, 'code': emp_code, 'country': country,
                'position': display_pos, 'hour_rate': hr, 'overtime_rate': ot_r,
                'rate_source': src, 'sar_rate': sar_rate, 'is_onsite': is_on,
                'months': full_months,
                'total_hours': round(seg_reg + seg_ram + seg_ot, 2),
                'total_cost_usd': _calc_cost(seg_reg, seg_ram, seg_ot, hr, ot_r),
                'current_mds': _calc_mds(seg_reg, seg_ram, seg_ot),
            })

    # Sort: name alpha, onsite after regular
    employees_out.sort(key=lambda x: (
        x['name'].lower().replace(' — onsite', ''),
        1 if x.get('is_onsite') else 0
    ))

    # Build per-month cost totals for profitability
    month_cost_totals = {m['key']: 0.0 for m in months}
    month_md_totals   = {m['key']: 0.0 for m in months}
    for emp in employees_out:
        hr  = emp.get('hour_rate') or 0
        otr = emp.get('overtime_rate') or hr * 1.5
        for m in months:
            cell = emp['months'].get(m['key'], {'regular':0,'ramadan':0,'overtime':0})
            reg, ram, ot = cell['regular'], cell['ramadan'], cell['overtime']
            month_md_totals[m['key']]   += (reg + ot) / 8 + ram / 6
            if hr > 0:
                month_cost_totals[m['key']] += (reg + ram) * hr + ot * otr
            elif emp.get('total_cost_usd') and emp.get('total_hours', 0) > 0:
                frac = (reg + ram + ot) / emp['total_hours']
                month_cost_totals[m['key']] += emp['total_cost_usd'] * frac

    return jsonify({
        'phase': phase_key,
        'months': months,
        'employees': employees_out,
        'month_cost_usd': month_cost_totals,
        'month_mds': month_md_totals,
        'project_start_month': months[0]['key'] if months else None,
        'total_employees': len(employees_out),
    })


@app.route('/api/plan-overrides', methods=['GET', 'POST'])
def api_plan_overrides_get():
    if request.method == 'POST':
        body = request.json or {}
        phase     = body.get('phase', '')
        month_key = body.get('month_key', '')
        field     = body.get('field', '')
        value     = body.get('value')
        if not phase or not month_key or not field:
            return jsonify({'error': 'phase, month_key, field required'}), 400
        # Save: namespace=plan, subkey=phase, key=month_key.field
        proj_set_override('plan', phase, f'{month_key}.{field}', value)
        return jsonify({'ok': True, 'phase': phase, 'month_key': month_key, 'field': field})
    return jsonify(load_plan_overrides())


@app.route('/api/position-overrides', methods=['POST'])
def api_position_overrides_save():
    """Manual position override for an employee (when Odoo doesn't have it)"""
    body = request.json or {}
    name = body.get('name')
    position = body.get('position')
    if not name:
        return jsonify({'error': 'name required'}), 400
    db.set_override('position', '', name, position if position else None)
    return jsonify({'ok': True})


@app.route('/api/plan-overrides/backup')
def api_overrides_backup():
    """Download all overrides as JSON for manual backup"""
    data = {
        'plan_overrides': load_plan_overrides(),
        'travel_records': load_travel(),
        'exported_at': datetime.now().isoformat(),
    }
    import json
    return Response(
        json.dumps(data, indent=2, ensure_ascii=False),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="dashboard_backup_{date.today().isoformat()}.json"'}
    )

@app.route('/api/plan-overrides/restore', methods=['POST'])
def api_overrides_restore():
    """Upload a backup JSON to restore overrides + travel records"""
    body = request.json or {}
    if not body:
        return jsonify({'error': 'Empty body'}), 400
    if 'plan_overrides' in body:
        save_plan_overrides(body['plan_overrides'])
    if 'travel_records' in body:
        save_travel(body['travel_records'])
    return jsonify({'ok': True, 'restored': {
        'plan_overrides': len(body.get('plan_overrides', {})),
        'travel_records': len(body.get('travel_records', [])),
    }})


@app.route('/api/storage-info')
def api_storage_info():
    """Diagnostic endpoint to verify storage is working"""
    info = {
        'persist_dir': PERSIST_DIR,
        'persist_dir_exists': os.path.isdir(PERSIST_DIR),
        'persist_dir_writable': os.access(PERSIST_DIR, os.W_OK),
        'data_path_env': os.environ.get('DATA_PATH'),
        'files': [],
        'db_stats': db.get_stats(),
    }
    if info['persist_dir_exists']:
        for f in os.listdir(PERSIST_DIR):
            full = os.path.join(PERSIST_DIR, f)
            if os.path.isfile(full):
                info['files'].append({
                    'name': f,
                    'size_bytes': os.path.getsize(full),
                    'modified': datetime.fromtimestamp(os.path.getmtime(full)).isoformat(),
                })
    return jsonify(info)


@app.route('/api/db/audit')
def api_db_audit():
    """Get recent changes for debugging"""
    limit = int(request.args.get('limit', 50))
    return jsonify({
        'recent_changes': db.get_recent_audit(limit=limit),
    })



@app.route('/api/project-employees')
def api_project_employees():
    """Get list of employees who have logged time on the project, with their positions"""
    raw = odoo.get_timesheets()
    if raw is None:
        return jsonify({'employees': [], 'connected': False})

    seen = {}
    import re as _re_emp
    for e in raw:
        emp = e.get('employee_id')
        if emp and emp[1]:
            full_name = emp[1]
            display_name = _re_emp.sub(r'^\s*\[[A-Z]\d+\s*\]\s*', '', full_name).strip()
            if full_name not in seen:
                seen[full_name] = {'name': display_name, 'full_name': full_name}

    # Try to fetch positions
    overrides = load_plan_overrides().get('position_overrides', {})
    for name, info in seen.items():
        if name in overrides:
            info['position'] = overrides[name]
            info['source'] = 'override'
        else:
            pos = get_odoo_position_for_employee(name)
            if pos:
                info['position'] = pos
                info['source'] = 'odoo'
            else:
                info['position'] = None
                info['source'] = None

    employees = sorted(seen.values(), key=lambda x: x['name'].lower())
    return jsonify({'employees': employees, 'connected': True, 'count': len(employees)})


# Main API endpoint for variance data
@app.route('/api/variance')
def api_variance():
    """Returns full variance structure"""
    out = {'tabs': {}, 'available': os.path.exists(VARIANCE_FILE)}
    if not out['available']:
        return jsonify(out)

    try:
        for tab_key, tab_info in VARIANCE_TABS.items():
            if tab_key == 'travel':
                continue  # handled separately
            tab_data = {'label': tab_info['label'], 'sections': []}
            for sect in tab_info['sections']:
                try:
                    df = pd.read_excel(VARIANCE_FILE, sheet_name=sect['sheet'], header=None)
                    parser = sect['parser']
                    if parser == 'budget':
                        data = parse_budget_sheet(df)
                        # Apply persistent overrides (auto-saved values)
                        data = apply_budget_overrides(data, tab_key)
                    elif parser == 'profitability':
                        data = parse_profitability_sheet(df)
                    elif parser == 'effort':
                        data = parse_effort_sheet(df)
                    elif parser == 'estimated':
                        data = parse_estimated_sheet(df)
                    else:
                        data = {}
                    tab_data['sections'].append({
                        'key': sect['key'],
                        'label': sect['label'],
                        'sheet': sect['sheet'],
                        'data': data,
                    })
                except Exception as e:
                    logger.error(f"Variance parse {sect['sheet']}: {e}")
                    tab_data['sections'].append({
                        'key': sect['key'],
                        'label': sect['label'],
                        'sheet': sect['sheet'],
                        'error': str(e),
                    })
            out['tabs'][tab_key] = tab_data
    except Exception as e:
        logger.error(f"Variance error: {e}\n{traceback.format_exc()}")
        out['error'] = str(e)
    return jsonify(out)

@app.route('/api/variance/export')
def api_variance_export():
    """Export variance Excel with current overrides applied"""
    from flask import send_file
    import shutil
    if not os.path.exists(VARIANCE_FILE):
        return jsonify({'error': 'File not found'}), 404

    # Copy original to a temp file in PERSIST_DIR
    out_path = os.path.join(PERSIST_DIR, f'variance_export_{date.today().isoformat()}.xlsx')
    try:
        shutil.copyfile(VARIANCE_FILE, out_path)

        # Apply overrides — open with openpyxl to preserve formatting
        try:
            from openpyxl import load_workbook
            wb = load_workbook(out_path)
            overrides_data = load_plan_overrides()
            plan_overrides = overrides_data.get('plan_overrides', {})

            # Map phase to sheet name
            phase_to_sheet = {
                'development': 'Profitability - Development',
                'consultation': 'Profitability - Consultation',
            }

            for phase_key, sheet_name in phase_to_sheet.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                phase_overrides = plan_overrides.get(phase_key, {})

                # Header at row 6 (1-indexed in openpyxl), data starts row 8
                # Find Month col (col A) and key columns
                # Reading row 6 to find column indices
                header_row = 6
                col_indices = {}
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=header_row, column=col).value
                    if val:
                        col_indices[str(val).strip()] = col

                completion_col = col_indices.get('% Completion from plan')
                remaining_col = col_indices.get('Estimated Remaining (MD)')

                # Walk rows 8+ and apply overrides
                for row in range(8, ws.max_row + 1):
                    month_val = ws.cell(row=row, column=1).value
                    if not month_val:
                        continue
                    if isinstance(month_val, datetime):
                        month_key = month_val.strftime('%Y-%m')
                    else:
                        month_key = str(month_val)[:7]

                    if month_key in phase_overrides:
                        mo = phase_overrides[month_key]
                        if isinstance(mo, dict):
                            if 'completion' in mo and completion_col:
                                # Stored as percentage, convert back to fraction
                                ws.cell(row=row, column=completion_col).value = mo['completion'] / 100
                            if 'remaining' in mo and remaining_col:
                                ws.cell(row=row, column=remaining_col).value = mo['remaining']

            wb.save(out_path)

            # Auto-fill "Current Effort - Development" from Odoo timesheets
            try:
                fill_current_effort_from_odoo(out_path)
            except Exception as e:
                logger.warning(f"Could not auto-fill current effort from Odoo: {e}")
        except Exception as e:
            logger.warning(f"Could not apply overrides to export: {e}")

    except Exception as e:
        logger.error(f"Export failed: {e}")
        return jsonify({'error': str(e)}), 500

    return send_file(
        out_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'BOG_Variance_{date.today().isoformat()}.xlsx'
    )


def fill_current_effort_from_odoo(xlsx_path):
    """Read Odoo timesheets and fill the 'Current Effort - Development' sheet.
    Sheet structure:
      - Row 4: Month names (March, April, May, ...) at columns G, J, M, P, S, V, Y, AB, AE, AH, AK
      - Row 5: Per-month sub-headers: Regular Time (MH), Ramdan Hours, Over Time (MH)
      - Row 6+: People rows (col B=Name, col D=Position)
    For each person, for each month, we fill:
      - Regular Time (MH) = total hours minus Ramadan and Overtime
      - Ramdan Hours = hours logged during Ramadan period (KSA: Feb 18-Mar 19; EGY: Feb 19-Mar 20)
      - Over Time (MH) = hours logged on weekends (Fri+Sat for KSA/EGY)
    """
    if not odoo.uid:
        if not odoo.connect():
            logger.warning("Odoo not available for auto-fill")
            return

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    sheet_name = 'Current Effort - Development'
    if sheet_name not in wb.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found")
        return
    ws = wb[sheet_name]

    # Map month name → starting column (Regular Time col index)
    # Row 4 has month names spaced every 3 columns starting from G (col 7)
    month_columns = {}  # 'March' -> col_index of "Regular Time (MH)"
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if v and isinstance(v, str):
            month_columns[v.strip()] = col

    if not month_columns:
        logger.warning("No month columns found in row 4")
        return

    logger.info(f"Found month columns: {list(month_columns.keys())}")

    # Read people from rows 6+ (col B = Name, col D = Position)
    people = []
    for r in range(6, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name or not isinstance(name, str):
            continue
        people.append({
            'row': r,
            'name': name.strip(),
            'position': (ws.cell(row=r, column=4).value or '').strip() if ws.cell(row=r, column=4).value else '',
        })

    if not people:
        logger.warning("No people found in sheet")
        return

    logger.info(f"Found {len(people)} people in Current Effort sheet")

    # Get all timesheets for this project for the development phase
    project_id = None
    if PROJECT_NAME:
        try:
            projects = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'search_read',
                [[('name', 'ilike', PROJECT_NAME)]],
                {'fields': ['id', 'name'], 'limit': 5}
            )
            if projects:
                project_id = projects[0]['id']
        except Exception as e:
            logger.warning(f"Could not find project: {e}")

    if not project_id:
        logger.warning("Could not resolve project_id, skipping auto-fill")
        return

    # Fetch all timesheets for this project (within Development phase)
    try:
        # Get tasks under Development Phase
        dev_phase_names = get_phase_mapping().get('development', ['Development Phase'])
        phases = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('name', 'in', dev_phase_names)]],
            {'fields': ['id', 'name'], 'limit': 50}
        )
        phase_ids = [p['id'] for p in phases]

        # Get parent tasks under these phases
        parent_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('phase_id', 'in', phase_ids), ('project_id', '=', project_id)]],
            {'fields': ['id', 'child_ids'], 'limit': 5000}
        )
        parent_task_ids = {t['id'] for t in parent_tasks}

        # Get all sub-tasks too (recursively walk down)
        all_relevant_task_ids = set(parent_task_ids)
        all_tasks_in_proj = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('project_id', '=', project_id)]],
            {'fields': ['id', 'parent_id'], 'limit': 10000}
        )
        # Walk parent chains
        task_by_id = {t['id']: t for t in all_tasks_in_proj}
        for t in all_tasks_in_proj:
            cur = t
            visited = set()
            while cur and cur['id'] not in visited:
                visited.add(cur['id'])
                if cur['id'] in parent_task_ids:
                    all_relevant_task_ids.add(t['id'])
                    break
                if not cur.get('parent_id'):
                    break
                pid = cur['parent_id'][0] if isinstance(cur['parent_id'], list) else cur['parent_id']
                cur = task_by_id.get(pid)

        logger.info(f"Found {len(all_relevant_task_ids)} relevant tasks for development phase")

        # Now fetch timesheets for these tasks
        timesheets = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.analytic.line', 'search_read',
            [[('task_id', 'in', list(all_relevant_task_ids))]],
            {'fields': ['employee_id', 'date', 'unit_amount'], 'limit': 50000}
        )
        logger.info(f"Fetched {len(timesheets)} timesheet entries")
    except Exception as e:
        logger.error(f"Could not fetch timesheets: {e}")
        return

    # Aggregate timesheets per person per month, splitting into Regular/Ramadan/Overtime
    from collections import defaultdict
    # person_data[name][month_name] = {'regular': h, 'ramadan': h, 'overtime': h}
    person_data = defaultdict(lambda: defaultdict(lambda: {'regular': 0, 'ramadan': 0, 'overtime': 0}))

    MONTH_NAMES = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']

    for ts in timesheets:
        emp = ts.get('employee_id', [None, ''])
        emp_name = emp[1] if isinstance(emp, list) and len(emp) > 1 else ''
        if not emp_name:
            continue
        date_str = ts.get('date')
        if not date_str:
            continue
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            continue
        hours = float(ts.get('unit_amount') or 0)
        if hours <= 0:
            continue

        month_name = MONTH_NAMES[d.month - 1]

        # Determine country from employee name (KSA/EGY/TUN)
        country = get_country_from_employee_name(emp_name)

        # Check if Ramadan (period stored in RAMADAN_RANGES)
        is_ramadan = False
        ramadan_range = RAMADAN_RANGES.get(country)
        if ramadan_range:
            try:
                ram_start = datetime.strptime(ramadan_range['start'], '%Y-%m-%d').date()
                ram_end = datetime.strptime(ramadan_range['end'], '%Y-%m-%d').date()
                if ram_start <= d <= ram_end:
                    is_ramadan = True
            except Exception:
                pass

        # Check if weekend → overtime
        # Tunis: Sat+Sun (5,6); KSA/EGY: Fri+Sat (4,5)
        weekend_days = TUNIS_WEEKEND if country == 'TUN' else WEEKEND_DAYS
        is_weekend = d.weekday() in weekend_days
        is_holiday = is_public_holiday(date_str, country)

        if is_weekend or is_holiday:
            person_data[emp_name][month_name]['overtime'] += hours
        elif is_ramadan:
            person_data[emp_name][month_name]['ramadan'] += hours
        else:
            person_data[emp_name][month_name]['regular'] += hours

    logger.info(f"Aggregated data for {len(person_data)} people")

    # Now write into the Excel sheet
    # For each person, find their row by name (fuzzy match - strip [code] prefix etc.)
    def normalize_name(s):
        # Remove [E123] prefix and lowercase
        import re
        return re.sub(r'\[[A-Z]\d+\]\s*', '', s).strip().lower()

    sheet_name_to_row = {}
    for p in people:
        sheet_name_to_row[normalize_name(p['name'])] = p['row']

    filled_count = 0
    for emp_name, months_data in person_data.items():
        norm = normalize_name(emp_name)
        # Try exact match
        target_row = sheet_name_to_row.get(norm)
        if not target_row:
            # Fuzzy match: any sheet name that contains the emp name (or vice versa)
            for sheet_norm, row_idx in sheet_name_to_row.items():
                if sheet_norm == norm:
                    target_row = row_idx
                    break
                # Match first 2 words
                emp_words = norm.split()[:2]
                sheet_words = sheet_norm.split()[:2]
                if emp_words and sheet_words and emp_words[0] == sheet_words[0]:
                    if len(emp_words) > 1 and len(sheet_words) > 1 and emp_words[1] == sheet_words[1]:
                        target_row = row_idx
                        break
        if not target_row:
            logger.info(f"  Skipping {emp_name} - not in sheet")
            continue

        for month_name, hrs in months_data.items():
            if month_name not in month_columns:
                continue
            base_col = month_columns[month_name]
            # base_col = Regular Time, base_col+1 = Ramadan, base_col+2 = Overtime
            if hrs['regular'] > 0:
                ws.cell(row=target_row, column=base_col).value = round(hrs['regular'], 2)
            if hrs['ramadan'] > 0:
                ws.cell(row=target_row, column=base_col + 1).value = round(hrs['ramadan'], 2)
            if hrs['overtime'] > 0:
                ws.cell(row=target_row, column=base_col + 2).value = round(hrs['overtime'], 2)
            filled_count += 1

    logger.info(f"Filled {filled_count} person-month cells")

    wb.save(xlsx_path)
    logger.info(f"Saved updated workbook to {xlsx_path}")

# ============================================================
# TRAVEL & ONSITE — stored in DB
# ============================================================
def load_travel():
    """Load travel records — reads from global storage (manage page).
    Falls back to db.list_travel() for legacy records already in the old store.
    """
    # Primary: global storage (set via /manage page)
    global_recs = _global_get('travel')
    if global_recs is not None:
        return global_recs
    # Fallback: old db.list_travel() (BOG legacy)
    try:
        old = db.list_travel()
        if old:
            return old
    except Exception:
        pass
    return []


def save_travel(records):
    """Bulk save (used by restore endpoint)."""
    if isinstance(records, list):
        for r in records:
            if isinstance(r, dict) and r.get('id') is not None:
                db.upsert_travel(str(r['id']), r)

# ============================================================
# PROMOTIONS — track when employees get promoted mid-project
# ============================================================
def load_promotions():
    """Load promotions — reads from global storage (manage page) first."""
    global_recs = _global_get('promotions')
    if global_recs is not None:
        # Normalize field names: manage page uses 'effective_date', old code uses 'promotion_date'
        for r in global_recs:
            # Normalize date: use effective_date, fallback to year-01-01
            eff = r.get('effective_date') or ''
            yr  = r.get('year')
            if not eff and yr:
                eff = f'{yr}-01-01'  # fallback: start of promotion year
            r['promotion_date'] = eff
            if 'effective_date' not in r or not r['effective_date']:
                r['effective_date'] = eff

            # Map title fields for compatibility
            if 'new_title' in r:
                r['new_position'] = r['new_title']
            if r.get('old_title'):
                r['old_position'] = r['old_title']
            if 'new_title' in r and 'position' not in r:
                r['position'] = r['new_title']
        global_recs.sort(key=lambda x: (x.get('name',''), x.get('promotion_date','')))
        return global_recs
    # Fallback: old promotions namespace
    try:
        raw = db.get_namespace_overrides('promotions', '')
        records = []
        for k, v in raw.items():
            if isinstance(v, dict):
                records.append(v)
        records.sort(key=lambda x: (x.get('name',''), x.get('promotion_date','')))
        return records
    except Exception:
        return []


@app.route('/api/promotions', methods=['GET'])
def api_promotions_list():
    return jsonify({'records': load_promotions()})


@app.route('/api/promotions/<int:rec_id>', methods=['PUT'])
def api_promotions_update(rec_id):
    body = request.json or {}
    records = load_promotions()
    for r in records:
        if int(r.get('id', 0)) == rec_id:
            for k in ['name', 'old_position', 'new_position', 'promotion_date', 'notes']:
                if k in body:
                    r[k] = body[k]
            r['updated_at'] = datetime.now().isoformat()
            db.set_override('promotions', '', str(rec_id), r)
            return jsonify({'ok': True, 'record': r})
    return jsonify({'error': 'not found'}), 404


@app.route('/api/promotions/employee-odoo-position')
def api_employee_odoo_position():
    """Get current Odoo position + suggest previous level."""
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    current_pos = get_odoo_position_for_employee(name)
    suggested_old = None
    if current_pos:
        import re as _rp
        p = current_pos.lower()
        if 'lead' in p:
            suggested_old = _rp.sub(r'(?i)lead\s+', 'Sr ', current_pos).strip()
        elif 'senior' in p:
            suggested_old = _rp.sub(r'(?i)senior\s+', '', current_pos).strip()
        elif _rp.search(r'(?i)\bsr\.?\s+', current_pos):
            suggested_old = _rp.sub(r'(?i)\bsr\.?\s+', '', current_pos).strip()
        else:
            suggested_old = current_pos
    return jsonify({
        'name': name,
        'current_position': current_pos,
        'suggested_old_position': suggested_old,
    })

@app.route('/api/budget-changes', methods=['GET', 'POST'])
def api_budget_changes_get():
    if request.method == 'POST':
        body = request.json or {}
        phase = body.get('phase', 'development')
        changes = body.get('changes', [])
        proj_set_override('budget_changes', '', phase, changes)
        return jsonify({'ok': True, 'phase': phase, 'saved': len(changes)})

    phase = request.args.get('phase', 'development')
    changes = proj_get_override('budget_changes', '', phase) or []
    if isinstance(changes, str):
        import json as _j
        try: changes = _j.loads(changes)
        except: changes = []
    history = db.get_override('planned_profit_history', '', phase) or {}
    return jsonify({'phase': phase, 'changes': changes if isinstance(changes, list) else [],
                    'planned_profit_history': history if isinstance(history, dict) else {}})


@app.route('/api/estimated-rows', methods=['GET', 'POST'])
def api_estimated_rows_get():
    """Get or save estimated cost rows for a phase."""
    if request.method == 'POST':
        body = request.json or {}
        phase = body.get('phase', 'development')
        rows  = body.get('rows', [])
        proj_set_override('estimated_rows', '', phase, rows)
        return jsonify({'ok': True, 'phase': phase, 'saved': len(rows)})

    phase = request.args.get('phase', 'development')
    rows = proj_get_override('estimated_rows', '', phase) or []
    if isinstance(rows, str):
        import json as _json
        try: rows = _json.loads(rows)
        except: rows = []
    return jsonify({'phase': phase, 'rows': rows if isinstance(rows, list) else []})


def _categorize_invoice_line_by_desc(desc_name):
    """
    Categorize a BOG invoice line using ONLY the line description.
    The description is the reliable field — product is too generic.

    Rules (exact match priority, then keyword):
      "Business Re Engineering"  → consultation
      "Development"              → development
      "Operation and Support"    → support
      "License"                  → license (skip)
      anything else              → other

    Keywords (case-insensitive):
      consultation: business re, process re, reengineering, re-engineering,
                    consultation, analysis, roadmap, assessment, design phase
      support:      operation, support
      development:  development
      license:      license
    """
    d = (desc_name or '').lower().strip()
    if not d:
        return 'other'

    CONSULT_KW = ['business re', 'process re', 'reengineering', 're-engineering',
                  'consultation', 'analysis', 'roadmap', 'assessment', 'design phase']
    SUPPORT_KW = ['operation and support', 'operation & support', 'support phase',
                  'operation', 'support']
    DEV_KW     = ['development']
    LIC_KW     = ['license']

    if any(k in d for k in LIC_KW):
        return 'license'
    if any(k in d for k in CONSULT_KW):
        return 'consultation'
    # Support checked before dev because "operation and support" has no dev keyword
    if any(k in d for k in SUPPORT_KW):
        return 'support'
    if any(k in d for k in DEV_KW):
        return 'development'
    return 'other'


@app.route('/api/invoices')
def api_invoices():
    """
    Get all validated (posted) invoices whose lines belong to project
    "BOG Digital Transformation 25", categorized by line description only.

    Strategy:
      1. Find the Odoo project by name → get its id
      2. Find all account.move.line linked to tasks/analytic of that project
         (via sale_line_ids.order_id.project_ids OR analytic_account_id)
         — actually simplest: find posted invoices linked to the known SO list
           S00201, S00166, S00204 + any SO whose project_ids contains our project id.
      3. For each invoice line, categorize by description ONLY.
      4. Only lines that categorize as dev/consultation/support are included;
         license lines and truly-other lines are excluded from totals.

    Returns: phase totals + per-invoice list + monthly + monthly_cumulative
    """
    try:
        if not odoo.uid:
            if not odoo.connect():
                return jsonify({'error': 'Odoo not connected',
                                'development': 0, 'consultation': 0, 'support': 0,
                                'invoices': [], 'monthly': {}}), 503

        # ── Step 1: find the project id ─────────────────────────────────────────
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', PROJECT_NAME)]],
            {'fields': ['id', 'name'], 'limit': 5}
        )
        project_ids = [p['id'] for p in projects]
        logger.info(f"BOG projects found: {[(p['id'], p['name']) for p in projects]}")

        # ── Step 2: find all SOs linked to this SPECIFIC project ──────────────
        # We use the project id for exact matching to avoid picking up SOs
        # from other projects that happen to share an analytic account.
        so_ids = set()
        so_names_found = []
        if project_ids:
            try:
                # Direct many2many: sale.order.project_ids
                sos = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'sale.order', 'search_read',
                    [[('project_ids', 'in', project_ids)]],
                    {'fields': ['name', 'id', 'project_ids'], 'limit': 200}
                )
                # Only keep SOs whose project_ids actually contain our project — exact match
                for so in sos:
                    proj_ids_on_so = so.get('project_ids', [])
                    if any(pid in proj_ids_on_so for pid in project_ids):
                        so_ids.add(so['id'])
                        so_names_found.append(so['name'])
            except Exception as _e:
                logger.warning(f"SO lookup by project_ids failed: {_e}")

        logger.info(f"SOs for BOG project (exact match): {so_names_found}")

        # ── Step 3: get posted invoices for those SOs ───────────────────────────
        if not so_ids:
            return jsonify({'error': 'No sale orders found for project',
                            'development': 0, 'consultation': 0, 'support': 0,
                            'invoices': [], 'monthly': {}, 'monthly_cumulative': {}}), 404

        domain = [
            ('move_type', '=', 'out_invoice'),
            ('state',     '=', 'posted'),
            ('invoice_line_ids.sale_line_ids.order_id', 'in', list(so_ids)),
        ]
        invoices_raw = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.move', 'search_read',
            [domain],
            {'fields': ['name', 'invoice_date', 'amount_untaxed',
                        'invoice_line_ids', 'invoice_origin', 'state'],
             'limit': 500,
             'order': 'invoice_date asc'}
        )
        logger.info(f"Posted invoices for BOG: {len(invoices_raw)}")

        # ── Step 4: fetch & categorize invoice lines by description only ─────────
        result = {
            'development':       0.0,
            'consultation':      0.0,
            'support':           0.0,
            'license_excluded':  0.0,   # informational — not in totals
            'other_unmatched':   0.0,   # lines with unknown desc — for debugging
            'invoices':          [],
            'monthly':           {},
            'monthly_cumulative': {},
            'sale_orders':       so_names_found,
        }

        for inv in invoices_raw:
            line_ids = inv.get('invoice_line_ids', [])
            if not line_ids:
                continue

            lines = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'account.move.line', 'search_read',
                [[('id', 'in', line_ids),
                  ('display_type', 'not in', ['line_section', 'line_note']),
                  # Only lines linked to our SOs (filters out lines from other projects
                  # that share the same invoice header)
                  ('sale_line_ids.order_id', 'in', list(so_ids))]],
                {'fields': ['name', 'price_subtotal']}
            )

            inv_dev = inv_con = inv_sup = inv_lic = inv_other = 0.0
            for line in lines:
                desc   = line.get('name') or ''
                amount = float(line.get('price_subtotal') or 0)
                cat    = _categorize_invoice_line_by_desc(desc)

                logger.info(f"  [{inv['name']}] desc='{desc}' → {cat} ({amount:.2f})")

                if   cat == 'license':      inv_lic   += amount; result['license_excluded'] += amount
                elif cat == 'consultation': inv_con   += amount; result['consultation']     += amount
                elif cat == 'support':      inv_sup   += amount; result['support']          += amount
                elif cat == 'development':  inv_dev   += amount; result['development']      += amount
                else:                       inv_other += amount; result['other_unmatched']  += amount

            inv_total = inv_dev + inv_con + inv_sup   # excl. license & other
            inv_date  = inv.get('invoice_date', '') or ''
            month_key = inv_date[:7] if inv_date else 'unknown'

            if inv_total > 0:
                result['invoices'].append({
                    'name':           inv['name'],
                    'date':           inv_date,
                    'month':          month_key,
                    'amount_untaxed': round(inv_total, 2),
                    'development':    round(inv_dev,   2),
                    'consultation':   round(inv_con,   2),
                    'support':        round(inv_sup,   2),
                    'license':        round(inv_lic,   2),
                    'other':          round(inv_other, 2),
                    'origin':         inv.get('invoice_origin', ''),
                })

            if month_key != 'unknown' and inv_total > 0:
                m = result['monthly'].setdefault(month_key,
                        {'development': 0.0, 'consultation': 0.0, 'support': 0.0, 'total': 0.0})
                m['development']  += inv_dev
                m['consultation'] += inv_con
                m['support']      += inv_sup
                m['total']        += inv_total

        # ── Step 5: cumulative per month ────────────────────────────────────────
        cum_dev = cum_con = cum_sup = cum_tot = 0.0
        for mk in sorted(result['monthly'].keys()):
            m = result['monthly'][mk]
            cum_dev += m['development']
            cum_con += m['consultation']
            cum_sup += m['support']
            cum_tot += m['total']
            result['monthly_cumulative'][mk] = {
                'development':  round(cum_dev, 2),
                'consultation': round(cum_con, 2),
                'support':      round(cum_sup, 2),
                'total':        round(cum_tot, 2),
            }

        result['development']      = round(result['development'],      2)
        result['consultation']     = round(result['consultation'],      2)
        result['support']          = round(result['support'],          2)
        result['license_excluded'] = round(result['license_excluded'], 2)
        result['other_unmatched']  = round(result['other_unmatched'],  2)
        result['total']            = round(result['development'] + result['consultation'] + result['support'], 2)

        logger.info(f"Invoice totals — dev:{result['development']} "
                    f"con:{result['consultation']} sup:{result['support']} "
                    f"lic_excl:{result['license_excluded']} other:{result['other_unmatched']}")

        return jsonify(result)

    except Exception as e:
        logger.error(f"Invoices API error: {e}", exc_info=True)
        return jsonify({'error': str(e),
                        'development': 0, 'consultation': 0, 'support': 0,
                        'invoices': [], 'monthly': {}, 'monthly_cumulative': {}}), 500


@app.route('/api/invoices/by-phase/<phase>')
def api_invoices_by_phase(phase):
    """
    Get total issued invoices (excl. License) for a specific phase.
    Returns: total_sar (cumulative), monthly breakdown, cumulative per month, invoice list.
    """
    try:
        data = api_invoices().get_json()
        if 'error' in data:
            return jsonify(data), 500

        phase_key = phase.lower()
        total = data.get(phase_key, 0)

        # Monthly for this phase
        monthly = {}
        for mk, m in data.get('monthly', {}).items():
            v = m.get(phase_key, 0)
            if v:
                monthly[mk] = round(v, 2)

        # Cumulative for this phase
        monthly_cumulative = {}
        for mk, m in data.get('monthly_cumulative', {}).items():
            v = m.get(phase_key, 0)
            if v:
                monthly_cumulative[mk] = round(v, 2)

        # Invoices that have any amount in this phase
        invoices = [inv for inv in data.get('invoices', []) if inv.get(phase_key, 0) > 0]

        return jsonify({
            'phase':               phase_key,
            'total_sar':           round(total, 2),
            'monthly':             monthly,
            'monthly_cumulative':  monthly_cumulative,
            'invoices':            invoices,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/debug/invoices')
def debug_invoices():
    """
    Debug: shows every invoice line with its desc and category.
    Also shows which SOs were found for the project.
    Visit /debug/invoices in browser to verify.
    """
    try:
        if not odoo.uid:
            if not odoo.connect():
                return jsonify({'error': 'Odoo not connected'}), 503

        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', PROJECT_NAME)]],
            {'fields': ['id', 'name'], 'limit': 5}
        )
        project_ids = [p['id'] for p in projects]

        so_ids = set()
        so_names = []
        if project_ids:
            sos = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'sale.order', 'search_read',
                [[('project_ids', 'in', project_ids)]],
                {'fields': ['name', 'id'], 'limit': 200}
            )
            for so in sos:
                so_ids.add(so['id']); so_names.append(so['name'])

        so_ids = list(so_ids)
        domain = [('move_type','=','out_invoice'), ('state','=','posted'),
                  ('invoice_line_ids.sale_line_ids.order_id','in', so_ids)]
        invoices_raw = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'account.move', 'search_read', [domain],
            {'fields': ['name','invoice_date','amount_untaxed','invoice_line_ids','invoice_origin'],
             'limit': 500, 'order': 'invoice_date asc'}
        )

        out = {'projects': [(p['id'],p['name']) for p in projects],
               'sale_orders': so_names,
               'invoice_count': len(invoices_raw),
               'detail': []}

        for inv in invoices_raw:
            line_ids = inv.get('invoice_line_ids', [])
            lines_raw = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'account.move.line', 'search_read',
                [[('id','in',line_ids),
                  ('display_type','not in',['line_section','line_note']),
                  ('sale_line_ids.order_id','in',so_ids)]],
                {'fields': ['name','price_subtotal']}
            ) if line_ids else []

            parsed_lines = []
            for line in lines_raw:
                desc   = line.get('name') or ''
                amount = float(line.get('price_subtotal') or 0)
                cat    = _categorize_invoice_line_by_desc(desc)
                parsed_lines.append({'desc': desc, 'amount': amount, 'category': cat})

            out['detail'].append({
                'name':    inv['name'],
                'date':    inv.get('invoice_date',''),
                'origin':  inv.get('invoice_origin',''),
                'amount_untaxed': inv.get('amount_untaxed'),
                'lines':   parsed_lines,
            })

        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500



@app.route('/api/travel', methods=['GET'])
def api_travel_list():
    records = load_travel()
    today_str = date.today().isoformat()
    # Compute current status for each record
    for r in records:
        end = r.get('end_date')
        if not end:
            r['status'] = 'Onsite (open-ended)'
            r['days_onsite'] = (date.today() - datetime.strptime(r['start_date'], '%Y-%m-%d').date()).days + 1 if r.get('start_date') else 0
        else:
            try:
                end_d = datetime.strptime(end, '%Y-%m-%d').date()
                start_d = datetime.strptime(r['start_date'], '%Y-%m-%d').date()
                if end_d < date.today():
                    r['status'] = 'Returned'
                else:
                    r['status'] = 'Onsite'
                r['days_onsite'] = (end_d - start_d).days + 1
            except Exception:
                r['status'] = 'Unknown'
                r['days_onsite'] = 0
    return jsonify({'records': records, 'today': today_str})

@app.route('/api/travel/<int:rec_id>', methods=['PUT'])
def api_travel_update(rec_id):
    body = request.json or {}
    records = load_travel()
    for r in records:
        if int(r.get('id', 0)) == rec_id:
            for k in ['name', 'position', 'start_date', 'end_date', 'notes']:
                if k in body:
                    r[k] = body[k] or None if k == 'end_date' else body[k]
            db.upsert_travel(str(rec_id), r)
            return jsonify({'ok': True, 'record': r})
    return jsonify({'error': 'not found'}), 404

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})

@app.route('/debug/task-fields/<int:task_id>')
def debug_task_fields(task_id):
    """Returns ALL fields of a specific task to help debug assignment issues."""
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'error': 'Odoo unreachable'}), 500
    try:
        # Try to get the task with ALL standard fields
        task = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'read',
            [[task_id]],
            {}  # no fields filter = all fields
        )
        if not task:
            return jsonify({'error': f'Task {task_id} not found'}), 404
        return jsonify({
            'task_id': task_id,
            'task': task[0],
            'available_fields': list(task[0].keys()),
            'assignment_related': {
                k: v for k, v in task[0].items()
                if 'user' in k.lower() or 'assign' in k.lower() or 'owner' in k.lower()
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/debug/task-by-name')
def debug_task_by_name():
    """Find a task by name and return its fields. Use ?name=Env"""
    name = request.args.get('name', '')
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'error': 'Odoo unreachable'}), 500
    try:
        tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('name', 'ilike', name)]],
            {'fields': [], 'limit': 5}  # all fields
        )
        result = []
        for t in tasks:
            result.append({
                'id': t.get('id'),
                'name': t.get('name'),
                'available_fields': list(t.keys()),
                'assignment_related': {
                    k: v for k, v in t.items()
                    if 'user' in k.lower() or 'assign' in k.lower() or 'owner' in k.lower()
                }
            })
        return jsonify({'count': len(result), 'tasks': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/debug/effort-phase')
def debug_effort_phase():
    """Debug: check what tasks/timesheets are found for a phase key.
    Usage: /debug/effort-phase?phase=consultation
    """
    phase_key = request.args.get('phase', 'consultation')
    if not odoo.uid:
        if not odoo.connect():
            return jsonify({'error': 'Odoo unreachable'})

    phase_names = get_phase_mapping().get(phase_key, [])
    info = {'phase_key': phase_key, 'phase_names': phase_names}

    try:
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', PROJECT_NAME)]],
            {'fields': ['id', 'name'], 'limit': 5}
        )
        info['projects'] = [{'id': p['id'], 'name': p['name']} for p in projects]
        if not projects:
            return jsonify(info)
        project_id = projects[0]['id']

        phases = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.phase', 'search_read',
            [[('name', 'in', phase_names)]],
            {'fields': ['id', 'name']}
        )
        info['phases_found'] = [{'id': p['id'], 'name': p['name']} for p in phases]
        phase_ids = [p['id'] for p in phases]

        if not phase_ids:
            # Try ilike instead of exact match
            phases_ilike = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.phase', 'search_read',
                [[('name', 'ilike', 'Consultation')]],
                {'fields': ['id', 'name']}
            )
            info['phases_ilike_consultation'] = [{'id': p['id'], 'name': p['name']} for p in phases_ilike]

            all_phases = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.phase', 'search_read',
                [[('project_id', '=', project_id)]],
                {'fields': ['id', 'name']}
            )
            info['all_project_phases'] = [{'id': p['id'], 'name': p['name']} for p in all_phases]
            return jsonify(info)

        phase_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('phase_id', 'in', phase_ids), ('project_id', '=', project_id)]],
            {'fields': ['id', 'name', 'phase_id'], 'limit': 20}
        )
        info['phase_tasks_count'] = len(phase_tasks)
        info['phase_tasks_sample'] = [{'id': t['id'], 'name': t['name']} for t in phase_tasks[:5]]

        parent_ids = {t['id'] for t in phase_tasks}

        all_proj_tasks = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.task', 'search_read',
            [[('project_id', '=', project_id)]],
            {'fields': ['id', 'parent_id', 'phase_id'], 'limit': 10000}
        )
        relevant_ids = set(parent_ids)
        for t in all_proj_tasks:
            ph = t.get('phase_id')
            if ph and isinstance(ph, list) and ph[0] in phase_ids:
                relevant_ids.add(t['id'])

        info['relevant_task_ids_count'] = len(relevant_ids)

        if relevant_ids:
            ts_sample = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'account.analytic.line', 'search_read',
                [[('task_id', 'in', list(relevant_ids)[:100])]],
                {'fields': ['employee_id', 'date', 'unit_amount'], 'limit': 10}
            )
            info['timesheets_sample_count'] = len(ts_sample)
            info['timesheets_sample'] = ts_sample[:3]

    except Exception as e:
        info['error'] = str(e)

    return jsonify(info)


@app.route('/debug/timesheets')
def debug_timesheets():
    """Test the actual timesheet fetching path used by the app"""
    date_from = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    date_to = request.args.get('to', date.today().isoformat())

    info = {
        'date_from': date_from,
        'date_to': date_to,
        'project_name': PROJECT_NAME,
        'odoo_uid': odoo.uid,
        'odoo_last_error': odoo.last_error,
    }

    # Try fetching
    raw = odoo.get_timesheets(date_from=date_from, date_to=date_to)
    if raw is None:
        info['result'] = 'FAILED — odoo.get_timesheets returned None'
        info['odoo_last_error_after'] = odoo.last_error
    else:
        info['result'] = f'SUCCESS — {len(raw)} entries'
        info['sample_raw'] = raw[:2] if raw else []
        normalized = [normalize_timesheet(e) for e in raw[:3]]
        info['sample_normalized'] = normalized

    # Test with no date filter
    raw_all = odoo.get_timesheets()
    info['no_filter_count'] = len(raw_all) if raw_all is not None else 'FAILED'

    return jsonify(info)


@app.route('/debug')
def debug():
    info = {
        'base_dir': BASE_DIR,
        'data_file_exists': os.path.exists(DATA_FILE),
        'templates': sorted(os.listdir(os.path.join(BASE_DIR, 'templates'))) if os.path.exists(os.path.join(BASE_DIR, 'templates')) else 'NA',
        'partials': sorted(os.listdir(os.path.join(BASE_DIR, 'templates', 'partials'))) if os.path.exists(os.path.join(BASE_DIR, 'templates', 'partials')) else 'NA',
        'env_vars': {
            'ODOO_URL': ODOO_URL,
            'ODOO_DB': ODOO_DB,
            'ODOO_USERNAME_set': bool(ODOO_USERNAME),
            'ODOO_USERNAME_preview': ODOO_USERNAME[:3] + '***' if ODOO_USERNAME else None,
            'ODOO_PASSWORD_set': bool(ODOO_PASSWORD),
            'ODOO_PASSWORD_length': len(ODOO_PASSWORD) if ODOO_PASSWORD else 0,
            'PROJECT_NAME': PROJECT_NAME,
        },
        'odoo_test': {}
    }

    # Test Odoo connection step by step
    try:
        common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
        version = common.version()
        info['odoo_test']['1_server_reachable'] = True
        info['odoo_test']['2_odoo_version'] = version.get('server_version', 'unknown') if version else 'unknown'
    except Exception as e:
        info['odoo_test']['1_server_reachable'] = False
        info['odoo_test']['error'] = f"Cannot reach {ODOO_URL}: {str(e)}"
        return jsonify(info)

    if not ODOO_USERNAME or not ODOO_PASSWORD:
        info['odoo_test']['3_credentials'] = 'NOT SET — add ODOO_USERNAME and ODOO_PASSWORD in Railway Variables'
        return jsonify(info)

    try:
        common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
        uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})
        if uid:
            info['odoo_test']['3_auth'] = f'SUCCESS — uid={uid}'
        else:
            info['odoo_test']['3_auth'] = 'FAILED — check ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD'
            info['odoo_test']['hint'] = 'Most common: wrong ODOO_DB name OR using regular password instead of API key'
            return jsonify(info)
    except Exception as e:
        info['odoo_test']['3_auth'] = f'ERROR — {type(e).__name__}: {str(e)}'
        return jsonify(info)

    # Test fetching timesheets
    try:
        models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
        # Check if project exists
        projects = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', PROJECT_NAME)]],
            {'fields': ['name', 'id'], 'limit': 5}
        )
        info['odoo_test']['4_project_search'] = {
            'matches_found': len(projects),
            'projects': [{'id': p['id'], 'name': p['name']} for p in projects]
        }

        # Sample timesheets
        ts = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            'account.analytic.line', 'search_read',
            [[('project_id.name', 'ilike', _proj_name)]],
            {'fields': ['date', 'employee_id', 'project_id'], 'limit': 3}
        )
        info['odoo_test']['5_sample_timesheets'] = {
            'count': len(ts),
            'sample': ts[:2] if ts else []
        }
    except Exception as e:
        info['odoo_test']['4_data_fetch'] = f'ERROR — {type(e).__name__}: {str(e)}'

    return jsonify(info)


@app.route('/api/sales-orders')
def api_sales_orders():
    """Return all sale orders for current project with amounts, delivery %, and invoices."""
    # Accept line→variance tab mapping from frontend (user overrides)
    line_var_map = {}
    try:
        # Load from DB (saved via plan-overrides)
        pfx = active_db_prefix()
        plan_ns = f'{pfx}_plan' if pfx else 'plan'
        so_map_raw = db.get_namespace_overrides(plan_ns, 'so_line_map') or {}
        for line_id_key, fields in so_map_raw.items():
            if isinstance(fields, dict) and fields.get('var_tab'):
                line_var_map[str(line_id_key)] = fields['var_tab']
            elif isinstance(fields, str):
                line_var_map[str(line_id_key)] = fields
    except Exception:
        pass
    try:
        if not odoo.uid: odoo.connect()
        _proj_name = active_project_name()
        _proj_id   = session.get('project_id')

        # Find project in Odoo
        projects = odoo.models.execute_kw(
            ODOO_DB, odoo.uid, ODOO_PASSWORD,
            'project.project', 'search_read',
            [[('name', 'ilike', _proj_name)]],
            {'fields': ['id', 'name'], 'limit': 5}
        )
        if not projects:
            return jsonify({'ok': True, 'orders': [], 'summary': {}})

        project_ids = [p['id'] for p in projects]

        # Step 1: Get the analytic account linked to this project
        analytic_account_id = None
        analytic_account_name = ''
        try:
            proj_data = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'project.project', 'read',
                [project_ids],
                {'fields': ['id', 'name', 'analytic_account_id']}
            )
            for p in proj_data:
                aa = p.get('analytic_account_id')
                if aa and isinstance(aa, list):
                    analytic_account_id = aa[0]
                    analytic_account_name = aa[1]
                    break
            logger.info(f"Project analytic account: {analytic_account_name} (id={analytic_account_id})")
        except Exception as e:
            logger.warning(f"analytic account fetch failed: {e}")

        if not analytic_account_id:
            return jsonify({'ok': True, 'orders': [], 'summary': {},
                           'note': 'No analytic account linked to this project'})

        # Step 2: Get SOs linked to this analytic account
        sos = []
        try:
            sos = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'sale.order', 'search_read',
                [[('analytic_account_id', '=', analytic_account_id)]],
                {'fields': ['id', 'name', 'partner_id', 'date_order', 'state',
                            'amount_untaxed', 'amount_tax', 'amount_total',
                            'invoice_status', 'currency_id', 'order_line'], 'limit': 500}
            )
            logger.info(f"SOs by analytic account {analytic_account_name}: {len(sos)}")
        except Exception as e:
            logger.warning(f"SO by analytic_account_id failed: {e}")

        # Fallback: try SO by project_id field
        if not sos:
            try:
                sos = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'sale.order', 'search_read',
                    [[('project_id', 'in', project_ids)]],
                    {'fields': ['id', 'name', 'partner_id', 'date_order', 'state',
                                'amount_untaxed', 'amount_tax', 'amount_total',
                                'invoice_status', 'currency_id', 'order_line'], 'limit': 500}
                )
                logger.info(f"SOs by project_id field: {len(sos)}")
            except Exception as _e: logger.warning(f"SO by project_id: {_e}")

        # Fallback: try project_ids (many2many)
        if not sos:
            try:
                sos = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'sale.order', 'search_read',
                    [[('project_ids', 'in', project_ids)]],
                    {'fields': ['id', 'name', 'partner_id', 'date_order', 'state',
                                'amount_untaxed', 'amount_tax', 'amount_total',
                                'invoice_status', 'currency_id', 'order_line'], 'limit': 500}
                )
                logger.info(f"SOs by project_ids m2m: {len(sos)}")
            except Exception as _e: logger.warning(f"SO by project_ids: {_e}")

        # Fallback: find SOs via invoice_origin on posted invoices for this analytic account
        if not sos and analytic_account_id:
            try:
                # Get posted invoices linked to the analytic account
                linked_invs = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'account.move', 'search_read',
                    [[('move_type', '=', 'out_invoice'),
                      ('state', '=', 'posted'),
                      ('invoice_line_ids.analytic_account_id', '=', analytic_account_id)]],
                    {'fields': ['invoice_origin'], 'limit': 500}
                )
                # Collect SO names from invoice_origin
                so_names = set()
                for inv in linked_invs:
                    for part in (inv.get('invoice_origin') or '').split(','):
                        part = part.strip()
                        if part and part.startswith('S'):
                            so_names.add(part)
                if so_names:
                    sos = odoo.models.execute_kw(
                        ODOO_DB, odoo.uid, ODOO_PASSWORD,
                        'sale.order', 'search_read',
                        [[('name', 'in', list(so_names))]],
                        {'fields': ['id', 'name', 'partner_id', 'date_order', 'state',
                                    'amount_untaxed', 'amount_tax', 'amount_total',
                                    'invoice_status', 'currency_id', 'order_line'], 'limit': 500}
                    )
                    logger.info(f"SOs via invoice_origin fallback: {len(sos)}")
            except Exception as _e: logger.warning(f"SO via invoice_origin: {_e}")

        # Fallback: find SOs via task/timesheet sale_line_id
        if not sos:
            try:
                tasks = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'project.task', 'search_read',
                    [[('project_id', 'in', project_ids), ('sale_order_id', '!=', False)]],
                    {'fields': ['sale_order_id'], 'limit': 500}
                )
                task_so_ids = list({t['sale_order_id'][0] for t in tasks if t.get('sale_order_id')})
                if task_so_ids:
                    sos = odoo.models.execute_kw(
                        ODOO_DB, odoo.uid, ODOO_PASSWORD,
                        'sale.order', 'read',
                        [task_so_ids],
                        {'fields': ['id', 'name', 'partner_id', 'date_order', 'state',
                                    'amount_untaxed', 'amount_tax', 'amount_total',
                                    'invoice_status', 'currency_id', 'order_line']}
                    )
                    logger.info(f"SOs via task.sale_order_id: {len(sos)}")
            except Exception as _e: logger.warning(f"SO via tasks: {_e}")

        # No SOs → fetch direct invoices by analytic account
        if not sos:
            direct_invoices = []
            if analytic_account_id:
                try:
                    direct_invoices = odoo.models.execute_kw(
                        ODOO_DB, odoo.uid, ODOO_PASSWORD,
                        'account.move', 'search_read',
                        [[('move_type', '=', 'out_invoice'),
                          ('state', 'in', ['posted', 'draft']),
                          ('invoice_line_ids.analytic_account_id', '=', analytic_account_id)]],
                        {'fields': ['id', 'name', 'invoice_date', 'invoice_date_due',
                                    'amount_untaxed', 'amount_tax', 'amount_total',
                                    'state', 'invoice_origin', 'payment_state',
                                    'narration', 'purpose', 'ref'],
                         'limit': 200, 'order': 'invoice_date asc'}
                    )
                except Exception as _e: logger.warning(f"Direct invoices: {_e}")

            # Get invoice lines for classification
            inv_ids = [i['id'] for i in direct_invoices]
            dir_inv_lines = []
            if inv_ids:
                try:
                    dir_inv_lines = odoo.models.execute_kw(
                        ODOO_DB, odoo.uid, ODOO_PASSWORD,
                        'account.move.line', 'search_read',
                        [[('move_id', 'in', inv_ids), ('display_type', 'not in', ['line_section','line_note'])]],
                        {'fields': ['move_id', 'name', 'price_subtotal'], 'limit': 2000}
                    )
                except Exception: pass

            # Load saved phase overrides for direct invoices from DB
            pfx2 = active_db_prefix()
            plan_ns2 = f'{pfx2}_plan' if pfx2 else 'plan'
            dir_inv_overrides = db.get_namespace_overrides(plan_ns2, 'direct_inv_phase') or {}

            # Classify each direct invoice by saved override or auto-detection
            def classify_direct_inv(inv_id, inv_name=''):
                key = (inv_name or '').replace('/', '_')
                override = dir_inv_overrides.get(key, {})
                if isinstance(override, dict) and override.get('phase'):
                    return override['phase']
                inv_lines_for = [l for l in dir_inv_lines
                                 if (l['move_id'][0] if isinstance(l['move_id'],list) else l['move_id']) == inv_id]
                for l in inv_lines_for:
                    n = (l.get('name') or '').lower()
                    if any(kw in n for kw in SUPPORT_KWS): return 'support'
                    if 'license' in n or '3rd party' in n or 'third party' in n: return 'license'
                return 'services'
            inv_list = [{
                'name':          i['name'],
                'date':          (i.get('invoice_date') or '')[:10],
                'due_date':      (i.get('invoice_date_due') or '')[:10],
                'amount_untaxed': round(i.get('amount_untaxed',0), 2),
                'amount_tax':    round(i.get('amount_tax',0), 2),
                'amount_total':  round(i.get('amount_total',0), 2),
                'state':         i.get('state',''),
                'payment_state': i.get('payment_state',''),
                'purpose':       i.get('narration') or i.get('purpose') or i.get('ref') or '',
                'phase':         classify_direct_inv(i['id'], i['name']),
            } for i in direct_invoices]

            total_inv = sum(i['amount_untaxed'] for i in inv_list if i['state']=='posted')
            return jsonify({'ok': True, 'orders': [], 'direct_invoices': inv_list,
                           'summary': {'total_orders': 0, 'total_untaxed': 0,
                                       'total_invoiced': round(total_inv,2),
                                       'total_remaining': 0, 'overall_invoiced_pct': 0},
                           'note': f'No sales orders found. Showing {len(inv_list)} direct invoices.'})

        so_ids = [s['id'] for s in sos]

        # Get order lines for qty delivered/invoiced
        all_line_ids = [lid for s in sos for lid in (s.get('order_line') or [])]
        line_map = {}
        if all_line_ids:
            lines = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'sale.order.line', 'read',
                [all_line_ids],
                {'fields': ['id', 'order_id', 'name', 'product_id',
                            'product_uom_qty', 'qty_delivered', 'qty_invoiced',
                            'price_unit', 'price_subtotal', 'discount']}
            )
            for l in lines:
                oid = l['order_id'][0] if isinstance(l['order_id'], list) else l['order_id']
                line_map.setdefault(oid, []).append(l)

        # Get invoices for these SOs
        invoices_raw = []
        try:
            invoices_raw = odoo.models.execute_kw(
                ODOO_DB, odoo.uid, ODOO_PASSWORD,
                'account.move', 'search_read',
                [[('move_type', '=', 'out_invoice'),
                  ('state', 'in', ['posted', 'draft']),
                  ('invoice_line_ids.sale_line_ids.order_id', 'in', so_ids)]],
                {'fields': ['id', 'name', 'invoice_date', 'invoice_date_due',
                            'amount_untaxed', 'amount_tax', 'amount_total',
                            'state', 'invoice_origin', 'payment_state',
                            'invoice_line_ids', 'narration', 'purpose', 'ref'],
                 'limit': 500, 'order': 'invoice_date asc'}
            )
        except Exception as e:
            logger.warning(f"Invoice fetch failed: {e}")

        # Fetch invoice lines with sale_line_ids linkage
        inv_line_map = {}  # invoice_line_id -> {sale_line_id, amount, name}
        all_inv_line_ids = [lid for inv in invoices_raw for lid in (inv.get('invoice_line_ids') or [])]
        if all_inv_line_ids:
            try:
                inv_lines = odoo.models.execute_kw(
                    ODOO_DB, odoo.uid, ODOO_PASSWORD,
                    'account.move.line', 'read',
                    [all_inv_line_ids],
                    {'fields': ['id', 'name', 'sale_line_ids', 'price_subtotal',
                                'quantity', 'move_id', 'display_type']}
                )
                for il in inv_lines:
                    if il.get('display_type') in ('line_section', 'line_note'):
                        continue
                    move_id = il['move_id'][0] if isinstance(il['move_id'], list) else il['move_id']
                    # Get invoice date from invoices_raw
                    inv_date = ''
                    for inv in invoices_raw:
                        if inv['id'] == move_id:
                            inv_date = inv.get('invoice_date') or ''
                            break
                    for slid in (il.get('sale_line_ids') or []):
                        inv_line_map.setdefault(slid, []).append({
                            'inv_line_id':   il['id'],
                            'move_id':       move_id,
                            'move_name':     il['move_id'][1] if isinstance(il['move_id'], list) else '',
                            'name':          il.get('name') or '',
                            'qty':           il.get('quantity') or 0,
                            'amount':        il.get('price_subtotal') or 0,
                            'inv_date':      inv_date,
                            'purpose':       next((inv.get('purpose') or inv.get('narration') or inv.get('ref') or '' for inv in invoices_raw if inv['id']==move_id), ''),
                        })
            except Exception as e:
                logger.warning(f"Invoice lines fetch failed: {e}")

        # Build invoice map by SO name
        inv_by_so = {}
        for inv in invoices_raw:
            origin = inv.get('invoice_origin') or ''
            for s in sos:
                if s['name'] in origin:
                    inv_by_so.setdefault(s['id'], []).append({
                        'name':          inv['name'],
                        'date':          inv.get('invoice_date') or '',
                        'due_date':      inv.get('invoice_date_due') or '',
                        'amount_untaxed': inv.get('amount_untaxed') or 0,
                        'amount_tax':    inv.get('amount_tax') or 0,
                        'amount_total':  inv.get('amount_total') or 0,
                        'state':         inv.get('state') or '',
                        'payment_state': inv.get('payment_state') or '',
                        'purpose':       inv.get('purpose') or inv.get('narration') or inv.get('ref') or '',
                    })

        # Build orders list
        orders = []
        total_untaxed = 0
        total_invoiced = 0
        total_delivered_pct = 0

        for s in sos:
            lines = line_map.get(s['id'], [])
            untaxed   = s.get('amount_untaxed') or 0
            total_ordered_qty  = sum(l.get('product_uom_qty', 0) or 0 for l in lines)
            total_delivered_qty= sum(l.get('qty_delivered',   0) or 0 for l in lines)
            total_invoiced_qty = sum(l.get('qty_invoiced',    0) or 0 for l in lines)
            delivered_pct = round(total_delivered_qty / total_ordered_qty * 100, 1) if total_ordered_qty else 0
            invoiced_pct  = round(total_invoiced_qty  / total_ordered_qty * 100, 1) if total_ordered_qty else 0

            # Invoiced amount from invoice lines
            so_invoices  = inv_by_so.get(s['id'], [])
            invoiced_amt = sum(i['amount_untaxed'] for i in so_invoices if i['state'] == 'posted')
            remaining    = untaxed - invoiced_amt

            partner = s.get('partner_id')
            partner_name = partner[1] if isinstance(partner, list) else ''

            total_untaxed  += untaxed
            total_invoiced += invoiced_amt

            # Add invoice links to each line
            enriched_lines = []
            for l in lines:
                line_invoices = inv_line_map.get(l['id'], [])
                # Calculate delivered/invoiced amounts from line price
                unit = l.get('price_unit') or 0
                disc = 1 - (l.get('discount') or 0) / 100
                delivered_amt = round(l.get('qty_delivered', 0) * unit * disc, 2)
                invoiced_amt_line = round(l.get('qty_invoiced', 0) * unit * disc, 2)
                remaining_amt = round(l.get('price_subtotal', 0) - invoiced_amt_line, 2)
                enriched_lines.append({
                    **l,
                    'delivered_amt':   delivered_amt,
                    'invoiced_amt':    invoiced_amt_line,
                    'remaining_amt':   remaining_amt,
                    'line_invoices':   line_invoices,
                })

            orders.append({
                'id':            s['id'],
                'name':          s['name'],
                'partner':       partner_name,
                'date':          (s.get('date_order') or '')[:10],
                'state':         s.get('state') or '',
                'invoice_status':s.get('invoice_status') or '',
                'amount_untaxed':round(untaxed, 2),
                'amount_tax':    round(s.get('amount_tax') or 0, 2),
                'amount_total':  round(s.get('amount_total') or 0, 2),
                'invoiced_amt':  round(invoiced_amt, 2),
                'remaining':     round(remaining, 2),
                'delivered_pct': delivered_pct,
                'invoiced_pct':  invoiced_pct,
                'lines':         enriched_lines,
                'invoices':      so_invoices,
            })

        orders.sort(key=lambda x: x['date'], reverse=True)

        summary = {
            'total_orders':   len(orders),
            'total_untaxed':  round(total_untaxed, 2),
            'total_invoiced': round(total_invoiced, 2),
            'total_remaining':round(total_untaxed - total_invoiced, 2),
            'overall_invoiced_pct': round(total_invoiced / total_untaxed * 100, 1) if total_untaxed else 0,
        }

        # Build invoices_by_phase: use FULL invoice amount (excl. VAT)
        # because some invoice lines may not be linked to analytic account
        invoices_by_phase = {}  # { phase_key: { 'YYYY-MM': amount_excl_vat } }

        # Step 1: Map each invoice → phase from SO line products + user overrides
        inv_phase_map = {}  # move_id → phase_key
        inv_phase_overrides = {}  # move_id → phase when user explicitly set it
        for order in orders:
            for line in order.get('lines', []):
                prod_name = ''
                if isinstance(line.get('product_id'), list):
                    prod_name = line['product_id'][1] if len(line['product_id']) > 1 else ''
                elif line.get('name'):
                    prod_name = line['name']
                prod_lower = prod_name.lower()
                if any(kw in prod_lower for kw in SUPPORT_KWS):
                    lp = 'support'
                elif 'license' in prod_lower or '3rd party' in prod_lower or 'third party' in prod_lower:
                    lp = 'license'
                else:
                    lp = 'development'
                # User override from dropdown (check both line id and string)
                line_id_str = str(line.get('id', ''))
                user_override = line_var_map.get(line_id_str)
                if user_override:
                    lp = user_override
                for li in line.get('line_invoices', []):
                    mid = li.get('move_id')
                    if mid:
                        if user_override:
                            # User explicitly chose — always wins
                            inv_phase_map[mid] = lp
                        elif mid not in inv_phase_map:
                            inv_phase_map[mid] = lp

        # Step 2: Accumulate FULL invoice amount_untaxed per phase per month
        processed_invs = set()
        for inv in invoices_raw:
            if inv['id'] in processed_invs:
                continue
            processed_invs.add(inv['id'])
            if inv.get('state') != 'posted':
                continue
            month = (inv.get('invoice_date') or '')[:7]
            if not month:
                continue
            lp = inv_phase_map.get(inv['id'], 'development')
            if lp not in invoices_by_phase:
                invoices_by_phase[lp] = {}
            invoices_by_phase[lp][month] = (
                invoices_by_phase[lp].get(month, 0) + (inv.get('amount_untaxed') or 0)
            )

        # Build cumulative per phase
        invoices_by_phase_cumulative = {}
        for phase_key, monthly in invoices_by_phase.items():
            cumulative = {}
            running = 0
            for month in sorted(monthly.keys()):
                running += monthly[month]
                cumulative[month] = {'month': monthly[month], 'cumulative': round(running, 2)}
            invoices_by_phase_cumulative[phase_key] = cumulative

        return jsonify({'ok': True, 'orders': orders, 'summary': summary,
                        'invoices_by_phase': invoices_by_phase_cumulative})

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
